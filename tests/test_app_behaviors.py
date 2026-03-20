import json
import io
import sys
import unittest
from pathlib import Path
from unittest import mock

from streamlit.testing.v1 import AppTest
from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parents[1]
APP_FILE = APP_DIR / "app.py"
sys.path.insert(0, str(APP_DIR))

import app  # noqa: E402


class TestPromptParsing(unittest.TestCase):
    def test_dashboard_request_supports_natural_language_without_keyword(self) -> None:
        request = app.parse_nl_dashboard_request("mostrame gasto por país y calidad de datos para Bancar ARG")
        self.assertTrue(request)
        self.assertTrue(request["show_spend"])
        self.assertTrue(request["show_geo"])
        self.assertTrue(request["show_quality"])
        self.assertEqual(request["filters"].get("company"), "Bancar ARG")
        self.assertNotIn("country", request["filters"])

    def test_compare_stock_prompt_does_not_invent_filters(self) -> None:
        filters = app.parse_filters_from_prompt("comparar stock entre países")
        self.assertEqual(filters, {})

    def test_email_only_prompt_maps_to_assignee_without_identifier(self) -> None:
        filters = app.parse_filters_from_prompt("matias.vazquez2024@gmail.com")
        self.assertEqual(filters.get("assignee"), "matias.vazquez2024@gmail.com")
        self.assertNotIn("identifier", filters)

    def test_bulk_update_accepts_natural_phrase(self) -> None:
        parsed = app.parse_bulk_location_action("actualizá en lote ISI-31645 ISI-32067 a México Bancar MEX")
        self.assertEqual(parsed, (["ISI-31645", "ISI-32067"], "Bancar MEX", "México"))

    def test_attribute_search_does_not_pollute_filters(self) -> None:
        self.assertEqual(app.detect_attribute_search("activos donde Nombre del modelo contiene thinkpad"), ("Nombre del modelo", "contiene", "thinkpad"))
        self.assertEqual(app.parse_filters_from_prompt("activos donde Nombre del modelo contiene thinkpad"), {})


class TestScriptInputs(unittest.TestCase):
    def _attr_map(self, attrs: list[dict]) -> dict[str, str]:
        out: dict[str, str] = {}
        for row in attrs:
            values = row.get("objectAttributeValues") or []
            if values:
                out[str(row.get("objectTypeAttributeId"))] = str(values[0].get("value", ""))
        return out

    def tearDown(self) -> None:
        app.st.session_state.clear()
        app.clear_process_fetch_cache()

    def test_build_asset_payload_accepts_flexible_headers(self) -> None:
        row = {
            "Nombre del activo": "NB-01",
            "Host name": "WKS-001",
            "Model": "ThinkPad X1",
            "Fecha de compra": "2026-01-01",
            "Estado del activo": "En uso",
            "Entidad del activo": "IT",
            "Fecha garantía": "2027-01-01",
            "Purchase Price": "1500",
            "Serial Number": "SER-001",
            "País": "México",
            "Usuario asignado": "ana@bancar.com",
            "Provider": "Lenovo",
            "Category": "laptops",
            "Compañía": "Bancar MEX",
        }
        type_id, attrs = app.build_asset_attributes_payload(row)
        attr_map = self._attr_map(attrs)

        self.assertEqual(type_id, app.CATEGORY_TO_TYPE_ID["portatiles"])
        self.assertEqual(attr_map[app.ID_NAME], "NB-01")
        self.assertEqual(attr_map[app.ID_HOSTNAME], "WKS-001")
        self.assertEqual(attr_map[app.ID_MODELO], "ThinkPad X1")
        self.assertEqual(attr_map[app.ID_PAIS], "México")
        self.assertEqual(attr_map[app.ID_COMPANIA], "Bancar MEX")
        self.assertEqual(attr_map[app.ID_ASIGNACION], "ana@bancar.com")

    def test_build_asset_payload_uses_tipo_when_category_is_missing(self) -> None:
        row = {
            "Nombre": "Dock-01",
            "Tipo": "tablets",
            "Hostname": "TAB-001",
        }
        type_id, attrs = app.build_asset_attributes_payload(row)
        attr_map = self._attr_map(attrs)

        self.assertEqual(type_id, app.CATEGORY_TO_TYPE_ID["tablets"])
        self.assertEqual(attr_map[app.ID_NAME], "Dock-01")
        self.assertEqual(attr_map[app.ID_HOSTNAME], "TAB-001")

    def test_build_asset_payload_derives_company_from_country(self) -> None:
        row = {
            "Nombre del activo": "MON-01",
            "Hostname": "MON-01",
            "Serial Number": "SER-777",
            "Tipo de activo": "Monitores",
            "País": "Colombia",
        }
        _, attrs = app.build_asset_attributes_payload(row)
        attr_map = self._attr_map(attrs)

        self.assertEqual(attr_map[app.ID_PAIS], "Colombia")
        self.assertEqual(attr_map[app.ID_COMPANIA], "Bancar COL")

    def test_answer_inventory_question_returns_assets_for_email_query(self) -> None:
        assets = [
            {
                "jira_key": "ISI-1",
                "name": "NB-01",
                "hostname": "NB-01",
                "serial_number": "SER-001",
                "status": "En uso",
                "country": "Argentina",
                "assigned_to": "matias.vazquez2024@gmail.com",
            },
            {
                "jira_key": "ISI-2",
                "name": "NB-02",
                "hostname": "NB-02",
                "serial_number": "SER-002",
                "status": "En uso",
                "country": "Argentina",
                "assigned_to": "matias.vazquez2024@gmail.com",
            },
            {
                "jira_key": "ISI-3",
                "name": "NB-03",
                "hostname": "NB-03",
                "serial_number": "SER-003",
                "status": "Stock nuevo",
                "country": "Argentina",
                "assigned_to": "otra.persona@bancar.com",
            },
        ]

        response = app.answer_inventory_question(assets, "matias.vazquez2024@gmail.com")

        self.assertIn("activos asignados a **matias.vazquez2024@gmail.com**", response)
        self.assertIn("SER-001", response)
        self.assertIn("SER-002", response)
        self.assertNotIn("SER-003", response)

    def test_mass_update_identifier_accepts_serial_aliases(self) -> None:
        self.assertEqual(app.resolve_mass_update_identifier({"Serial": "SER-001"}), "SER-001")
        self.assertEqual(app.resolve_mass_update_identifier({"Número de serie": "SER-002"}), "SER-002")

    def test_mass_update_identifier_uses_serial_only(self) -> None:
        self.assertEqual(app.resolve_mass_update_identifier({"Serial Number": "SER-001", "Jira Key": "ISI-31645"}), "SER-001")
        self.assertEqual(app.resolve_mass_update_identifier({"Jira Key": "ISI-31645", "Host name": "WKS-001"}), "")

    def test_mass_upload_template_contains_expected_sheets_and_headers(self) -> None:
        raw = app.build_mass_upload_template_bytes()
        workbook = load_workbook(io.BytesIO(raw))

        self.assertEqual(workbook.sheetnames, ["Carga masiva", "Listas", "Instrucciones"])
        ws = workbook["Carga masiva"]
        headers = [ws.cell(row=1, column=idx).value for idx in range(1, len(app.MASS_UPLOAD_TEMPLATE_HEADERS) + 1)]

        self.assertEqual(headers, app.MASS_UPLOAD_TEMPLATE_HEADERS)
        self.assertEqual(ws["A2"].value, app.MASS_UPLOAD_TEMPLATE_EXAMPLE_ROW["Tipo de activo"])
        self.assertEqual(ws["G2"].value, app.MASS_UPLOAD_TEMPLATE_EXAMPLE_ROW["País"])
        self.assertEqual(workbook["Listas"].sheet_state, "hidden")

    def test_mass_update_template_contains_expected_sheets_and_headers(self) -> None:
        raw = app.build_mass_update_template_bytes()
        workbook = load_workbook(io.BytesIO(raw))

        self.assertEqual(workbook.sheetnames, ["Modificación masiva", "Listas", "Instrucciones"])
        ws = workbook["Modificación masiva"]
        headers = [ws.cell(row=1, column=idx).value for idx in range(1, len(app.MASS_UPDATE_TEMPLATE_HEADERS) + 1)]

        self.assertEqual(headers, app.MASS_UPDATE_TEMPLATE_HEADERS)
        self.assertEqual(ws["A2"].value, app.MASS_UPDATE_TEMPLATE_EXAMPLE_ROW["Serial Number"])
        self.assertEqual(ws["E2"].value, app.MASS_UPDATE_TEMPLATE_EXAMPLE_ROW["Estado del activo"])
        self.assertEqual(workbook["Listas"].sheet_state, "hidden")

    def test_build_asset_create_payload_resolves_special_attribute_types(self) -> None:
        config = app.AppConfig(
            jira_email="jira@example.com",
            jira_api_token="token",
            workspace_id="workspace",
            site="https://bancar.atlassian.net",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            rovo_api_key="",
            rovo_enabled=False,
        )
        row = {
            "Nombre del activo": "NB-01",
            "Hostname": "NB-01",
            "Serial Number": "SER-001",
            "Tipo de activo": "laptops",
            "Estado del activo": "Stock nuevo",
            "Usuario asignado": "ana@bancar.com",
            "Compañía": "Bancar ARG",
        }
        attr_defs = [
            {"id": app.ID_NAME, "name": "Name", "defaultType": {"name": "Text"}},
            {"id": app.ID_HOSTNAME, "name": "Hostname", "defaultType": {"name": "Text"}},
            {"id": app.ID_SERIAL, "name": "Serial Number", "defaultType": {"name": "Text"}},
            {"id": app.ID_ESTADO, "name": "Estado del activo", "defaultType": {"name": "Status"}},
            {"id": app.ID_ASIGNACION, "name": "Usuario asignado", "defaultType": {"name": "User"}},
            {"id": app.ID_CATEGORIA, "name": "Categoria", "referenceObjectType": {"id": "1300-ref"}},
            {"id": app.ID_COMPANIA, "name": "Compañía", "referenceObjectType": {"id": "1337-ref"}},
        ]

        def fake_ref_resolver(_config, reference_type_id, raw_value, _auth, *, attr_id="", headers=None):
            if reference_type_id == "1300-ref":
                return "CAT-1"
            if reference_type_id == "1337-ref" and raw_value == "Bancar ARG":
                return "COM-1"
            return None

        def fake_option_lookup(_config, attr_id, _auth, _headers):
            return {"stock nuevo": "STATUS-1"} if attr_id == app.ID_ESTADO else {}

        with (
            mock.patch.object(app, "fetch_object_type_attributes", return_value=attr_defs),
            mock.patch.object(app, "resolve_reference_object_key", side_effect=fake_ref_resolver),
            mock.patch.object(app, "fetch_attribute_option_lookup", side_effect=fake_option_lookup),
            mock.patch.object(app, "resolve_user_account_id", return_value="acc-123"),
        ):
            type_id, attrs, issues = app.build_asset_create_payload(config, row)

        attr_map = self._attr_map(attrs)
        self.assertEqual(type_id, app.CATEGORY_TO_TYPE_ID["portatiles"])
        self.assertEqual(issues, [])
        self.assertEqual(attr_map[app.ID_ESTADO], "STATUS-1")
        self.assertEqual(attr_map[app.ID_ASIGNACION], "acc-123")
        self.assertEqual(attr_map[app.ID_CATEGORIA], "CAT-1")
        self.assertEqual(attr_map[app.ID_COMPANIA], "COM-1")

    def test_build_asset_update_payload_skips_identifier_serial_and_updates_only_filled_columns(self) -> None:
        config = app.AppConfig(
            jira_email="jira@example.com",
            jira_api_token="token",
            workspace_id="workspace",
            site="https://bancar.atlassian.net",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            rovo_api_key="",
            rovo_enabled=False,
        )
        row = {
            "Serial Number": "SER-001",
            "Hostname": "NB-01-NEW",
            "Estado del activo": "Stock usado",
        }
        attr_defs = [
            {"id": app.ID_SERIAL, "name": "Serial Number", "defaultType": {"name": "Text"}},
            {"id": app.ID_HOSTNAME, "name": "Hostname", "defaultType": {"name": "Text"}},
            {"id": app.ID_ESTADO, "name": "Estado del activo", "defaultType": {"name": "Status"}},
        ]

        def fake_option_lookup(_config, attr_id, _auth, _headers):
            return {"stock usado": "STATUS-2"} if attr_id == app.ID_ESTADO else {}

        with (
            mock.patch.object(app, "fetch_object_type_attributes", return_value=attr_defs),
            mock.patch.object(app, "fetch_attribute_option_lookup", side_effect=fake_option_lookup),
        ):
            type_id, attrs, issues = app.build_asset_update_payload(config, "213", row)

        attr_map = self._attr_map(attrs)
        self.assertEqual(type_id, "213")
        self.assertEqual(issues, [])
        self.assertNotIn(app.ID_SERIAL, attr_map)
        self.assertEqual(attr_map[app.ID_HOSTNAME], "NB-01-NEW")
        self.assertEqual(attr_map[app.ID_ESTADO], "STATUS-2")

    def test_cached_fetch_assets_prefers_live_empty_result_over_stale_snapshot(self) -> None:
        config = app.AppConfig(
            jira_email="jira@example.com",
            jira_api_token="token",
            workspace_id="workspace",
            site="https://bancar.atlassian.net",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            rovo_api_key="",
            rovo_enabled=False,
        )
        future = mock.Mock()
        future.result.return_value = ([], {"last_base_records_count": 0}, [])
        app.st.session_state.clear()
        app.PROCESS_FETCH_RESULTS.clear()
        app.PROCESS_FETCH_JOBS.clear()

        with (
            mock.patch.object(app.PROCESS_FETCH_EXECUTOR, "submit", return_value=future),
            mock.patch.object(app, "load_assets_snapshot", return_value=([{"jira_key": "OLD-1"}], {"last_base_records_count": 1}, app.datetime.now())),
            mock.patch.object(app, "save_assets_snapshot", return_value=True) as save_snapshot_mock,
            mock.patch.object(app, "apply_fetch_metadata"),
            mock.patch.object(app, "append_error_events"),
        ):
            assets = app.cached_fetch_assets(config, "", 10, force_live=False)

        self.assertEqual(assets, [])
        save_snapshot_mock.assert_called_once_with([], {"last_base_records_count": 0})
        self.assertEqual(app.st.session_state.get("assets_source"), "live")

    def test_cached_fetch_assets_uses_snapshot_only_when_live_fetch_fails(self) -> None:
        config = app.AppConfig(
            jira_email="jira@example.com",
            jira_api_token="token",
            workspace_id="workspace",
            site="https://bancar.atlassian.net",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            rovo_api_key="",
            rovo_enabled=False,
        )
        future = mock.Mock()
        future.result.side_effect = RuntimeError("boom")
        snapshot_time = app.datetime.now()
        snapshot_assets = [{"jira_key": "OLD-1"}]
        app.st.session_state.clear()
        app.PROCESS_FETCH_RESULTS.clear()
        app.PROCESS_FETCH_JOBS.clear()

        with (
            mock.patch.object(app.PROCESS_FETCH_EXECUTOR, "submit", return_value=future),
            mock.patch.object(app, "load_assets_snapshot", return_value=(snapshot_assets, {"last_base_records_count": 1}, snapshot_time)),
            mock.patch.object(app, "apply_fetch_metadata"),
            mock.patch.object(app, "append_error_events"),
            mock.patch.object(app, "save_assets_snapshot") as save_snapshot_mock,
        ):
            assets = app.cached_fetch_assets(config, "", 10, force_live=False)

        self.assertEqual(assets, snapshot_assets)
        save_snapshot_mock.assert_not_called()
        self.assertEqual(app.st.session_state.get("assets_source"), "snapshot_fallback")

    def test_clear_process_fetch_cache_empties_global_caches(self) -> None:
        app.PROCESS_FETCH_RESULTS["hash-1"] = (0.0, [], {}, [])
        app.PROCESS_FETCH_JOBS["hash-1"] = mock.Mock(cancel=mock.Mock())

        app.clear_process_fetch_cache()

        self.assertEqual(app.PROCESS_FETCH_RESULTS, {})
        self.assertEqual(app.PROCESS_FETCH_JOBS, {})

    def test_build_asset_create_payload_creates_missing_model_and_skips_unknown_assignment(self) -> None:
        config = app.AppConfig(
            jira_email="jira@example.com",
            jira_api_token="token",
            workspace_id="workspace",
            site="https://bancar.atlassian.net",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            rovo_api_key="",
            rovo_enabled=False,
        )
        row = {
            "Nombre del activo": "NB-02",
            "Hostname": "NB-02",
            "Serial Number": "SER-002",
            "Tipo de activo": "laptops",
            "Nombre del modelo": "MacBook Air M4 16GB 512GB",
            "Usuario asignado": "matias.vazquez@gmail.com",
        }
        attr_defs = [
            {"id": app.ID_NAME, "name": "Name", "defaultType": {"name": "Text"}},
            {"id": app.ID_HOSTNAME, "name": "Hostname", "defaultType": {"name": "Text"}},
            {"id": app.ID_SERIAL, "name": "Serial Number", "defaultType": {"name": "Text"}},
            {"id": app.ID_MODELO, "name": "Nombre del modelo", "referenceObjectType": {"id": "994-ref"}},
            {"id": app.ID_ASIGNACION, "name": "Asignacion", "defaultType": {"name": "User"}},
            {"id": app.ID_CATEGORIA, "name": "Categoria", "referenceObjectType": {"id": "1300-ref"}},
        ]

        def fake_ref_resolver(_config, reference_type_id, raw_value, _auth, *, attr_id="", headers=None):
            if reference_type_id == "1300-ref":
                return "CAT-1"
            if reference_type_id == "994-ref":
                return None
            return None

        with (
            mock.patch.object(app, "fetch_object_type_attributes", return_value=attr_defs),
            mock.patch.object(app, "resolve_reference_object_key", side_effect=fake_ref_resolver),
            mock.patch.object(app, "create_reference_object", return_value="MOD-123"),
            mock.patch.object(app, "fetch_attribute_option_lookup", return_value={}),
            mock.patch.object(app, "resolve_user_account_id", return_value=None),
        ):
            type_id, attrs, issues = app.build_asset_create_payload(config, row)

        attr_map = self._attr_map(attrs)
        self.assertEqual(type_id, app.CATEGORY_TO_TYPE_ID["portatiles"])
        self.assertEqual(issues, [])
        self.assertEqual(attr_map[app.ID_MODELO], "MOD-123")
        self.assertNotIn(app.ID_ASIGNACION, attr_map)

    def test_build_asset_create_payload_keeps_assignment_raw_when_schema_is_not_user_type(self) -> None:
        config = app.AppConfig(
            jira_email="jira@example.com",
            jira_api_token="token",
            workspace_id="workspace",
            site="https://bancar.atlassian.net",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            rovo_api_key="",
            rovo_enabled=False,
        )
        row = {
            "Nombre del activo": "NB-03",
            "Hostname": "NB-03",
            "Serial Number": "SER-003",
            "Tipo de activo": "laptops",
            "Usuario asignado": "matias.vazquez@gmail.com",
        }
        attr_defs = [
            {"id": app.ID_NAME, "name": "Name", "defaultType": {"name": "Text"}},
            {"id": app.ID_HOSTNAME, "name": "Hostname", "defaultType": {"name": "Text"}},
            {"id": app.ID_SERIAL, "name": "Serial Number", "defaultType": {"name": "Text"}},
            {"id": app.ID_CATEGORIA, "name": "Categoria", "referenceObjectType": {"id": "1300-ref"}},
            {"id": app.ID_ASIGNACION, "name": "Asignacion", "defaultType": {"name": "Text"}},
        ]

        def fake_ref_resolver(_config, reference_type_id, raw_value, _auth, *, attr_id="", headers=None):
            if reference_type_id == "1300-ref":
                return "CAT-1"
            return None

        with (
            mock.patch.object(app, "fetch_object_type_attributes", return_value=attr_defs),
            mock.patch.object(app, "resolve_reference_object_key", side_effect=fake_ref_resolver),
            mock.patch.object(app, "fetch_attribute_option_lookup", return_value={}),
            mock.patch.object(app, "resolve_user_account_id") as resolve_user_mock,
        ):
            type_id, attrs, issues = app.build_asset_create_payload(config, row)

        attr_map = self._attr_map(attrs)
        self.assertEqual(type_id, app.CATEGORY_TO_TYPE_ID["portatiles"])
        self.assertEqual(issues, [])
        self.assertEqual(attr_map[app.ID_ASIGNACION], "matias.vazquez@gmail.com")
        resolve_user_mock.assert_not_called()

    def test_detects_template_example_row(self) -> None:
        self.assertTrue(app.is_mass_upload_example_row(dict(app.MASS_UPLOAD_TEMPLATE_EXAMPLE_ROW)))
        altered = dict(app.MASS_UPLOAD_TEMPLATE_EXAMPLE_ROW)
        altered["Serial Number"] = "OTRO-SERIAL"
        self.assertFalse(app.is_mass_upload_example_row(altered))

    def test_detects_mass_update_template_example_row(self) -> None:
        self.assertTrue(app.is_mass_update_example_row(dict(app.MASS_UPDATE_TEMPLATE_EXAMPLE_ROW)))
        altered = dict(app.MASS_UPDATE_TEMPLATE_EXAMPLE_ROW)
        altered["Estado del activo"] = "Stock usado"
        self.assertFalse(app.is_mass_update_example_row(altered))

    def test_resolve_reference_object_key_matches_email_inside_label(self) -> None:
        config = app.AppConfig(
            jira_email="jira@example.com",
            jira_api_token="token",
            workspace_id="workspace",
            site="https://bancar.atlassian.net",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            rovo_api_key="",
            rovo_enabled=False,
        )
        lookup = {
            app.normalize_lookup_key("Matias Vazquez (matias.vazquez@gmail.com)"): "USR-1",
        }
        with mock.patch.object(app, "fetch_reference_object_lookup", return_value=lookup):
            resolved = app.resolve_reference_object_key(
                config,
                "1232-ref",
                "matias.vazquez@gmail.com",
                _auth := mock.Mock(),
                attr_id=app.ID_ASIGNACION,
                headers={},
            )
        self.assertEqual(resolved, "USR-1")

    def test_fetch_reference_object_lookup_indexes_raw_email_value(self) -> None:
        config = app.AppConfig(
            jira_email="jira@example.com",
            jira_api_token="token",
            workspace_id="workspace",
            site="https://bancar.atlassian.net",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            rovo_api_key="",
            rovo_enabled=False,
        )
        body = {
            "values": [
                {
                    "objectKey": "USR-1",
                    "label": "Matias Vazquez",
                    "attributes": [
                        {
                            "objectAttributeValues": [
                                {
                                    "displayValue": "Matias Vazquez",
                                    "value": "matias.vazquez@gmail.com",
                                }
                            ]
                        }
                    ],
                }
            ]
        }
        response = mock.Mock()
        response.json.return_value = body
        app.st.session_state.clear()
        with mock.patch.object(app, "jira_request_with_retry", return_value=response):
            lookup = app.fetch_reference_object_lookup(config, "1232-ref", mock.Mock(), {})

        self.assertEqual(lookup[app.normalize_lookup_key("matias.vazquez@gmail.com")], "USR-1")

    def test_fetch_reference_object_lookup_falls_back_to_later_endpoint_when_first_is_empty(self) -> None:
        config = app.AppConfig(
            jira_email="jira@example.com",
            jira_api_token="token",
            workspace_id="workspace",
            site="https://bancar.atlassian.net",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            rovo_api_key="",
            rovo_enabled=False,
        )
        first_response = mock.Mock()
        first_response.json.return_value = {"values": []}
        second_response = mock.Mock()
        second_response.json.return_value = {
            "values": [
                {
                    "objectKey": "USR-1",
                    "label": "Matias Vazquez",
                    "attributes": [
                        {
                            "objectAttributeValues": [
                                {
                                    "displayValue": "Matias Vazquez",
                                    "value": "matias.vazquez@uala.com.ar",
                                }
                            ]
                        }
                    ],
                }
            ]
        }
        app.st.session_state.clear()
        with mock.patch.object(
            app,
            "jira_request_with_retry",
            side_effect=[first_response, second_response],
        ):
            lookup = app.fetch_reference_object_lookup(config, "1232-ref", mock.Mock(), {})

        self.assertEqual(lookup[app.normalize_lookup_key("matias.vazquez@uala.com.ar")], "USR-1")

    def test_fetch_reference_object_lookup_uses_batch_script_payload_shape(self) -> None:
        config = app.AppConfig(
            jira_email="jira@example.com",
            jira_api_token="token",
            workspace_id="workspace",
            site="https://bancar.atlassian.net",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            rovo_api_key="",
            rovo_enabled=False,
        )
        response = mock.Mock()
        response.json.return_value = {"values": []}
        app.st.session_state.clear()
        with mock.patch.object(app, "jira_request_with_retry", return_value=response) as request_mock:
            app.fetch_reference_object_lookup(config, "1232-ref", mock.Mock(), {})

        first_call = request_mock.call_args_list[0]
        self.assertNotIn("params", first_call.kwargs)
        self.assertEqual(first_call.kwargs["json_payload"]["resultsPerPage"], 1000)
        self.assertTrue(first_call.kwargs["json_payload"]["includeAttributes"])
        self.assertEqual(first_call.kwargs["json_payload"]["qlQuery"], "objectTypeId = 1232-ref")

    def test_warranty_date_is_sent_as_jira_datetime(self) -> None:
        config = app.AppConfig(
            jira_email="jira@example.com",
            jira_api_token="token",
            workspace_id="workspace",
            site="https://bancar.atlassian.net",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            rovo_api_key="",
            rovo_enabled=False,
        )
        row = {
            "Nombre del activo": "NB-04",
            "Hostname": "NB-04",
            "Serial Number": "SER-004",
            "Tipo de activo": "laptops",
            "Fecha garantía": "2027-03-20",
        }
        attr_defs = [
            {"id": app.ID_NAME, "name": "Name", "defaultType": {"name": "Text"}},
            {"id": app.ID_HOSTNAME, "name": "Hostname", "defaultType": {"name": "Text"}},
            {"id": app.ID_SERIAL, "name": "Serial Number", "defaultType": {"name": "Text"}},
            {"id": app.ID_CATEGORIA, "name": "Categoria", "referenceObjectType": {"id": "1300-ref"}},
            {"id": app.ID_FECHA_GARANTIA, "name": "Fecha soporte garantia", "defaultType": {"name": "DateTime"}},
        ]

        def fake_ref_resolver(_config, reference_type_id, raw_value, _auth, *, attr_id="", headers=None):
            if reference_type_id == "1300-ref":
                return "CAT-1"
            return None

        with (
            mock.patch.object(app, "fetch_object_type_attributes", return_value=attr_defs),
            mock.patch.object(app, "resolve_reference_object_key", side_effect=fake_ref_resolver),
            mock.patch.object(app, "fetch_attribute_option_lookup", return_value={}),
        ):
            _, attrs, issues = app.build_asset_create_payload(config, row)

        attr_map = self._attr_map(attrs)
        self.assertEqual(issues, [])
        self.assertEqual(attr_map[app.ID_FECHA_GARANTIA], "2027-03-20T00:00:00.000Z")

    def test_create_asset_from_payload_returns_real_error_body(self) -> None:
        config = app.AppConfig(
            jira_email="jira@example.com",
            jira_api_token="token",
            workspace_id="workspace",
            site="https://bancar.atlassian.net",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            rovo_api_key="",
            rovo_enabled=False,
        )
        fake_response = mock.Mock(status_code=404, text='{"error":"not found"}')
        fake_response.json.return_value = {}
        fake_client = mock.MagicMock()
        fake_client.__enter__.return_value = fake_client
        fake_client.post.return_value = fake_response

        with mock.patch.object(app.httpx, "Client", return_value=fake_client):
            ok, msg = app.create_asset_from_payload(
                config,
                "213",
                [{"objectTypeAttributeId": app.ID_NAME, "objectAttributeValues": [{"value": "NB-01"}]}],
            )

        self.assertFalse(ok)
        self.assertIn("404", msg)
        self.assertIn("not found", msg)


class TestChatDashboardIntegration(unittest.TestCase):
    def test_chat_renders_dashboard_section_below_history(self) -> None:
        snapshot_path = APP_DIR / "assets_snapshot.json"
        previous_snapshot = snapshot_path.read_text(encoding="utf-8") if snapshot_path.exists() else None
        sample_snapshot = {
            "saved_at": "2026-03-20T10:00:00",
            "assets": [
                {
                    "jira_key": "ISI-1",
                    "name": "NB-01",
                    "hostname": "NB-01",
                    "serial_number": "SER-001",
                    "category": "Portátiles",
                    "status": "En uso",
                    "country": "Argentina",
                    "company": "Bancar ARG",
                    "purchase_price": "1500",
                    "warranty_date": "2027-03-20",
                    "assigned_to": "ana@bancar.com",
                    "attrs_by_name": {},
                    "attrs_by_id": {},
                },
                {
                    "jira_key": "ISI-2",
                    "name": "NB-02",
                    "hostname": "NB-02",
                    "serial_number": "SER-002",
                    "category": "Portátiles",
                    "status": "Stock nuevo",
                    "country": "Argentina",
                    "company": "Bancar ARG",
                    "purchase_price": "1800",
                    "warranty_date": "2027-05-01",
                    "assigned_to": "",
                    "attrs_by_name": {},
                    "attrs_by_id": {},
                },
            ],
            "metadata": {"last_base_records_count": 2},
        }
        snapshot_path.write_text(json.dumps(sample_snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
        try:
            at = AppTest.from_file(str(APP_FILE), default_timeout=180)
            at.session_state["assets"] = sample_snapshot["assets"]
            at.session_state["last_sync"] = None
            at.session_state["assets_source"] = "snapshot_fallback"
            at.query_params["page"] = "Chat"
            at.run()
            at.chat_input[0].set_value("mostrame gasto por país y calidad de datos para Bancar ARG").run()

            markdown_values = [str(getattr(item, "value", "")) for item in at.markdown]
            self.assertEqual(len(at.exception), 0)
            self.assertGreaterEqual(len(at.chat_message), 2)
            self.assertTrue(any("### Dashboard solicitado" in value for value in markdown_values))
            self.assertGreaterEqual(sum("**Dashboard —" in value for value in markdown_values), 2)
        finally:
            if previous_snapshot is None:
                snapshot_path.unlink(missing_ok=True)
            else:
                snapshot_path.write_text(previous_snapshot, encoding="utf-8")


if __name__ == "__main__":
    unittest.main()
