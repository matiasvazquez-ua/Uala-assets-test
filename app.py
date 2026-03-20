# ── 1. IMPORTS & CONFIGURACIÓN ──────────────────────────────────────────
import asyncio
import copy
import concurrent.futures
import hashlib
import html
import io
import json
import os
import random
import re
import threading
import time
import unicodedata
from calendar import monthrange
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import httpx
import streamlit as st
from httpx import BasicAuth
from pydantic import BaseModel, Field, field_validator

try:
    from dotenv import load_dotenv
except ModuleNotFoundError:
    load_dotenv = None

try:
    from apscheduler.schedulers.background import BackgroundScheduler
except ModuleNotFoundError:
    BackgroundScheduler = None

try:
    import pandas as pd
except ModuleNotFoundError:
    pd = None

try:
    import plotly.express as px
except ModuleNotFoundError:
    px = None

try:
    import plotly.io as pio
except ModuleNotFoundError:
    pio = None

try:
    from openai import OpenAI
except ModuleNotFoundError:
    OpenAI = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ModuleNotFoundError:
    Workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None
    DataValidation = None


def load_env_tolerant(env_path: str = ".env") -> None:
    path = Path(env_path)
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        raw = line.strip()
        if not raw or raw.startswith("#") or "=" not in raw:
            continue
        key, value = raw.split("=", 1)
        key = key.strip()
        if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", key):
            continue
        value = value.strip().strip('"').strip("'")
        os.environ.setdefault(key, value)


# Este proyecto suele mezclar notas/JSON dentro de `.env`.
# `python-dotenv` emite warnings ruidosos en ese caso, así que usamos
# únicamente el loader tolerante que toma solo líneas `KEY=VALUE`.
load_env_tolerant()


def debug_log(message: str) -> None:
    """Log opcional para diagnosticar el flujo de carga sin romper la app."""
    if os.getenv("APP_DEBUG", "").strip().lower() not in {"1", "true", "yes"}:
        return
    print(f"[APP_DEBUG] {datetime.now().isoformat(timespec='seconds')} {message}", flush=True)

# ── 2. CONSTANTES & MAPEOS ───────────────────────────────────────────────
SITE = os.getenv("JIRA_SITE", "https://bancar.atlassian.net").rstrip("/")
WORKSPACE_ID = os.getenv("ASSETS_WORKSPACE_ID") or os.getenv("JIRA_WORKSPACE_ID", "")
SCHEMA_ID = os.getenv("ASSETS_SCHEMA_ID", "40")
REQUEST_TIMEOUT = (10, 60)
PAGE_SIZE = int(os.getenv("ASSETS_PAGE_SIZE", "1000"))
FORCE_FETCH_MIN_ASSETS = int(os.getenv("ASSETS_FORCE_FETCH_MIN", "200"))
FORCE_FETCH_ENABLED = os.getenv("ASSETS_FORCE_FETCH_ENABLED", "false").strip().lower() in {"1", "true", "yes", "si", "sí"}
ASSETS_SCOPE_MODE = os.getenv("ASSETS_SCOPE_MODE", "hardware").strip().lower()
TYPE_SCAN_ENABLED = os.getenv("ASSETS_TYPE_SCAN_ENABLED", "false").strip().lower() in {"1", "true", "yes", "si", "sí"}
TYPE_SCAN_START = int(os.getenv("ASSETS_TYPE_SCAN_START", "200"))
TYPE_SCAN_END = int(os.getenv("ASSETS_TYPE_SCAN_END", "500"))
SEGMENTED_FETCH_ENABLED = os.getenv("ASSETS_SEGMENTED_FETCH_ENABLED", "false").strip().lower() in {"1", "true", "yes", "si", "sí"}
KNOWN_OBJECT_TYPE_IDS = ["213", "217", "225", "229", "231", "235", "238"]
GENERAL_HARDWARE_TYPE_ID = os.getenv("JIRA_GENERAL_HARDWARE_TYPE_ID", "211")
BASE_DIR = Path(__file__).resolve().parent
RULES_FILE = BASE_DIR / "reglas_normalizacion.json"
AUTO_ASSIGN_RULES_FILE = BASE_DIR / "auto_assign_rules.json"
AUTO_ASSIGN_LOG_FILE = BASE_DIR / "auto_assign_log.json"
AUTO_ASSIGN_SNAPSHOT_FILE = BASE_DIR / "auto_assign_snapshot_prev.json"
MOVEMENTS_FILE = BASE_DIR / "movimientos_uala.jsonl"
ASSETS_SNAPSHOT_FILE = BASE_DIR / "assets_snapshot.json"
PROCESS_FETCH_CACHE_TTL_SECONDS = 600
PROCESS_FETCH_EXECUTOR = concurrent.futures.ThreadPoolExecutor(max_workers=2)
PROCESS_FETCH_LOCK = threading.Lock()
PROCESS_FETCH_JOBS: dict[str, concurrent.futures.Future] = {}
PROCESS_FETCH_RESULTS: dict[str, tuple[float, list[dict[str, Any]], dict[str, int], list[dict[str, Any]]]] = {}

# IDs de asset_masivo.py
ID_NAME = "991"
ID_HOSTNAME = "993"
ID_MODELO = "994"
ID_FECHA_COMPRA = "999"
ID_ESTADO = "1005"
ID_ENTIDAD = "1088"
ID_FECHA_GARANTIA = "1089"
ID_COSTO = "1090"
ID_SERIAL = "1091"
ID_PAIS = "1092"
ID_ASIGNACION = "1232"
ID_PROVEEDOR = "1265"
ID_CATEGORIA = "1300"
ID_COMPANIA = "1337"

CATEGORY_ALIAS_TO_CANONICAL = {
    "laptop": "portatiles",
    "laptops": "portatiles",
    "notebook": "portatiles",
    "notebooks": "portatiles",
    "portable": "portatiles",
    "portatil": "portatiles",
    "portatiles": "portatiles",
    "portátiles": "portatiles",
    "impresora": "impresoras",
    "impresoras": "impresoras",
    "printer": "impresoras",
    "printers": "impresoras",
    "activo": "consumibles",
    "activos": "consumibles",
    "consumible": "consumibles",
    "consumibles": "consumibles",
    "mobile phone": "celulares",
    "mobile phones": "celulares",
    "phone": "celulares",
    "phones": "celulares",
    "telefono": "celulares",
    "telefonos": "celulares",
    "teléfono": "celulares",
    "teléfonos": "celulares",
    "celular": "celulares",
    "celulares": "celulares",
    "monitor": "monitores",
    "monitores": "monitores",
    "desktop": "monitores",
    "acces point": "accesorios de conectividad",
    "access point": "accesorios de conectividad",
    "access points": "accesorios de conectividad",
    "acces points": "accesorios de conectividad",
    "ap": "accesorios de conectividad",
    "docking": "accesorios de conectividad",
    "dock": "accesorios de conectividad",
    "router": "accesorios de conectividad",
    "switch": "accesorios de conectividad",
    "firewall": "accesorios de conectividad",
    "servidor": "accesorios de conectividad",
    "servidores": "accesorios de conectividad",
    "server": "accesorios de conectividad",
    "servers": "accesorios de conectividad",
    "network": "accesorios de conectividad",
    "networking": "accesorios de conectividad",
    "conectividad": "accesorios de conectividad",
    "accesorios de conectividad": "accesorios de conectividad",
    "tablet": "tablets",
    "tablets": "tablets",
}

ESTADO_NORMALIZATION = {
    "en uso": "En uso",
    "asignado": "En uso",
    "asignado al edificio": "Asignado al edificio",
    "asignado (deployed)": "En uso",
    "asignado (deployable)": "En uso",
    "stock nuevo": "Stock nuevo",
    "nuevo": "Stock nuevo",
    "stock usado": "Stock usado",
    "usado": "Stock usado",
    "stock nuevo (deployable)": "Stock nuevo",
    "stock usado (deployable)": "Stock usado",
    "stock nuevo (deployed)": "Stock nuevo",
    "stock usado (deployed)": "Stock usado",
    "asignado para testing (deployed)": "En uso",
    "asignado para testing (deployable)": "En uso",
}

PAIS_KEYWORDS = {
    "Argentina": {"arg", "ar", "argentina", "buenos aires", "palermo", "ceibo", "nicaragua", "uala ar"},
    "Colombia": {"col", "co", "colombia", "bogota", "bogotá", "wework colombia", "uala co"},
    "México": {"mx", "mex", "mexico", "méxico", "cdmx", "ciudad de mexico", "ciudad de méxico", "uala mx"},
}

COMPANIA_CANONICAS = {
    "bancar arg": "Bancar ARG",
    "bancar col": "Bancar COL",
    "bancar mex": "Bancar MEX",
}

COMPANIA_KEYWORDS = {
    "Bancar ARG": {"arg", "ar", "argentina", "bancar arg", "bancar argentina"},
    "Bancar COL": {"col", "co", "colombia", "bancar col", "bancar colombia"},
    "Bancar MEX": {"mex", "mx", "mexico", "méxico", "bancar mex", "bancar mexico", "bancar méxico"},
}

CATEGORY_TO_TYPE_ID = {
    "portatiles": "213",
    "impresoras": "217",
    "consumibles": "225",
    "celulares": "229",
    "monitores": "231",
    "accesorios de conectividad": "235",
    "tablets": "238",
}

IDENTIFIER_REGEX = re.compile(r"\b[A-Za-z]{3,}[0-9]*[A-Za-z0-9._-]*\b")
IDENTIFIER_STOPWORDS = {
    "todos", "todas", "activo", "activos", "equipo", "equipos", "asset", "assets", "stock", "estado", "estado",
    "modelo", "model", "pais", "país", "compania", "compañia", "company", "usuario", "asignado", "asignada",
    "argentina", "colombia", "mexico", "méxico", "bancar", "dashboard", "grafico", "gráfico", "panel",
    "quiero", "mostrar", "mostrame", "muéstrame", "necesito", "desde", "hasta", "para", "sobre",
    "resumen", "ejecutivo", "comparar", "top", "usuarios", "garantia", "garantía", "costo", "costos",
    "entre", "datos", "calidad", "critico", "crítico", "paises", "países", "gasto", "inversion", "inversión",
}
GENERIC_CATEGORY_PROMPT_ALIASES = {"activo", "activos"}
DASHBOARD_EXPLICIT_HINTS = {"dashboard", "insight", "grafico", "grafica", "gráfico", "gráfica", "panel"}
DASHBOARD_VISUAL_HINTS = {"mostrar", "mostrame", "muéstrame", "ver", "quiero", "armar", "arma", "generar", "grafico", "grafica", "panel"}
DASHBOARD_AGGREGATION_HINTS = {"comparar", "distribucion", "distribución", "ranking", "mix", "apertura", "brecha", "score"}
MASS_UPLOAD_COLUMN_ALIASES = {
    "name": ["Nombre", "Name", "Nombre del activo", "Asset Name"],
    "hostname": ["Hostname", "Host name", "Nombre de host", "Computer Name"],
    "model": ["Modelo", "Model", "Nombre del modelo"],
    "purchase_date": ["Fecha compra", "Fecha de compra", "Purchase Date"],
    "status": ["Estado", "Estado del activo", "Status"],
    "entity": ["Entidad", "Entidad del activo", "Entity"],
    "warranty_date": ["Fecha garantia", "Fecha garantía", "Garantia", "Garantía", "Warranty", "Warranty End"],
    "cost": ["Costo", "Cost", "Precio", "Purchase Price"],
    "serial": ["Serial", "Serial Number", "Número de serie", "Numero de serie"],
    "country": ["Pais", "País", "Country"],
    "assignment": ["Asignacion", "Asignación", "Assigned To", "Usuario asignado", "Asignado a"],
    "provider": ["Proveedor", "Provider"],
    "category": ["Categoria", "Categoría", "Category", "Tipo", "Tipo de activo", "Object Type"],
    "company": ["Compania", "Compañía", "Company"],
}
MASS_UPLOAD_TEMPLATE_HEADERS = [
    "Tipo de activo",
    "Nombre del activo",
    "Hostname",
    "Serial Number",
    "Nombre del modelo",
    "Estado del activo",
    "País",
    "Compañía",
    "Entidad del activo",
    "Usuario asignado",
    "Fecha de compra",
    "Fecha garantía",
    "Purchase Price",
    "Provider",
]
MASS_UPLOAD_REQUIRED_HEADERS = [
    "Tipo de activo",
    "Nombre del activo",
    "Hostname",
    "Serial Number",
    "Nombre del modelo",
    "Estado del activo",
    "País",
]
MASS_UPLOAD_REQUIRED_HEADER_SET = set(MASS_UPLOAD_REQUIRED_HEADERS)
MASS_UPLOAD_TEMPLATE_EXAMPLE_ROW = {
    "Tipo de activo": "Portátiles",
    "Nombre del activo": "WKSAR0001L",
    "Hostname": "WKSAR0001L",
    "Serial Number": "ABC123XYZ",
    "Nombre del modelo": "MacBook Air M4 16GB 512GB",
    "Estado del activo": "Stock nuevo",
    "País": "Argentina",
    "Compañía": "Bancar ARG",
    "Entidad del activo": "ALAU Tecnología S.A.U.",
    "Usuario asignado": "",
    "Fecha de compra": "2026-03-20",
    "Fecha garantía": "2027-03-20",
    "Purchase Price": "1500",
    "Provider": "Macstation",
}
MASS_UPLOAD_TEMPLATE_LISTS = {
    "Tipo de activo": ["Portátiles", "Impresoras", "Consumibles", "Celulares", "Monitores", "Accesorios de conectividad", "Tablets"],
    "Estado del activo": ["En uso", "Stock nuevo", "Stock usado", "Asignado al edificio"],
    "País": ["Argentina", "Colombia", "México"],
    "Compañía": ["Bancar ARG", "Bancar COL", "Bancar MEX"],
}
MASS_UPDATE_IDENTIFIER_ALIASES = [
    "Serial",
    "Serial Number",
    "Hostname",
    "Host name",
    "Jira",
    "Jira Key",
    "Object Key",
    "Identificador",
    "Identifier",
    "Asset Key",
]
CHAT_PAYLOAD_PREFIX = "__CHAT_PAYLOAD__::"

THEMES = {
    "Oscuro ": {
        "bg": "#003262",
        "accent": "#D4A12A",
        "text": "#F8F9FA",
        "card": "#0A3A73",
        "muted": "#D8DEE4",
    },
    "Claro": {
        "bg": "#FFFFFF",
        "accent": "#003262",
        "text": "#1A1A1A",
        "card": "#F6F8FC",
        "muted": "#5B6270",
    },
    "Cyberpunk/Futurista": {
        "bg": "#0b0f19",
        "accent": "#00f3ff",
        "accent_alt": "#ff003c",
        "text": "#e6fbff",
        "card": "rgba(12, 18, 33, 0.72)",
        "muted": "#86a9c7",
    },
}

SCHEMA_MINI = {
    "objectSchemaId": SCHEMA_ID,
    "knownObjectTypeIds": KNOWN_OBJECT_TYPE_IDS,
    "attributes": [
        {"id": ID_HOSTNAME, "name": "Hostname"},
        {"id": ID_SERIAL, "name": "Serial Number"},
        {"id": ID_ESTADO, "name": "Estado del activo"},
        {"id": ID_ASIGNACION, "name": "Usuario asignado"},
        {"id": ID_MODELO, "name": "Nombre del modelo"},
        {"id": ID_PAIS, "name": "Pais"},
        {"id": ID_COMPANIA, "name": "Compañía"},
        {"id": ID_ENTIDAD, "name": "Entidad del activo"},
    ],
}


@dataclass
class AppConfig:
    jira_email: str
    jira_api_token: str
    workspace_id: str
    site: str
    openai_api_key: str
    openai_model: str
    rovo_api_key: str
    rovo_enabled: bool


@dataclass
class StructuredError:
    """Representa un error HTTP estructurado para auditoría y soporte."""
    url: str
    method: str
    status_code: int
    response_body: str
    timestamp: str
    context: str


@dataclass
class MovimientoAsset:
    """Representa un movimiento de inventario persistido para auditoría."""
    timestamp: str
    tipo_accion: str
    identificador: str
    jira_key: str
    campo_modificado: str
    valor_anterior: str
    valor_nuevo: str
    usuario_asignado: str
    ejecutado_por: str
    resultado: str
    detalle: str


@dataclass
class ReglaNormalizacion:
    """Define una regla de normalización masiva sobre activos."""
    campo_condicion: str
    operador: str
    valor_condicion: str
    campo_a_modificar: str
    valor_nuevo: str
    descripcion: str


@dataclass
class ReglaAsignacionAuto:
    nombre: str
    activa: bool
    campo_condicion: str
    operador: str
    valor_condicion: str
    tipo_accion: str
    usuario_destino: str
    perfil_destino: str
    pais_destino: str
    cola_usuarios: list[str]
    estado_destino: str
    prioridad: int
    descripcion: str


# ── 4. UTILIDADES GENERALES ──────────────────────────────────────────────
def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def normalize_lookup_key(value: Any) -> str:
    text = normalize_text(value)
    text = "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return " ".join(text.split())


def canonical_model_key(value: Any) -> str:
    text = normalize_text(value)
    text = text.replace('"', "")
    text = re.sub(r"\s+", " ", text)
    return text.rstrip(" /")


def compact_lookup_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_lookup_key(value))


def lookup_tokens(value: Any) -> set[str]:
    return set(re.findall(r"[a-z0-9]+", normalize_lookup_key(value)))


def matches_lookup_keyword(lookup_text: str, keyword: str, *, lookup_token_set: set[str] | None = None) -> bool:
    normalized_keyword = normalize_lookup_key(keyword)
    if not normalized_keyword:
        return False
    tokens = lookup_token_set if lookup_token_set is not None else set(re.findall(r"[a-z0-9]+", lookup_text))
    if len(normalized_keyword.split()) == 1 and len(normalized_keyword) <= 3:
        return normalized_keyword in tokens
    return bool(re.search(rf"(?<![a-z0-9]){re.escape(normalized_keyword)}(?![a-z0-9])", lookup_text))


def normalize_tabular_value(value: Any) -> str:
    if pd is not None:
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value or "").strip()
    if text.lower() in {"", "nan", "nat", "none"}:
        return ""
    if re.fullmatch(r"-?\d+\.0", text):
        return text[:-2]
    return text


def build_row_lookup(row: dict[str, Any]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for key, value in row.items():
        normalized_key = normalize_lookup_key(key)
        normalized_value = normalize_tabular_value(value)
        if normalized_key and normalized_value and normalized_key not in lookup:
            lookup[normalized_key] = normalized_value
    return lookup


def get_row_value_by_aliases(row_lookup: dict[str, str], aliases: list[str]) -> str:
    alias_keys = [normalize_lookup_key(alias) for alias in aliases if normalize_lookup_key(alias)]
    for alias_key in alias_keys:
        if alias_key in row_lookup:
            return row_lookup[alias_key]
    for alias_key in alias_keys:
        for key, value in row_lookup.items():
            if alias_key and (alias_key in key or key in alias_key):
                return value
    return ""


def company_for_country(country: str) -> str:
    canonical = canonical_country(country)
    return {
        "Argentina": "Bancar ARG",
        "Colombia": "Bancar COL",
        "México": "Bancar MEX",
    }.get(canonical, "")


def parse_date(value: str) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    formats = [
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.000Z",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d/%b/%Y",
        "%d/%b/%y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def format_jira_datetime(value: str) -> str:
    parsed = parse_date(value)
    if parsed is None:
        return str(value or "").strip()
    return parsed.strftime("%Y-%m-%dT%H:%M:%S.000Z")


def canonical_category(raw_value: str) -> str:
    normalized = normalize_text(raw_value)
    if not normalized:
        return "Sin categoría"
    if normalized in CATEGORY_ALIAS_TO_CANONICAL:
        return CATEGORY_ALIAS_TO_CANONICAL[normalized]
    for alias, canonical in CATEGORY_ALIAS_TO_CANONICAL.items():
        if alias in normalized:
            return canonical
    return raw_value.strip() or "Sin categoría"


def canonical_status(raw_value: str) -> str:
    normalized = normalize_text(raw_value)
    if not normalized:
        return "Sin estado"
    if normalized in ESTADO_NORMALIZATION:
        return ESTADO_NORMALIZATION[normalized]
    return raw_value.strip() or "Sin estado"


def canonical_country(raw_value: str) -> str:
    normalized = normalize_text(raw_value)
    if not normalized:
        return "Sin país"
    for country, keywords in PAIS_KEYWORDS.items():
        if normalized in keywords or any(keyword in normalized for keyword in keywords):
            return country
    return raw_value.strip() or "Sin país"


def get_env_or_secret(key: str, default: str = "") -> str:
    value = os.getenv(key, "")
    if value:
        return value
    try:
        return str(st.secrets.get(key, default))
    except Exception:
        return default


# ── 3. MODELOS DE DATOS (Pydantic) ───────────────────────────────────────
class AssetRecord(BaseModel):
    """Modelo tipado de un activo de Jira Assets."""
    object_id: str = ""
    object_type_id: str = ""
    name: str = ""
    object_type: str = ""
    jira_key: str = ""
    category: str = "Sin categoría"
    status: str = "Sin estado"
    country: str = "Sin país"
    company: str = ""
    entity: str = ""
    assigned_to: str = ""
    serial_number: str = ""
    hostname: str = ""
    model: str = ""
    provider: str = ""
    purchase_date: str = ""
    warranty_date: str = ""
    purchase_price: str = ""
    created: str = ""
    attrs_by_name: dict[str, str] = Field(default_factory=dict)
    attrs_by_id: dict[str, str] = Field(default_factory=dict)
    attr_name_to_id: dict[str, str] = Field(default_factory=dict)

    @field_validator("category", mode="before")
    @classmethod
    def _normalize_category(cls, v: Any) -> str:
        return canonical_category(str(v or ""))

    @field_validator("status", mode="before")
    @classmethod
    def _normalize_status(cls, v: Any) -> str:
        return canonical_status(str(v or ""))

    @field_validator("country", mode="before")
    @classmethod
    def _normalize_country(cls, v: Any) -> str:
        return canonical_country(str(v or ""))

    def to_dict(self) -> dict[str, Any]:
        """Devuelve un dict para compatibilidad con código legado."""
        return self.model_dump()


def enforce_assignment_status_rules(status_value: str, assigned_to: str) -> tuple[str, str]:
    status = (status_value or "").strip() or "Sin estado"
    assigned = (assigned_to or "").strip()
    status_norm = normalize_text(status)

    # Regla dura: "Asignado al edificio" siempre sin usuario asignado.
    if status_norm == normalize_text("Asignado al edificio"):
        return "Asignado al edificio", ""

    # Regla dura: con usuario asignado => En uso.
    if assigned:
        return "En uso", assigned

    # Regla dura: desasignado y en uso => Stock usado.
    if status_norm == normalize_text("En uso"):
        return "Stock usado", ""

    return status, assigned


def load_config() -> AppConfig:
    rovo_enabled = normalize_text(get_env_or_secret("ROVO_ENABLED", "false")) in {"1", "true", "yes", "si", "sí"}
    jira_site = get_env_or_secret("JIRA_SITE", SITE).rstrip("/")
    workspace = get_env_or_secret("ASSETS_WORKSPACE_ID", "") or get_env_or_secret("JIRA_WORKSPACE_ID", WORKSPACE_ID)
    return AppConfig(
        jira_email=get_env_or_secret("JIRA_EMAIL", ""),
        jira_api_token=get_env_or_secret("JIRA_API_TOKEN", ""),
        workspace_id=workspace,
        site=jira_site,
        openai_api_key=get_env_or_secret("OPENAI_API_KEY", ""),
        openai_model=get_env_or_secret("OPENAI_MODEL", "gpt-4o-mini"),
        rovo_api_key=get_env_or_secret("ROVO_API_KEY", ""),
        rovo_enabled=rovo_enabled,
    )


# ── 5. CLIENTE JIRA ASSETS (fetch, AQL, retry) ───────────────────────────
def build_auth_headers(config: AppConfig) -> tuple[BasicAuth, dict[str, str]]:
    """Construye autenticación y headers estándar para Jira Assets."""
    auth = BasicAuth(config.jira_email, config.jira_api_token)
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "X-Atlassian-Token": "no-check",
    }
    return auth, headers


def config_to_cache_key(config: AppConfig) -> dict[str, str]:
    """Genera clave serializable del config para operaciones auxiliares."""
    return {
        "jira_email": config.jira_email,
        "jira_api_token": config.jira_api_token,
        "workspace_id": config.workspace_id,
        "site": config.site,
        "openai_api_key": config.openai_api_key,
        "openai_model": config.openai_model,
        "rovo_api_key": config.rovo_api_key,
        "rovo_enabled": "true" if config.rovo_enabled else "false",
    }


def config_from_cache_key(data: dict[str, str]) -> AppConfig:
    """Reconstruye AppConfig desde una clave serializable."""
    return AppConfig(
        jira_email=str(data.get("jira_email") or ""),
        jira_api_token=str(data.get("jira_api_token") or ""),
        workspace_id=str(data.get("workspace_id") or ""),
        site=str(data.get("site") or ""),
        openai_api_key=str(data.get("openai_api_key") or ""),
        openai_model=str(data.get("openai_model") or "gpt-4o-mini"),
        rovo_api_key=str(data.get("rovo_api_key") or ""),
        rovo_enabled=str(data.get("rovo_enabled") or "").strip().lower() in {"1", "true", "yes", "si", "sí"},
    )


def append_error_events(events: list[dict[str, Any]]) -> None:
    """Inserta eventos al log de errores de sesión sin asumir contexto fuera del hilo principal."""
    if not events:
        return
    try:
        log = st.session_state.setdefault("error_log", [])
        for event in events:
            log.insert(0, event)
        st.session_state["error_log"] = log[:20]
    except Exception:
        return


def push_structured_error(error: StructuredError, sink: list[dict[str, Any]] | None = None) -> None:
    """Guarda error HTTP estructurado; en workers puede acumularlo en un sink local."""
    if int(error.status_code or 0) < 400:
        return
    event = dict(error.__dict__)
    if sink is not None:
        sink.append(event)
        return
    append_error_events([event])


def push_app_error(context: str, detail: str, sink: list[dict[str, Any]] | None = None) -> None:
    """Registra errores locales no HTTP para soporte y debugging."""
    event = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "context": context,
        "detail": detail,
        "error_type": "APP",
    }
    if sink is not None:
        sink.append(event)
        return
    append_error_events([event])


def escape_html_text(value: Any) -> str:
    """Escapa texto dinámico antes de insertarlo en bloques HTML de Streamlit."""
    return html.escape(str(value or ""), quote=True)


def append_movimiento_to_file(movimiento: MovimientoAsset) -> None:
    """Persiste un movimiento al archivo JSONL local."""
    try:
        with MOVEMENTS_FILE.open("a", encoding="utf-8") as handle:
            handle.write(f"{json.dumps(movimiento.__dict__, ensure_ascii=False)}\n")
    except OSError as exc:
        push_app_error("append_movimiento_to_file", f"{MOVEMENTS_FILE.name}: {exc}")


def log_movimiento(
    config: AppConfig,
    asset: dict[str, Any] | None,
    tipo_accion: str,
    campo: str,
    valor_anterior: str,
    valor_nuevo: str,
    resultado: str,
    detalle: str,
    identificador: str = "",
) -> None:
    """Registra movimiento en memoria, archivo y action_log."""
    source = asset or {}
    movement = MovimientoAsset(
        timestamp=datetime.now().isoformat(timespec="seconds"),
        tipo_accion=tipo_accion,
        identificador=identificador or str(source.get("serial_number") or source.get("hostname") or source.get("jira_key") or ""),
        jira_key=str(source.get("jira_key") or ""),
        campo_modificado=campo,
        valor_anterior=str(valor_anterior or ""),
        valor_nuevo=str(valor_nuevo or ""),
        usuario_asignado=str(source.get("assigned_to") or ""),
        ejecutado_por=config.jira_email,
        resultado=resultado,
        detalle=detalle,
    )
    moves = st.session_state.setdefault("movimientos", [])
    moves.append(movement.__dict__)
    st.session_state["movimientos"] = moves[-3000:]
    append_movimiento_to_file(movement)
    action_log = st.session_state.setdefault("action_log", [])
    action_log.append(
        {
            "timestamp": movement.timestamp,
            "tipo_accion": movement.tipo_accion,
            "identificador": movement.identificador,
            "resultado": movement.resultado,
            "detalle": movement.detalle,
        }
    )
    st.session_state["action_log"] = action_log[-1000:]


def apply_global_filter(assets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Aplica filtros globales de país y compañía al listado de activos."""
    countries = set(st.session_state.get("global_filter_countries", []))
    companies = set(st.session_state.get("global_filter_companies", []))
    if not countries and not companies:
        return assets
    filtered: list[dict[str, Any]] = []
    for asset in assets:
        if countries and str(asset.get("country") or "") not in countries:
            continue
        if companies and str(asset.get("company") or "") not in companies:
            continue
        filtered.append(asset)
    return filtered


def run_anomaly_detection(assets: list[dict[str, Any]]) -> dict[str, Any]:
    """Detecta anomalías operativas y de calidad en el inventario."""
    by_assignee = Counter(str(a.get("assigned_to") or "").strip() for a in assets if str(a.get("assigned_to") or "").strip())
    duplicates = detect_duplicates(assets)
    serial_groups = duplicates.get("serial_duplicates", [])
    host_groups = duplicates.get("hostname_duplicates", [])
    today = datetime.now().date()
    garantia_vencida_en_uso = 0
    en_uso_sin_asignado = 0
    sin_tipo = 0
    for asset in assets:
        if normalize_text(asset.get("status")) == normalize_text("en uso") and not str(asset.get("assigned_to") or "").strip():
            en_uso_sin_asignado += 1
        if not str(asset.get("object_type_id") or "").strip():
            sin_tipo += 1
        w = parse_date(str(asset.get("warranty_date", "")).split("|")[0].strip())
        if w and w.date() < today and normalize_text(asset.get("status")) == normalize_text("en uso"):
            garantia_vencida_en_uso += 1
    asignado_multiple = sum(1 for _, c in by_assignee.items() if c > 3)
    return {
        "en_uso_sin_asignado": en_uso_sin_asignado,
        "asignado_multiple": asignado_multiple,
        "garantia_vencida_en_uso": garantia_vencida_en_uso,
        "serial_duplicado": len(serial_groups),
        "hostname_duplicado": len(host_groups),
        "sin_tipo": sin_tipo,
        "total": en_uso_sin_asignado + asignado_multiple + garantia_vencida_en_uso + len(serial_groups) + len(host_groups) + sin_tipo,
    }


def parse_chat_response_for_table(text: str) -> tuple[str, Any]:
    """Extrae tablas simples desde respuestas markdown con listas separadas por `|`."""
    if pd is None:
        return text, None
    lines = [line for line in str(text or "").splitlines()]
    asset_lines = [line.strip() for line in lines if line.strip().startswith("- ") and "|" in line and len(line.strip()) > 5]
    if len(asset_lines) < 3:
        return text, None

    def clean_cell(cell: str) -> str:
        value = re.sub(r"\*\*(.+?)\*\*", r"\1", str(cell or ""))
        value = re.sub(r"`(.+?)`", r"\1", value)
        return value.strip()

    rows: list[list[str]] = []
    for line in asset_lines:
        content = line[2:].strip()
        cols = [clean_cell(c) for c in content.split("|")]
        rows.append(cols)

    max_cols = max(len(row) for row in rows)
    normalized_rows = [row + [""] * (max_cols - len(row)) for row in rows]
    header_candidates = {
        2: ["Activo", "Detalle"],
        3: ["Activo", "Estado", "Asignado"],
        4: ["Activo", "Hostname", "Estado", "País"],
        5: ["Activo", "Hostname", "Serial", "Estado", "Asignado"],
        6: ["Activo", "Hostname", "Serial", "Estado", "País", "Compañía"],
        7: ["Activo", "Hostname", "Serial", "Estado", "Asignado", "País", "Modelo"],
    }
    columns = header_candidates.get(max_cols, [f"Col {i+1}" for i in range(max_cols)])
    if len(columns) < max_cols:
        columns += [f"Col {i+1}" for i in range(len(columns), max_cols)]
    df = pd.DataFrame(normalized_rows, columns=columns[:max_cols])
    df = df.replace("", pd.NA).dropna(how="all").fillna("")

    asset_line_set = set(asset_lines)
    intro_lines = [line for line in lines if line.strip() not in asset_line_set]
    intro = "\n".join(line for line in intro_lines if line.strip()).strip()
    return intro or "Resultado:", df


def encode_chat_payload(text: str, charts: list[dict[str, Any]] | None = None) -> str:
    payload = {"text": text, "charts": charts or []}
    return CHAT_PAYLOAD_PREFIX + json.dumps(payload, ensure_ascii=False)


def decode_chat_payload(content: Any) -> tuple[str, list[dict[str, Any]]]:
    raw = str(content or "")
    if not raw.startswith(CHAT_PAYLOAD_PREFIX):
        return raw, []
    try:
        payload = json.loads(raw[len(CHAT_PAYLOAD_PREFIX):])
    except json.JSONDecodeError:
        return raw, []
    text = str(payload.get("text") or "")
    charts = payload.get("charts") or []
    return text, charts if isinstance(charts, list) else []


def remember_dashboard_response(prompt: str, response: str) -> None:
    text, charts = decode_chat_payload(response)
    if not charts:
        return
    st.session_state["last_dashboard_prompt"] = prompt
    st.session_state["last_dashboard_text"] = text
    st.session_state["last_dashboard_charts"] = charts
    st.session_state["last_dashboard_updated_at"] = datetime.now().isoformat(timespec="seconds")


def restore_dashboard_state_from_history() -> None:
    if st.session_state.get("last_dashboard_charts"):
        return
    history = st.session_state.get("chat_history", [])
    for idx in range(len(history) - 1, -1, -1):
        message = history[idx]
        if message.get("role") != "assistant":
            continue
        text, charts = decode_chat_payload(message.get("content"))
        if not charts:
            continue
        prompt = ""
        for prev_idx in range(idx - 1, -1, -1):
            previous = history[prev_idx]
            if previous.get("role") == "user":
                prompt = str(previous.get("content") or "")
                break
        st.session_state["last_dashboard_prompt"] = prompt
        st.session_state["last_dashboard_text"] = text
        st.session_state["last_dashboard_charts"] = charts
        return


def generate_html_report(assets_filtered: list[dict[str, Any]]) -> str:
    """Genera reporte HTML imprimible del inventario filtrado."""
    total = len(assets_filtered)
    in_use = sum(1 for a in assets_filtered if normalize_text(a.get("status")) == normalize_text("en uso"))
    stock = sum(1 for a in assets_filtered if normalize_text(a.get("status")) in {normalize_text("stock nuevo"), normalize_text("stock usado")})
    by_country = Counter((a.get("country") or "Sin país") for a in assets_filtered)
    by_company = Counter((a.get("company") or "Sin compañía") for a in assets_filtered)
    rows = []
    for a in assets_filtered[:2000]:
        rows.append(
            f"<tr><td>{escape_html_text(a.get('jira_key',''))}</td><td>{escape_html_text(a.get('name',''))}</td><td>{escape_html_text(a.get('status',''))}</td>"
            f"<td>{escape_html_text(a.get('assigned_to',''))}</td><td>{escape_html_text(a.get('country',''))}</td><td>{escape_html_text(a.get('company',''))}</td></tr>"
        )
    countries_html = "".join(f"<li>{escape_html_text(k)}: {v}</li>" for k, v in by_country.items())
    companies_html = "".join(f"<li>{escape_html_text(k)}: {v}</li>" for k, v in by_company.items())
    return f"""
    <html><head><meta charset='utf-8'><style>
    body{{font-family:Arial,sans-serif;padding:16px}} .kpis{{display:flex;gap:12px}}
    .card{{border:1px solid #ddd;padding:8px 12px;border-radius:8px}}
    table{{width:100%;border-collapse:collapse}} th,td{{border:1px solid #ddd;padding:6px;font-size:12px}}
    th{{background:#003262;color:white}} @media print{{.noprint{{display:none}}}}
    </style></head><body>
    <h1>Reporte Inventario Uala</h1>
    <div class='kpis'><div class='card'>Total: {total}</div><div class='card'>En uso: {in_use}</div><div class='card'>Stock: {stock}</div></div>
    <h3>Resumen por país</h3><ul>{countries_html}</ul>
    <h3>Resumen por compañía</h3><ul>{companies_html}</ul>
    <h3>Activos</h3>
    <table><thead><tr><th>Jira</th><th>Nombre</th><th>Estado</th><th>Asignado</th><th>País</th><th>Compañía</th></tr></thead>
    <tbody>{''.join(rows)}</tbody></table>
    <p>Generado: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
    </body></html>
    """


def backoff_sleep(response: httpx.Response | None, attempt: int) -> None:
    """Aplica espera exponencial respetando Retry-After cuando existe."""
    if response is not None:
        retry_after = response.headers.get("Retry-After")
        if retry_after:
            try:
                time.sleep(float(retry_after))
                return
            except ValueError:
                pass
    time.sleep(min(2**attempt, 30))


def jira_request_with_retry(
    method: str,
    url: str,
    *,
    auth: BasicAuth,
    headers: dict[str, str],
    json_payload: dict[str, Any] | None = None,
    params: dict[str, Any] | None = None,
    max_attempts: int = 5,
    error_sink: list[dict[str, Any]] | None = None,
) -> httpx.Response:
    """Ejecuta request HTTP con reintentos y backoff para Jira."""
    last_error = None
    for attempt in range(max_attempts):
        response = None
        try:
            with httpx.Client(timeout=REQUEST_TIMEOUT) as client:
                response = client.request(
                    method=method,
                    url=url,
                    auth=auth,
                    headers=headers,
                    json=json_payload,
                    params=params,
                )
            if response.status_code == 429:
                backoff_sleep(response, attempt)
                continue
            response.raise_for_status()
            return response
        except httpx.HTTPError as exc:
            last_error = exc
            status_code = int(response.status_code) if response is not None else 0
            body = response.text[:2000] if response is not None else ""
            push_structured_error(
                StructuredError(
                    url=url,
                    method=method,
                    status_code=status_code,
                    response_body=body,
                    timestamp=datetime.now().isoformat(timespec="seconds"),
                    context="jira_request_with_retry",
                ),
                sink=error_sink,
            )
            should_retry = response is None or response.status_code >= 500 or response.status_code == 429
            if attempt < max_attempts - 1 and should_retry:
                backoff_sleep(response, attempt)
                continue
            raise RuntimeError(f"Error consultando Jira Assets: {exc}") from exc
    raise RuntimeError(f"Error consultando Jira Assets: {last_error}")


async def jira_request_with_retry_async(
    client: httpx.AsyncClient,
    method: str,
    url: str,
    *,
    auth: BasicAuth,
    headers: dict[str, str],
    json_payload: dict[str, Any] | None = None,
    params: dict[str, Any] | None = None,
    max_attempts: int = 5,
    error_sink: list[dict[str, Any]] | None = None,
) -> httpx.Response:
    """Ejecuta request async con reintentos y backoff para Jira."""
    last_error: Exception | None = None
    for attempt in range(max_attempts):
        response: httpx.Response | None = None
        try:
            response = await client.request(
                method=method,
                url=url,
                auth=auth,
                headers=headers,
                json=json_payload,
                params=params,
            )
            if response.status_code == 429:
                await asyncio.sleep(min(2**attempt, 30))
                continue
            response.raise_for_status()
            return response
        except httpx.HTTPError as exc:
            last_error = exc
            status_code = int(response.status_code) if response is not None else 0
            body = response.text[:2000] if response is not None else ""
            push_structured_error(
                StructuredError(
                    url=url,
                    method=method,
                    status_code=status_code,
                    response_body=body,
                    timestamp=datetime.now().isoformat(timespec="seconds"),
                    context="jira_request_with_retry_async",
                ),
                sink=error_sink,
            )
            should_retry = response is None or response.status_code >= 500 or response.status_code == 429
            if attempt < max_attempts - 1 and should_retry:
                await asyncio.sleep(min(2**attempt, 30))
                continue
            raise RuntimeError(f"Error consultando Jira Assets: {exc}") from exc
    raise RuntimeError(f"Error consultando Jira Assets: {last_error}")


async def _fetch_type_async(
    config: AppConfig,
    auth: BasicAuth,
    headers: dict[str, str],
    type_id: str,
    aql_query: str,
    scope_type_ids: list[str] | None = None,
) -> list[dict[str, Any]]:
    """Consulta de forma asíncrona un objectType puntual y devuelve registros limpios."""
    urls = [
        f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/aql",
        f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/navlist/aql",
        f"{config.site}/rest/servicedeskapi/assets/workspace/{config.workspace_id}/v1/object/aql",
    ]
    ql = combine_schema_aql(
        f"({aql_query}) AND objectTypeId = {type_id}" if aql_query.strip() else f"objectTypeId = {type_id}",
        type_ids=scope_type_ids,
    )
    payload_templates = [
        {"resultsPerPage": PAGE_SIZE, "page": 1},
        {"resultPerPage": PAGE_SIZE, "page": 1},
        {"maxResults": PAGE_SIZE, "startAt": 0},
    ]
    async with httpx.AsyncClient(timeout=REQUEST_TIMEOUT) as client:
        for template in payload_templates:
            records: list[dict[str, Any]] = []
            seen: set[str] = set()
            for url in urls:
                page = 1
                try:
                    while page <= 500:
                        payload: dict[str, Any] = {"qlQuery": ql, "includeAttributes": True}
                        if "page" in template:
                            payload["page"] = page
                            if "resultsPerPage" in template:
                                payload["resultsPerPage"] = template["resultsPerPage"]
                            if "resultPerPage" in template:
                                payload["resultPerPage"] = template["resultPerPage"]
                        else:
                            payload["maxResults"] = template["maxResults"]
                            payload["startAt"] = (page - 1) * int(template["maxResults"])

                        response = await jira_request_with_retry_async(
                            client,
                            "POST",
                            url,
                            auth=auth,
                            headers=headers,
                            json_payload=payload,
                        )
                        body = response.json()
                        values = body.get("values") or body.get("objectEntries") or []
                        if not values:
                            break

                        added = 0
                        for asset in values:
                            object_id = str(asset.get("id") or asset.get("objectId") or asset.get("objectKey") or "")
                            if object_id and object_id in seen:
                                continue
                            if object_id:
                                seen.add(object_id)
                            records.append(clean_asset_object(asset).to_dict())
                            added += 1

                        if body.get("isLast", False) or added == 0:
                            break
                        if "page" not in template and len(values) < int(template["maxResults"]):
                            break
                        page += 1

                    if records:
                        return records
                except RuntimeError:
                    continue
    return []


def fetch_type_sync(
    config: AppConfig,
    auth: BasicAuth,
    headers: dict[str, str],
    type_id: str,
    aql_query: str,
    scope_type_ids: list[str] | None = None,
) -> list[dict[str, Any]]:
    """Consulta síncrona robusta por objectType con paginación defensiva."""
    ql = combine_schema_aql(
        f"({aql_query}) AND objectTypeId = {type_id}" if aql_query.strip() else f"objectTypeId = {type_id}",
        type_ids=scope_type_ids,
    )
    return paginate_aql_sync(config, auth, headers, ql)


def extract_values_from_aql_body(body: dict[str, Any]) -> list[dict[str, Any]]:
    values = body.get("values") or body.get("objectEntries") or body.get("objects") or []
    return values if isinstance(values, list) else []


def extract_total_from_aql_body(body: dict[str, Any]) -> int | None:
    for key in ("totalFilterCount", "totalCount", "total", "count"):
        value = body.get(key)
        if isinstance(value, int) and value >= 0:
            return value
        if isinstance(value, str) and value.isdigit():
            return int(value)
    return None


def extract_has_more_from_aql_body(body: dict[str, Any]) -> bool | None:
    for key in ("hasMoreResults", "hasMore", "isLast"):
        value = body.get(key)
        if isinstance(value, bool):
            if key == "isLast":
                return not value
            return value
    return None


def paginate_aql_sync(
    config: AppConfig,
    auth: BasicAuth,
    headers: dict[str, str],
    ql_query: str,
) -> list[dict[str, Any]]:
    """Ejecuta AQL contra múltiples endpoints y formatos de paginación, devolviendo el mejor set."""
    legacy_get_url = f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/aql/objects"
    page_size = max(50, min(int(PAGE_SIZE), 200))
    try:
        records: list[dict[str, Any]] = []
        seen: set[str] = set()
        with httpx.Client(timeout=REQUEST_TIMEOUT) as client:
            for page in range(1, 501):
                response = client.request(
                    "GET",
                    legacy_get_url,
                    auth=auth,
                    headers={"Accept": "application/json"},
                    params={
                        "qlQuery": ql_query,
                        "page": page,
                        "resultPerPage": page_size,
                        "includeAttributes": "true",
                    },
                )
                response.raise_for_status()
                body = response.json()
                values = extract_values_from_aql_body(body)
                if not values:
                    break

                added = 0
                for asset in values:
                    object_id = str(asset.get("id") or asset.get("objectId") or asset.get("objectKey") or "")
                    if object_id and object_id in seen:
                        continue
                    if object_id:
                        seen.add(object_id)
                    records.append(clean_asset_object(asset).to_dict())
                    added += 1

                if added == 0 or len(values) < page_size:
                    break
        if records:
            return records
    except Exception:
        pass

    urls = [
        f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/aql",
        f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/navlist/aql",
        f"{config.site}/rest/servicedeskapi/assets/workspace/{config.workspace_id}/v1/object/aql",
    ]
    payload_templates = [
        {"mode": "page", "size_key": "resultsPerPage", "page_key": "page", "size": page_size},
        {"mode": "page", "size_key": "resultPerPage", "page_key": "page", "size": page_size},
        {"mode": "offset", "size_key": "maxResults", "offset_key": "startAt", "size": page_size},
    ]
    best_records: list[dict[str, Any]] = []
    with httpx.Client(timeout=REQUEST_TIMEOUT) as client:
        for url in urls:
            for template in payload_templates:
                records: list[dict[str, Any]] = []
                seen: set[str] = set()
                repeated_pages = 0
                cursor = 1
                try:
                    while cursor <= 500:
                        payload: dict[str, Any] = {"qlQuery": ql_query, "includeAttributes": True}
                        if template["mode"] == "page":
                            payload[template["page_key"]] = cursor
                            payload[template["size_key"]] = template["size"]
                        else:
                            payload[template["size_key"]] = template["size"]
                            payload[template["offset_key"]] = (cursor - 1) * int(template["size"])

                        response = client.request("POST", url, auth=auth, headers=headers, json=payload)
                        response.raise_for_status()
                        body = response.json()
                        values = extract_values_from_aql_body(body)
                        if not values:
                            break

                        added = 0
                        for asset in values:
                            object_id = str(asset.get("id") or asset.get("objectId") or asset.get("objectKey") or "")
                            if object_id and object_id in seen:
                                continue
                            if object_id:
                                seen.add(object_id)
                            records.append(clean_asset_object(asset).to_dict())
                            added += 1

                        # Jira Cloud puede reportar hasMore/total inconsistentes (ej: total=25 fijo).
                        # Avanzamos por cursor hasta que no entren filas nuevas de forma consecutiva.
                        if added == 0:
                            repeated_pages += 1
                        else:
                            repeated_pages = 0
                        if repeated_pages >= 2:
                            break

                        cursor += 1

                    if len(records) > len(best_records):
                        best_records = records
                except Exception:
                    continue
    return best_records


def fetch_objects_by_type_bruteforce(
    config: AppConfig,
    auth: BasicAuth,
    headers: dict[str, str],
    type_id: str,
    error_sink: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Fallback agresivo: lista objetos por objectType usando endpoint dedicado."""
    urls = [
        f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/objecttype/{type_id}/objects",
        f"{config.site}/rest/servicedeskapi/assets/workspace/{config.workspace_id}/v1/objecttype/{type_id}/objects",
    ]
    page_size = max(50, min(int(PAGE_SIZE), 200))
    templates = [
        {"mode": "page", "size_key": "resultsPerPage", "page_key": "page", "size": page_size},
        {"mode": "page", "size_key": "resultPerPage", "page_key": "page", "size": page_size},
        {"mode": "offset", "size_key": "maxResults", "offset_key": "startAt", "size": page_size},
    ]

    best: list[dict[str, Any]] = []
    for url in urls:
        for template in templates:
            seen: set[str] = set()
            rows: list[dict[str, Any]] = []
            repeated_pages = 0
            cursor = 1
            try:
                while cursor <= 800:
                    payload: dict[str, Any] = {"includeAttributes": True}
                    if template["mode"] == "page":
                        payload[template["page_key"]] = cursor
                        payload[template["size_key"]] = template["size"]
                    else:
                        payload[template["size_key"]] = template["size"]
                        payload[template["offset_key"]] = (cursor - 1) * int(template["size"])

                    response = jira_request_with_retry(
                        "POST",
                        url,
                        auth=auth,
                        headers=headers,
                        json_payload=payload,
                        error_sink=error_sink,
                    )
                    body = response.json()
                    values = extract_values_from_aql_body(body) if isinstance(body, dict) else (body if isinstance(body, list) else [])
                    if not values:
                        break

                    added = 0
                    for asset in values:
                        object_id = str(asset.get("id") or asset.get("objectId") or asset.get("objectKey") or "")
                        if object_id and object_id in seen:
                            continue
                        if object_id:
                            seen.add(object_id)
                        rows.append(clean_asset_object(asset).to_dict())
                        added += 1

                    # Evita confiar en hasMore/total cuando Jira responde metadatos inconsistentes.
                    if added == 0:
                        repeated_pages += 1
                    else:
                        repeated_pages = 0
                    if repeated_pages >= 2:
                        break
                    cursor += 1

                if len(rows) > len(best):
                    best = rows
            except Exception:
                continue
    return best


def normalize_type_id_list(values: Any) -> list[str]:
    """Normaliza listas de type ids y elimina duplicados conservando orden."""
    normalized: list[str] = []
    for value in values or []:
        type_id = str(value or "").strip()
        if type_id and type_id not in normalized:
            normalized.append(type_id)
    return normalized


def build_hardware_scope_type_ids(discovered_type_ids: list[str] | None = None) -> list[str]:
    """Resuelve el scope activo de hardware sin depender de session_state."""
    root_id = str(GENERAL_HARDWARE_TYPE_ID).strip()
    fallback = normalize_type_id_list(KNOWN_OBJECT_TYPE_IDS)
    combined = [root_id] if root_id else []
    for type_id in normalize_type_id_list(discovered_type_ids) or fallback:
        if type_id not in combined:
            combined.append(type_id)
    return combined or fallback


def get_active_hardware_type_ids(discovered_type_ids: list[str] | None = None) -> list[str]:
    """Resuelve el scope activo de hardware: parent 211 + descendientes descubiertos."""
    if discovered_type_ids is None:
        discovered_type_ids = st.session_state.get("discovered_type_ids") or []
    return build_hardware_scope_type_ids(discovered_type_ids)


def get_schema_scan_type_ids(
    all_schema_type_ids: list[str] | None = None,
    discovered_type_ids: list[str] | None = None,
) -> list[str]:
    """Devuelve todos los objectTypeIds del esquema para escaneos de recuperación."""
    normalized_all = normalize_type_id_list(
        all_schema_type_ids if all_schema_type_ids is not None else st.session_state.get("all_schema_type_ids") or []
    )
    if normalized_all:
        return normalized_all
    return build_hardware_scope_type_ids(discovered_type_ids if discovered_type_ids is not None else st.session_state.get("discovered_type_ids") or [])


def build_schema_only_aql(aql_query: str) -> str:
    """Replica la estrategia legacy: primero barrer el esquema completo."""
    query = (aql_query or "").strip()
    base = f"objectSchemaId = {SCHEMA_ID}"
    if not query:
        return base
    return f"{base} AND ({query})"


def build_primary_aql(aql_query: str, type_ids: list[str] | None = None) -> str:
    """Define la consulta primaria según el scope activo."""
    if ASSETS_SCOPE_MODE in {"hardware", "strict_hardware"}:
        return combine_schema_aql(aql_query, type_ids=type_ids)
    return build_schema_only_aql(aql_query)


def combine_schema_aql(aql_query: str, type_ids: list[str] | None = None) -> str:
    type_ids = normalize_type_id_list(type_ids) or get_active_hardware_type_ids()
    hardware_scope = " OR ".join(f"objectTypeId = {type_id}" for type_id in type_ids)
    base = f"objectSchemaId = {SCHEMA_ID} AND ({hardware_scope})"
    query = (aql_query or "").strip()
    if not query:
        return base
    return f"{base} AND ({query})"


def fetch_schema_object_type_rows(config: AppConfig, schema_id: str, auth: BasicAuth, headers: dict[str, str]) -> list[dict[str, Any]]:
    """Descubre filas de object types del esquema con paginación defensiva."""
    if not config.workspace_id:
        return []
    endpoints = [
        f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/objectschema/{schema_id}/objecttypes",
        f"{config.site}/rest/servicedeskapi/assets/workspace/{config.workspace_id}/v1/objectschema/{schema_id}/objecttypes",
    ]
    page_size = 200
    templates = [
        {"mode": "page", "size_key": "resultsPerPage", "page_key": "page", "size": page_size},
        {"mode": "page", "size_key": "resultPerPage", "page_key": "page", "size": page_size},
        {"mode": "offset", "size_key": "maxResults", "offset_key": "startAt", "size": page_size},
    ]

    def _collect_flat_rows(node: Any, collector: list[dict[str, Any]]) -> None:
        if isinstance(node, dict):
            type_id = str(node.get("id") or node.get("objectTypeId") or "").strip()
            if type_id:
                collector.append(node)
            for key in ("values", "objectTypes", "children", "childObjectTypes", "items"):
                child = node.get(key)
                if isinstance(child, (list, dict)):
                    _collect_flat_rows(child, collector)
            return
        if isinstance(node, list):
            for item in node:
                _collect_flat_rows(item, collector)

    best_rows: list[dict[str, Any]] = []
    for url in endpoints:
        for template in templates:
            rows: list[dict[str, Any]] = []
            seen_ids: set[str] = set()
            cursor = 1
            repeated_pages = 0
            try:
                while cursor <= 500:
                    params: dict[str, Any] = {}
                    if template["mode"] == "page":
                        params[template["page_key"]] = cursor
                        params[template["size_key"]] = template["size"]
                    else:
                        params[template["size_key"]] = template["size"]
                        params[template["offset_key"]] = (cursor - 1) * int(template["size"])

                    response = jira_request_with_retry("GET", url, auth=auth, headers=headers, params=params)
                    body = response.json()
                    page_rows: list[dict[str, Any]] = []
                    _collect_flat_rows(body, page_rows)
                    if not page_rows:
                        break

                    added = 0
                    for row in page_rows:
                        type_id = str(row.get("id") or row.get("objectTypeId") or "").strip()
                        if not type_id or type_id in seen_ids:
                            continue
                        seen_ids.add(type_id)
                        rows.append(row)
                        added += 1

                    # No confiar en metadatos de paginación; cortar solo cuando no aparecen filas nuevas.
                    if added == 0:
                        repeated_pages += 1
                    else:
                        repeated_pages = 0
                    if repeated_pages >= 2:
                        break
                    cursor += 1
            except Exception:
                continue

            if len(rows) > len(best_rows):
                best_rows = rows
    return best_rows


def fetch_schema_object_type_ids(config: AppConfig, schema_id: str) -> list[str]:
    """Descubre objectTypeIds del árbol bajo GENERAL_HARDWARE_TYPE_ID (excluye el root)."""
    auth, headers = build_auth_headers(config)
    flat_rows = fetch_schema_object_type_rows(config, schema_id, auth, headers)
    if not flat_rows:
        return []
    root_id = str(GENERAL_HARDWARE_TYPE_ID).strip()

    children_by_parent: dict[str, list[str]] = defaultdict(list)
    all_ids: set[str] = set()
    for row in flat_rows:
        type_id = str(row.get("id") or row.get("objectTypeId") or "").strip()
        if not type_id:
            continue
        all_ids.add(type_id)
        parent_id = str(
            row.get("parentObjectTypeId")
            or (row.get("parentObjectType") or {}).get("id")
            or (row.get("parentObjectTypeBean") or {}).get("id")
            or ""
        ).strip()
        if parent_id:
            children_by_parent[parent_id].append(type_id)

    if root_id not in all_ids:
        # Si el root configurado no existe en el esquema, devolvemos todos para no perder cobertura.
        return sorted(all_ids)

    queue = list(children_by_parent.get(root_id, []))
    descendants: list[str] = []
    seen_desc: set[str] = set()
    while queue:
        current = queue.pop(0)
        if current in seen_desc:
            continue
        seen_desc.add(current)
        descendants.append(current)
        queue.extend(children_by_parent.get(current, []))
    return descendants


def fetch_all_schema_object_type_ids(config: AppConfig, schema_id: str) -> list[str]:
    """Devuelve todos los objectTypeIds del esquema, sin filtrar por parent."""
    auth, headers = build_auth_headers(config)
    rows = fetch_schema_object_type_rows(config, schema_id, auth, headers)
    return sorted({str(row.get("id") or row.get("objectTypeId") or "").strip() for row in rows if str(row.get("id") or row.get("objectTypeId") or "").strip()})


def fetch_object_type_attributes(config: AppConfig, object_type_id: str, auth: BasicAuth, headers: dict[str, str]) -> list[dict[str, Any]]:
    urls = [
        f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/objecttype/{object_type_id}/attributes",
        f"{config.site}/rest/servicedeskapi/assets/workspace/{config.workspace_id}/v1/objecttype/{object_type_id}/attributes",
    ]
    for url in urls:
        try:
            response = jira_request_with_retry("GET", url, auth=auth, headers=headers)
            body = response.json()
            if isinstance(body, list):
                return body
            if isinstance(body, dict):
                values = body.get("values")
                if isinstance(values, list):
                    return values
        except Exception:
            continue
    return []


def parse_attribute_options_payload(payload: Any) -> list[tuple[str, str]]:
    """Normaliza respuestas de Jira que exponen opciones en distintos shapes."""
    if isinstance(payload, list):
        raw_items = payload
    elif isinstance(payload, dict):
        raw_items = payload.get("values") or payload.get("options") or payload.get("items") or []
    else:
        raw_items = []

    parsed: list[tuple[str, str]] = []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        option_id = str(item.get("id") or item.get("optionId") or item.get("value") or "").strip()
        option_name = str(item.get("name") or item.get("label") or item.get("displayValue") or "").strip()
        if option_id and option_name:
            parsed.append((option_name, option_id))
    return parsed


def fetch_attribute_option_lookup(config: AppConfig, attr_id: str, auth: BasicAuth, headers: dict[str, str]) -> dict[str, str]:
    """Trae opciones válidas de un atributo y las indexa por label normalizado."""
    cache = st.session_state.setdefault("mass_upload_option_cache", {})
    cache_key = str(attr_id).strip()
    if cache_key in cache:
        return dict(cache[cache_key])

    urls = [
        f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/objecttypeattribute/{attr_id}/values",
        f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/objecttypeattribute/{attr_id}",
        f"{config.site}/rest/servicedeskapi/assets/workspace/{config.workspace_id}/v1/objecttypeattribute/{attr_id}/values",
        f"{config.site}/rest/servicedeskapi/assets/workspace/{config.workspace_id}/v1/objecttypeattribute/{attr_id}",
    ]
    lookup: dict[str, str] = {}
    for url in urls:
        try:
            response = jira_request_with_retry("GET", url, auth=auth, headers=headers)
            for option_name, option_id in parse_attribute_options_payload(response.json()):
                normalized = normalize_lookup_key(option_name)
                if normalized and option_id:
                    lookup[normalized] = option_id
        except Exception:
            continue

    cache[cache_key] = dict(lookup)
    return lookup


def fetch_reference_object_lookup(
    config: AppConfig,
    reference_object_type_id: str,
    auth: BasicAuth,
    headers: dict[str, str],
) -> dict[str, str]:
    """Trae objetos de referencia y arma un lookup label/key -> objectKey."""
    cache = st.session_state.setdefault("mass_upload_reference_cache", {})
    cache_key = str(reference_object_type_id).strip()
    if cache_key in cache:
        return dict(cache[cache_key])

    urls = [
        f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/aql",
        f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/navlist/aql",
        f"{config.site}/rest/servicedeskapi/assets/workspace/{config.workspace_id}/v1/object/aql",
    ]
    lookup: dict[str, str] = {}
    for url in urls:
        start_at = 0
        while True:
            try:
                response = jira_request_with_retry(
                    "POST",
                    url,
                    auth=auth,
                    headers=headers,
                    json_payload={
                        "qlQuery": f"objectTypeId = {cache_key}",
                        "includeAttributes": True,
                        "maxResults": 200,
                        "startAt": start_at,
                    },
                )
            except Exception:
                break

            body = response.json() if isinstance(response.json(), dict) else {}
            values = body.get("values") or body.get("objectEntries") or []
            if not isinstance(values, list) or not values:
                break

            for item in values:
                if not isinstance(item, dict):
                    continue
                object_key = str(item.get("objectKey") or item.get("key") or "").strip()
                label = str(item.get("label") or item.get("name") or "").strip()
                if object_key:
                    lookup[normalize_lookup_key(object_key)] = object_key
                if label and object_key:
                    lookup[normalize_lookup_key(label)] = object_key
                    compact_label = compact_lookup_key(label)
                    if compact_label:
                        lookup[compact_label] = object_key
                for attribute in item.get("attributes", []) or []:
                    attr_value = extract_attr_text(attribute)
                    if not attr_value or not object_key:
                        continue
                    lookup[normalize_lookup_key(attr_value)] = object_key
                    compact_attr = compact_lookup_key(attr_value)
                    if compact_attr:
                        lookup[compact_attr] = object_key

            if len(values) < 200:
                break
            start_at += 200

        if lookup:
            break

    cache[cache_key] = dict(lookup)
    return lookup


def create_reference_object(
    config: AppConfig,
    reference_object_type_id: str,
    display_name: str,
    auth: BasicAuth,
    headers: dict[str, str],
) -> str | None:
    """Crea un objeto de referencia simple cuando el catálogo no lo tiene todavía."""
    attrs = fetch_object_type_attributes(config, reference_object_type_id, auth, headers)
    candidate = None
    for attr in attrs:
        default_name = normalize_lookup_key((attr.get("defaultType") or {}).get("name") or attr.get("type") or "")
        if default_name != "text":
            continue
        if int(attr.get("minimumCardinality", 0) or 0) > 0:
            candidate = attr
            if "name" in normalize_lookup_key(attr.get("name") or ""):
                break
    if candidate is None:
        for attr in attrs:
            default_name = normalize_lookup_key((attr.get("defaultType") or {}).get("name") or attr.get("type") or "")
            if default_name == "text" and "name" in normalize_lookup_key(attr.get("name") or ""):
                candidate = attr
                break
    if candidate is None:
        return None

    url = f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/create"
    payload = {
        "objectTypeId": str(reference_object_type_id),
        "attributes": [
            {
                "objectTypeAttributeId": str(candidate.get("id") or ""),
                "objectAttributeValues": [{"value": display_name}],
            }
        ],
    }
    try:
        with httpx.Client(timeout=REQUEST_TIMEOUT) as client:
            response = client.post(url, auth=auth, headers=headers, json=payload)
        if response.status_code not in (200, 201):
            return None
        body = response.json() if response.text else {}
        created_key = str(body.get("objectKey") or body.get("key") or "").strip()
        if not created_key:
            return None
        cache = st.session_state.setdefault("mass_upload_reference_cache", {})
        cache_key = str(reference_object_type_id).strip()
        ref_lookup = dict(cache.get(cache_key, {}))
        ref_lookup[normalize_lookup_key(display_name)] = created_key
        ref_lookup[canonical_model_key(display_name)] = created_key
        ref_lookup[normalize_lookup_key(created_key)] = created_key
        cache[cache_key] = ref_lookup
        return created_key
    except httpx.HTTPError:
        return None


def resolve_reference_object_key(
    config: AppConfig,
    reference_object_type_id: str,
    raw_value: str,
    auth: BasicAuth,
    *,
    attr_id: str = "",
    headers: dict[str, str],
) -> str | None:
    """Resuelve texto visible a objectKey para atributos de referencia."""
    raw = str(raw_value or "").strip()
    if not raw:
        return None
    if re.fullmatch(r"[A-Z]+-\d+", raw):
        return raw

    lookup = fetch_reference_object_lookup(config, reference_object_type_id, auth, headers)
    if not lookup:
        return None

    candidates = [
        raw,
        canonical_category(raw),
        canonical_country(raw),
        normalize_company(raw),
    ]
    for candidate in candidates:
        normalized = normalize_lookup_key(candidate)
        if normalized and normalized in lookup:
            return lookup[normalized]
        compact = compact_lookup_key(candidate)
        if compact and compact in lookup:
            return lookup[compact]

    if attr_id == ID_MODELO:
        target_key = canonical_model_key(raw)
        if target_key:
            for key, object_key in lookup.items():
                if canonical_model_key(key) == target_key:
                    return object_key

    normalized_raw = normalize_lookup_key(raw)
    compact_raw = compact_lookup_key(raw)
    for key, object_key in lookup.items():
        if normalized_raw and (normalized_raw in key or key in normalized_raw):
            return object_key
        if compact_raw:
            compact_key = compact_lookup_key(key)
            if compact_key and (compact_raw in compact_key or compact_key in compact_raw):
                return object_key
    return None


def resolve_mass_upload_attribute_value(
    config: AppConfig,
    attr_id: str,
    raw_value: Any,
    attr_def: dict[str, Any],
    auth: BasicAuth,
    headers: dict[str, str],
) -> tuple[Any | None, str]:
    """Convierte el valor tabular al formato esperado por Jira Assets."""
    raw = normalize_tabular_value(raw_value)
    if not raw:
        return None, ""

    value = raw
    default_type = normalize_lookup_key((attr_def.get("defaultType") or {}).get("name") or attr_def.get("type") or "")
    reference_object_type_id = str((attr_def.get("referenceObjectType") or {}).get("id") or "").strip()

    if attr_id == ID_ESTADO:
        value = canonical_status(raw)
    elif attr_id == ID_PAIS:
        value = canonical_country(raw)
    elif attr_id == ID_COMPANIA:
        value = normalize_company(raw)
    elif attr_id == ID_COSTO:
        return parse_cost(raw), ""
    elif attr_id == ID_FECHA_COMPRA:
        parsed_date = parse_date(raw)
        return (parsed_date.strftime("%Y-%m-%d") if parsed_date else raw), ""
    elif attr_id == ID_FECHA_GARANTIA:
        return format_jira_datetime(raw), ""

    if reference_object_type_id:
        resolved = resolve_reference_object_key(
            config,
            reference_object_type_id,
            str(value),
            auth,
            attr_id=attr_id,
            headers=headers,
        )
        if not resolved and attr_id == ID_MODELO:
            resolved = create_reference_object(config, reference_object_type_id, str(value), auth, headers)
        if not resolved:
            return None, f"No pude resolver la referencia `{value}` para el atributo `{attr_def.get('name') or attr_id}`."
        return resolved, ""

    if "user" in default_type:
        account_id = resolve_user_account_id(config, str(value), auth)
        if not account_id:
            return None, ""
        return account_id, ""

    option_lookup = fetch_attribute_option_lookup(config, attr_id, auth, headers)
    if option_lookup:
        normalized = normalize_lookup_key(str(value))
        if normalized in option_lookup:
            return option_lookup[normalized], ""

    return value, ""


def build_asset_create_payload(config: AppConfig, row: dict[str, Any]) -> tuple[str, list[dict[str, Any]], list[str]]:
    """Arma payload de alta resolviendo referencias, usuarios y opciones."""
    type_id, attrs = build_asset_attributes_payload(row)
    auth, headers = build_auth_headers(config)
    attr_defs = fetch_object_type_attributes(config, type_id, auth, headers)
    attr_defs_by_id = {str(attr.get("id") or "").strip(): attr for attr in attr_defs if str(attr.get("id") or "").strip()}

    resolved_attrs: list[dict[str, Any]] = []
    issues: list[str] = []

    for attr in attrs:
        attr_id = str(attr.get("objectTypeAttributeId") or "").strip()
        raw_values = attr.get("objectAttributeValues") or []
        raw_value = raw_values[0].get("value") if raw_values else ""
        attr_def = attr_defs_by_id.get(attr_id, {})

        resolved_value, issue = resolve_mass_upload_attribute_value(config, attr_id, raw_value, attr_def, auth, headers)
        if issue:
            issues.append(issue)
            continue
        if resolved_value in (None, ""):
            continue
        resolved_attrs.append({"objectTypeAttributeId": attr_id, "objectAttributeValues": [{"value": resolved_value}]})

    return type_id, resolved_attrs, issues


def create_asset_from_payload(config: AppConfig, object_type_id: str, attrs: list[dict[str, Any]]) -> tuple[bool, str]:
    """Crea un asset usando la ruta de alta que sí utiliza el script batch existente."""
    auth, headers = build_auth_headers(config)
    url = f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/create"
    payload = {"objectTypeId": str(object_type_id), "attributes": attrs}
    try:
        with httpx.Client(timeout=REQUEST_TIMEOUT) as client:
            response = client.post(url, auth=auth, headers=headers, json=payload)
        if response.status_code in (200, 201):
            body = response.json() if response.text else {}
            created_key = str(body.get("objectKey") or body.get("key") or "").strip()
            return True, f"Creado{f' ({created_key})' if created_key else ''}"
        detail = response.text.strip()[:1000] or "sin detalle"
        return False, f"Error creando asset ({response.status_code}): {detail}"
    except httpx.HTTPError as exc:
        return False, f"Error creando asset: {exc}"


def fetch_schema_bridge(config: AppConfig) -> dict[str, list[dict[str, Any]]]:
    auth, headers = build_auth_headers(config)
    bridge: dict[str, list[dict[str, Any]]] = {}
    for type_id in KNOWN_OBJECT_TYPE_IDS:
        bridge[type_id] = fetch_object_type_attributes(config, type_id, auth, headers)
    return bridge


def resolve_user_account_id(config: AppConfig, email_or_name: str, auth: BasicAuth) -> str | None:
    query = (email_or_name or "").strip()
    if not query:
        return None
    url = f"{config.site}/rest/api/3/user/search"
    try:
        response = jira_request_with_retry(
            "GET",
            url,
            auth=auth,
            headers={"Accept": "application/json"},
            params={"query": query, "maxResults": 50},
        )
        users = response.json() if isinstance(response.json(), list) else []
        if not users:
            return None
        target = normalize_text(query)
        for user in users:
            mail = normalize_text(user.get("emailAddress", ""))
            name = normalize_text(user.get("displayName", ""))
            if target and (target == mail or target == name):
                return str(user.get("accountId") or "")
        return str(users[0].get("accountId") or "")
    except Exception:
        return None


def update_asset_assignment(config: AppConfig, object_id: str, object_type_id: str, assignee_text: str) -> tuple[bool, str]:
    auth, headers = build_auth_headers(config)
    assignee_value: str | dict[str, str] = assignee_text
    account_id = resolve_user_account_id(config, assignee_text, auth)
    if account_id:
        assignee_value = account_id

    payload = {
        "objectTypeId": str(object_type_id),
        "attributes": [
            {"objectTypeAttributeId": ID_ASIGNACION, "objectAttributeValues": [{"value": assignee_value}]},
            {"objectTypeAttributeId": ID_ESTADO, "objectAttributeValues": [{"value": "En uso"}]},
        ],
    }

    put_url = f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/{object_id}"
    post_url = f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/{object_id}/attributes"
    for method, url in [("PUT", put_url), ("POST", post_url)]:
        try:
            response = jira_request_with_retry(method, url, auth=auth, headers=headers, json_payload=payload)
            if response.status_code in (200, 201):
                return True, "Asignación actualizada"
        except Exception as exc:
            last = str(exc)
    return False, f"No se pudo actualizar asignación: {last if 'last' in locals() else 'error desconocido'}"


def update_asset_status(config: AppConfig, object_id: str, object_type_id: str, new_status: str) -> tuple[bool, str]:
    auth, headers = build_auth_headers(config)
    payload = {
        "objectTypeId": str(object_type_id),
        "attributes": [
            {"objectTypeAttributeId": ID_ESTADO, "objectAttributeValues": [{"value": new_status}]},
        ],
    }
    put_url = f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/{object_id}"
    post_url = f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/{object_id}/attributes"
    for method, url in [("PUT", put_url), ("POST", post_url)]:
        try:
            response = jira_request_with_retry(method, url, auth=auth, headers=headers, json_payload=payload)
            if response.status_code in (200, 201):
                return True, f"Estado actualizado a {new_status}"
        except Exception as exc:
            last = str(exc)
    return False, f"No se pudo actualizar estado: {last if 'last' in locals() else 'error desconocido'}"


# ── 7. MOTOR DE CHAT NL (parsers, filtros, respuestas) ───────────────────
def parse_assignment_action(prompt: str) -> tuple[str, str] | None:
    text = (prompt or "").strip()
    patterns = [
        r"(?:asign(?:a|ar|ame|ale|á))\s+(?:(?:el|la|este|esta)\s+)?(?:laptop|notebook|equipo|activo)?\s*([A-Za-z0-9._\-/]+)\s+(?:a|para)\s+(.+)$",
        r"(?:asign(?:a|ar|á))\s+serial\s+([A-Za-z0-9._\-/]+)\s+(?:a|para)\s+(.+)$",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip(), match.group(2).strip()
    return None


def parse_unassign_action(prompt: str) -> tuple[str, str] | None:
    text = (prompt or "").strip()
    nt = normalize_text(text)
    intent_words = [
        "desasign",
        "quitar usuario",
        "quita usuario",
        "sacar usuario",
        "liberar equipo",
        "libera equipo",
    ]
    if not any(w in nt for w in intent_words):
        return None
    candidates = extract_identifier_candidates(text)
    if not candidates:
        return None
    ident = candidates[-1]
    status = "Stock nuevo" if "stock nuevo" in nt else "Stock usado"
    return ident, status


def parse_status_change_action(prompt: str) -> tuple[str, str] | None:
    text = (prompt or "").strip()
    patterns = [
        r"(?:cambi(?:a|ar|á)|pas(?:a|ar|á)|pon(?:e|er|é))\s+(?:el\s+)?estado\s+(?:de\s+)?([A-Za-z0-9._\-/]+)\s+(?:a|en)\s+(.+)$",
        r"([A-Za-z0-9._\-/]+)\s+(?:a|en)\s+(stock nuevo|stock usado|en uso|asignado al edificio)$",
    ]
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.IGNORECASE)
        if not m:
            continue
        ident = m.group(1).strip()
        raw_status = normalize_text(m.group(2))
        if "stock" == raw_status:
            status = "Stock nuevo"
        else:
            status = ESTADO_NORMALIZATION.get(raw_status, m.group(2).strip().title())
            if status == "En uso":
                status = "En uso"
            elif normalize_text(status) == "stock nuevo":
                status = "Stock nuevo"
            elif normalize_text(status) == "stock usado":
                status = "Stock usado"
            elif normalize_text(status) == "asignado al edificio":
                status = "Asignado al edificio"
        return ident, status
    return None


def parse_assignee_query(prompt: str) -> str | None:
    text = (prompt or "").strip()
    patterns = [
        r"(?:que|qué)\s+tiene\s+asignad[oa]\s+(.+)$",
        r"asignad[oa]s?\s+a\s+(.+)$",
        r"(?:equipos|activos)\s+de\s+(.+)$",
    ]
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip(" ?.")
    return None


def parse_assignee_of_identifier_query(prompt: str) -> str | None:
    text = (prompt or "").strip()
    patterns = [
        r"(?:a\s+quien|a\s+quién)\s+esta\s+asignad[oa]\s*\??\s*([A-Za-z0-9._\-/]+)$",
        r"(?:quien|quién)\s+tiene\s+(?:el\s+)?(?:equipo|activo|hostname|serial)?\s*([A-Za-z0-9._\-/]+)$",
        r"(?:de\s+quien|de\s+quién)\s+es\s+(?:el\s+)?(?:equipo|activo|hostname|serial)?\s*([A-Za-z0-9._\-/]+)$",
    ]
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip(" ?.")
    return None


def looks_like_inventory_identifier(token: str) -> bool:
    raw = str(token or "").strip()
    if not raw:
        return False
    norm = normalize_lookup_key(raw)
    if not norm or norm in IDENTIFIER_STOPWORDS:
        return False
    if not IDENTIFIER_REGEX.fullmatch(raw):
        return False
    has_digit = bool(re.search(r"\d", raw))
    has_separator = bool(re.search(r"[._-]", raw))
    is_upperish = raw.upper() == raw and any(ch.isalpha() for ch in raw)
    if has_digit or has_separator or is_upperish:
        return True
    return raw.isalpha() and 4 <= len(raw) <= 15


def extract_identifier_candidates(prompt: str) -> list[str]:
    text = prompt or ""
    short_prompt = len(normalize_text(prompt).split()) <= 4
    matches = IDENTIFIER_REGEX.findall(text)
    out: list[str] = []
    for token in matches:
        if "@" in token:
            continue
        if token.isalpha() and not short_prompt:
            continue
        if not looks_like_inventory_identifier(token):
            continue
        out.append(token.strip())
    # dedupe conservando orden
    seen = set()
    unique = []
    for t in out:
        key = normalize_text(t)
        if key in seen:
            continue
        seen.add(key)
        unique.append(t)
    return unique


def build_aql_from_prompt(prompt: str) -> tuple[str, list[str]]:
    notes: list[str] = []
    clauses: list[str] = []
    t = normalize_text(prompt)

    # Si es comando de acción, no forzar ruta NL->AQL.
    if parse_assignment_action(prompt) or parse_unassign_action(prompt) or parse_status_change_action(prompt) or parse_bulk_location_action(prompt):
        return "", notes

    category = detect_category_from_prompt(prompt)
    if category and category in CATEGORY_TO_TYPE_ID:
        clauses.append(f"objectTypeId = {CATEGORY_TO_TYPE_ID[category]}")
        notes.append(f"tipo={category}")

    if any(token in t for token in ["libre", "libres", "disponible", "disponibles"]):
        clauses.append('"Estado del activo" IN ("Stock nuevo", "Stock usado")')
        notes.append("estado=disponible(logico)")
    else:
        status = detect_status_from_prompt(prompt)
        if status:
            clauses.append(f'"Estado del activo" = "{status}"')
            notes.append(f"estado={status}")

    serial_candidate = extract_serial_candidate(prompt)
    if serial_candidate and (any(x in t for x in ["serial", "sn-", "equipo", "hostname", "host"]) or len(t.split()) <= 5):
        ident = serial_candidate.replace('"', '\\"')
        clauses.append(f'("Serial Number" = "{ident}" OR "Hostname" = "{ident}")')
        notes.append(f"id~{ident}")

    host_match = re.search(r"(?:hostname|host)\s+([A-Za-z0-9._\-/]+)", prompt, flags=re.IGNORECASE)
    if host_match:
        host = host_match.group(1).strip()
        clauses.append(f'"Hostname" = "{host}"')
        notes.append(f"hostname={host}")

    model_match = re.search(r"(?:modelo|model)\s+(.+)$", prompt, flags=re.IGNORECASE)
    if model_match:
        model = model_match.group(1).strip().strip("\"'")
        clauses.append(f'"Nombre del modelo" LIKE "{model}"')
        notes.append(f"modelo~{model}")
    elif any(x in t for x in ["macbook", "thinkpad", "elitebook", "probook", "latitude", "ideapad"]):
        guess = re.search(r"(macbook[^,.;]*|thinkpad[^,.;]*|elitebook[^,.;]*|probook[^,.;]*|latitude[^,.;]*|ideapad[^,.;]*)", t)
        if guess:
            model = guess.group(1).strip()
            clauses.append(f'"Nombre del modelo" LIKE "{model}"')
            notes.append(f"modelo~{model}")

    country = detect_country_from_prompt(prompt)
    if country:
        clauses.append(f'"Pais" LIKE "{country}"')
        notes.append(f"pais={country}")

    company = detect_company_from_prompt(prompt)
    if company:
        clauses.append(f'"Compañía" LIKE "{company}"')
        notes.append(f"compania={company}")

    entity = detect_entity_from_prompt(prompt)
    if entity:
        clauses.append(f'"Entidad del activo" LIKE "{entity}"')
        notes.append(f"entidad={entity}")

    today = datetime.now().date()
    if "comprados este año" in t:
        start = f"{today.year}-01-01"
        clauses.append(f'"Fecha compra" >= "{start}"')
        notes.append(f"fecha_compra>={start}")
    if "comprados el año pasado" in t:
        start = f"{today.year - 1}-01-01"
        end = f"{today.year - 1}-12-31"
        clauses.append(f'"Fecha compra" >= "{start}" AND "Fecha compra" <= "{end}"')
        notes.append(f"fecha_compra={today.year-1}")
    if "garantia vence en 30 dias" in normalize_lookup_key(prompt) or "garantía vence en 30 días" in normalize_text(prompt):
        end = (today + timedelta(days=30)).strftime("%Y-%m-%d")
        clauses.append(f'"Fecha garantia" <= "{end}"')
        notes.append("garantia<=30d")
    m_before = re.search(r"comprad[oa]s?\s+antes\s+de\s+(\d{4})", t)
    if m_before:
        year = int(m_before.group(1))
        clauses.append(f'"Fecha compra" < "{year}-01-01"')
        notes.append(f"fecha_compra<{year}")
    m_month = re.search(
        r"comprad[oa]s?\s+en\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre)\s+(\d{4})",
        t,
    )
    if m_month:
        month_names = {
            "enero": 1,
            "febrero": 2,
            "marzo": 3,
            "abril": 4,
            "mayo": 5,
            "junio": 6,
            "julio": 7,
            "agosto": 8,
            "septiembre": 9,
            "setiembre": 9,
            "octubre": 10,
            "noviembre": 11,
            "diciembre": 12,
        }
        month_num = month_names[m_month.group(1)]
        year = int(m_month.group(2))
        last_day = monthrange(year, month_num)[1]
        start = f"{year}-{month_num:02d}-01"
        end = f"{year}-{month_num:02d}-{last_day:02d}"
        clauses.append(f'"Fecha compra" >= "{start}" AND "Fecha compra" <= "{end}"')
        notes.append(f"fecha_compra={m_month.group(1)} {year}")
    if "garantia vencida" in normalize_lookup_key(prompt) or "garantía vencida" in t:
        today_s = today.strftime("%Y-%m-%d")
        clauses.append(f'"Fecha garantia" < "{today_s}"')
        notes.append("garantia_vencida")

    person_query = parse_assignee_query(prompt)
    if person_query:
        clauses.append(f'"Usuario asignado" LIKE "{person_query}"')
        notes.append(f"asignado~{person_query}")

    if not clauses:
        return "", notes
    return " AND ".join(clauses), notes


def summarize_for_ai(assets: list[dict[str, Any]], max_rows: int = 14) -> str:
    lines = ["name|type|status|model|country|assignee|warranty"]
    for a in assets[:max_rows]:
        lines.append(
            f"{a.get('name','')}|{a.get('object_type','')}|{a.get('status','')}|{a.get('model','')}|{a.get('country','')}|{a.get('assigned_to','')}|{a.get('warranty_date','')}"
        )
    return "\n".join(lines)


def push_openai_history(user_text: str, assistant_text: str) -> None:
    """Guarda memoria deslizante de 6 pares user/assistant para contexto LLM."""
    history = st.session_state.setdefault("openai_history", [])
    history.append({"role": "user", "content": user_text})
    history.append({"role": "assistant", "content": assistant_text})
    st.session_state["openai_history"] = history[-12:]


def ai_compact_answer(config: AppConfig, prompt: str, assets: list[dict[str, Any]], notes: list[str], prefiltered: bool = False) -> str:
    if OpenAI is None or not config.openai_api_key:
        return local_chat_answer(assets, prompt, prefiltered=prefiltered)
    try:
        client = OpenAI(api_key=config.openai_api_key)
        context = summarize_for_ai(assets)
        sys = "Analista IT de inventario Jira. Responde corto en español y accionable. Si faltan datos, dilo."
        usr = f"notas={'; '.join(notes[:6])}\nrows={len(assets)}\n{context}\n\npregunta:{prompt}"
        messages: list[dict[str, str]] = [{"role": "system", "content": sys}]
        for item in st.session_state.get("openai_history", [])[-12:]:
            if item.get("role") in {"user", "assistant"} and item.get("content"):
                messages.append({"role": item["role"], "content": str(item["content"])})
        messages.append({"role": "user", "content": usr})
        resp = client.chat.completions.create(
            model=config.openai_model,
            temperature=0,
            max_tokens=220,
            messages=messages,
        )
        return resp.choices[0].message.content or local_chat_answer(assets, prompt, prefiltered=prefiltered)
    except Exception:
        return local_chat_answer(assets, prompt, prefiltered=prefiltered)


def rovo_resolve_identity_context(config: AppConfig, assets: list[dict[str, Any]], user_query: str) -> dict[str, Any]:
    auth, _ = build_auth_headers(config)
    account_id = resolve_user_account_id(config, user_query, auth)
    linked_assets = filter_assets_by_assignee(assets, user_query)
    return {
        "source": "rovo-conceptual" if config.rovo_enabled else "local-fallback",
        "input": user_query,
        "accountId": account_id or "",
        "matched_assets": [
            {
                "jira": a.get("jira_key"),
                "hostname": a.get("hostname"),
                "serial": a.get("serial_number"),
                "status": a.get("status"),
            }
            for a in linked_assets[:10]
        ],
        "count": len(linked_assets),
    }


def compact_assets_rows(assets: list[dict[str, Any]], limit: int = 15) -> list[dict[str, Any]]:
    rows = []
    for a in assets[:limit]:
        rows.append(
            {
                "jira": a.get("jira_key") or a.get("name"),
                "hostname": a.get("hostname"),
                "serial": a.get("serial_number"),
                "status": a.get("status"),
                "assigned_to": a.get("assigned_to"),
                "country": a.get("country"),
                "model": a.get("model"),
            }
        )
    return rows


# ── 8. SUPER AGENTE (OpenAI function calling) ─────────────────────────────
def run_super_agente(
    config: AppConfig,
    prompt: str,
    assets: list[dict[str, Any]],
) -> tuple[bool, str, bool]:
    if OpenAI is None or not config.openai_api_key:
        return False, "", False

    client = OpenAI(api_key=config.openai_api_key)
    tools = [
        {
            "type": "function",
            "function": {
                "name": "query_assets_aql",
                "description": "Consulta Jira Assets por AQL preciso. No usar lista completa local si puede evitarse.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "aql": {"type": "string"},
                        "limit": {"type": "integer", "minimum": 1, "maximum": 50},
                    },
                    "required": ["aql"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "rovo_resolve_identity",
                "description": "Valida identidad/contexto de usuario (Rovo conceptual) antes de accionar.",
                "parameters": {
                    "type": "object",
                    "properties": {"user_query": {"type": "string"}},
                    "required": ["user_query"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "assign_asset_action",
                "description": "Asigna un activo a un usuario y lo pasa a En uso.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "identifier": {"type": "string"},
                        "assignee": {"type": "string"},
                    },
                    "required": ["identifier", "assignee"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "unassign_asset_action",
                "description": "Desasigna usuario de activo y lo lleva a stock.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "identifier": {"type": "string"},
                        "target_status": {"type": "string"},
                    },
                    "required": ["identifier"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "update_status_action",
                "description": "Actualiza estado de activo.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "identifier": {"type": "string"},
                        "new_status": {"type": "string"},
                    },
                    "required": ["identifier", "new_status"],
                },
            },
        },
    ]
    sys = (
        "Sos un super agente de inventario IT. "
        "Primero validá identidad con rovo_resolve_identity cuando el prompt mencione persona/email. "
        "Para responder datos usa query_assets_aql con AQL preciso. "
        "No pidas lista completa de activos. Sé breve, claro y en español."
    )
    user_msg = (
        f"schema={json.dumps(SCHEMA_MINI, ensure_ascii=False)}\n"
        f"prompt={prompt}\n"
        f"context_count={len(assets)}"
    )
    try:
        history_messages: list[dict[str, str]] = []
        for item in st.session_state.get("openai_history", [])[-12:]:
            if item.get("role") in {"user", "assistant"} and item.get("content"):
                history_messages.append({"role": item["role"], "content": str(item["content"])})
        first = client.chat.completions.create(
            model=config.openai_model,
            temperature=0,
            max_tokens=280,
            tool_choice="auto",
            tools=tools,
            messages=[{"role": "system", "content": sys}, *history_messages, {"role": "user", "content": user_msg}],
        )
        msg = first.choices[0].message
        tool_calls = msg.tool_calls or []
        if not tool_calls:
            content = msg.content or ""
            if content.strip():
                return True, content, False
            return False, "", False

        action_success = False
        action_outputs: list[dict[str, Any]] = []
        for call in tool_calls:
            fn = call.function.name
            try:
                args = json.loads(call.function.arguments or "{}")
            except Exception:
                args = {}
            result: Any
            if fn == "query_assets_aql":
                aql = str(args.get("aql", "")).strip()
                limit = int(args.get("limit", 15) or 15)
                queried = fetch_assets(config, aql) if aql else []
                result = {
                    "aql": aql,
                    "count": len(queried),
                    "rows": compact_assets_rows(queried, limit=max(1, min(limit, 50))),
                }
            elif fn == "rovo_resolve_identity":
                result = rovo_resolve_identity_context(config, assets, str(args.get("user_query", "")).strip())
            elif fn == "assign_asset_action":
                ok, text = assign_asset(config, assets, str(args.get("identifier", "")).strip(), str(args.get("assignee", "")).strip())
                result = {"ok": ok, "message": text}
                action_success = action_success or ok
            elif fn == "unassign_asset_action":
                ok, text = unassign_asset(
                    config,
                    assets,
                    str(args.get("identifier", "")).strip(),
                    str(args.get("target_status", "Stock usado") or "Stock usado"),
                )
                result = {"ok": ok, "message": text}
                action_success = action_success or ok
            elif fn == "update_status_action":
                ok, text = update_status(config, assets, str(args.get("identifier", "")).strip(), str(args.get("new_status", "")).strip())
                result = {"ok": ok, "message": text}
                action_success = action_success or ok
            else:
                result = {"error": f"tool no soportada: {fn}"}
            action_outputs.append(
                {
                    "tool_call_id": call.id,
                    "role": "tool",
                    "name": fn,
                    "content": json.dumps(result, ensure_ascii=False),
                }
            )

        tool_calls_payload = [
            {
                "id": tc.id,
                "type": "function",
                "function": {
                    "name": tc.function.name,
                    "arguments": tc.function.arguments,
                },
            }
            for tc in tool_calls
        ]
        final_messages: list[dict[str, Any]] = [
            {"role": "system", "content": sys},
            {"role": "user", "content": user_msg},
            {
                "role": "assistant",
                "content": msg.content or "",
                "tool_calls": tool_calls_payload,
            },
        ]
        final_messages.extend(action_outputs)
        second = client.chat.completions.create(
            model=config.openai_model,
            temperature=0,
            max_tokens=320,
            messages=final_messages,
        )
        answer = (second.choices[0].message.content or "").strip()
        return True, answer or "Operación procesada.", action_success
    except Exception:
        return False, "", False


def extract_attr_text(attribute: dict[str, Any]) -> str:
    values = []
    for item in attribute.get("objectAttributeValues", []):
        display = item.get("displayValue")
        raw = item.get("value")
        if display not in (None, ""):
            values.append(str(display).strip())
            continue
        if isinstance(raw, dict):
            candidate = raw.get("label") or raw.get("name") or raw.get("value")
            if candidate not in (None, ""):
                values.append(str(candidate).strip())
                continue
        if raw not in (None, ""):
            values.append(str(raw).strip())
    unique = []
    seen = set()
    for value in values:
        if value and value not in seen:
            seen.add(value)
            unique.append(value)
    return " | ".join(unique)


def get_attr_value(attrs_by_id: dict[str, str], attrs_by_name: dict[str, str], preferred_id: str, alias_names: list[str]) -> str:
    value_by_id = str(attrs_by_id.get(preferred_id, "")).strip()
    if value_by_id:
        return value_by_id

    normalized_name_map = {normalize_lookup_key(name): value for name, value in attrs_by_name.items()}
    for alias in alias_names:
        alias_norm = normalize_lookup_key(alias)
        if alias_norm in normalized_name_map and str(normalized_name_map[alias_norm]).strip():
            return str(normalized_name_map[alias_norm]).strip()

    for alias in alias_names:
        alias_norm = normalize_lookup_key(alias)
        for attr_name, attr_value in attrs_by_name.items():
            if alias_norm and alias_norm in normalize_lookup_key(attr_name) and str(attr_value).strip():
                return str(attr_value).strip()
    return ""


def clean_asset_object(asset: dict[str, Any]) -> AssetRecord:
    """Normaliza un objeto Jira Assets a un registro tipado."""
    attrs_by_id: dict[str, str] = {}
    attrs_by_name: dict[str, str] = {}
    attr_name_to_id: dict[str, str] = {}

    for attribute in asset.get("attributes", []):
        meta = attribute.get("objectTypeAttribute", {})
        attr_id = str(meta.get("id", "")).strip()
        attr_name = str(meta.get("name", "")).strip()
        attr_value = extract_attr_text(attribute)
        if not attr_value:
            continue
        if attr_id:
            attrs_by_id[attr_id] = attr_value
        if attr_name:
            attrs_by_name[attr_name] = attr_value
            attr_name_to_id[normalize_lookup_key(attr_name)] = attr_id

    purchase_date = get_attr_value(attrs_by_id, attrs_by_name, ID_FECHA_COMPRA, ["Fecha compra", "Fecha de compra", "Purchase Date"])
    warranty_date = get_attr_value(attrs_by_id, attrs_by_name, ID_FECHA_GARANTIA, ["Fecha garantia", "Garantia", "Warranty", "Warranty End"])
    category_raw = get_attr_value(attrs_by_id, attrs_by_name, ID_CATEGORIA, ["Categoria", "Category"])
    status_raw = get_attr_value(attrs_by_id, attrs_by_name, ID_ESTADO, ["Estado", "Estado del activo", "Status"])
    country_raw = get_attr_value(attrs_by_id, attrs_by_name, ID_PAIS, ["Pais", "País", "Country"])
    assigned_raw = get_attr_value(
        attrs_by_id,
        attrs_by_name,
        ID_ASIGNACION,
        ["Asignacion", "Asignación", "Assigned To", "Asignado a", "Usuario asignado", "User assigned"],
    )
    status_final, assigned_final = enforce_assignment_status_rules(canonical_status(status_raw), assigned_raw)

    return AssetRecord(
        object_id=str(asset.get("id") or asset.get("objectId") or ""),
        object_type_id=str(asset.get("objectTypeId") or (asset.get("objectType") or {}).get("id") or ""),
        name=asset.get("label") or attrs_by_id.get(ID_NAME, "") or asset.get("name", ""),
        object_type=(asset.get("objectType") or {}).get("name", ""),
        created=asset.get("created", ""),
        jira_key=str(asset.get("objectKey") or ""),
        category=canonical_category(category_raw),
        status=status_final,
        country=canonical_country(country_raw),
        company=get_attr_value(attrs_by_id, attrs_by_name, ID_COMPANIA, ["Compañia", "Compania", "Company"]),
        entity=get_attr_value(attrs_by_id, attrs_by_name, ID_ENTIDAD, ["Entidad del activo", "Entidad", "Entity"]),
        assigned_to=assigned_final,
        serial_number=get_attr_value(attrs_by_id, attrs_by_name, ID_SERIAL, ["Serial Number", "Serial"]),
        hostname=get_attr_value(attrs_by_id, attrs_by_name, ID_HOSTNAME, ["Hostname"]),
        model=get_attr_value(attrs_by_id, attrs_by_name, ID_MODELO, ["Modelo", "Model"]),
        provider=get_attr_value(attrs_by_id, attrs_by_name, ID_PROVEEDOR, ["Proveedor", "Provider"]),
        purchase_date=purchase_date,
        warranty_date=warranty_date,
        purchase_price=get_attr_value(attrs_by_id, attrs_by_name, ID_COSTO, ["Costo", "Cost"]),
        attrs_by_name=attrs_by_name,
        attrs_by_id=attrs_by_id,
        attr_name_to_id=attr_name_to_id,
    )


def enrich_assets_with_object_details(
    config: AppConfig,
    auth: BasicAuth,
    headers: dict[str, str],
    records: list[dict[str, Any]],
    error_sink: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    if not records:
        return records

    urls = [
        f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object",
        f"{config.site}/rest/servicedeskapi/assets/workspace/{config.workspace_id}/v1/object",
    ]
    results_by_index: dict[int, dict[str, Any]] = {}
    grouped_to_fetch: dict[str, list[tuple[int, dict[str, Any]]]] = {}

    for idx, record in enumerate(records):
        if bool(record.get("attrs_by_name") or record.get("attrs_by_id")):
            results_by_index[idx] = record
            continue
        object_id = str(record.get("object_id") or "").strip()
        if not object_id:
            results_by_index[idx] = record
            continue
        grouped_to_fetch.setdefault(object_id, []).append((idx, record))

    def fetch_detail_for_object(object_id: str, seed_record: dict[str, Any]) -> tuple[str, dict[str, Any], list[dict[str, Any]]]:
        local_errors: list[dict[str, Any]] = []
        detail_body = None
        for base_url in urls:
            try:
                response = jira_request_with_retry(
                    "GET",
                    f"{base_url}/{object_id}",
                    auth=auth,
                    headers=headers,
                    params={"includeAttributes": "true"},
                    error_sink=local_errors,
                )
                detail_body = response.json()
                break
            except Exception:
                continue

        if isinstance(detail_body, dict) and detail_body.get("attributes"):
            detailed = clean_asset_object(detail_body).to_dict()
            if not detailed.get("jira_key"):
                detailed["jira_key"] = seed_record.get("jira_key", "")
            return object_id, detailed, local_errors
        return object_id, seed_record, local_errors

    if grouped_to_fetch:
        max_workers = max(1, min(len(grouped_to_fetch), 6))
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_map = {
                executor.submit(fetch_detail_for_object, object_id, grouped[0][1]): object_id
                for object_id, grouped in grouped_to_fetch.items()
            }
            for future in concurrent.futures.as_completed(future_map):
                object_id = future_map[future]
                try:
                    _, detailed_record, local_errors = future.result()
                except Exception:
                    detailed_record = grouped_to_fetch[object_id][0][1]
                    local_errors = []
                if error_sink is not None and local_errors:
                    error_sink.extend(local_errors)
                for idx, original in grouped_to_fetch[object_id]:
                    result = dict(detailed_record)
                    if not result.get("jira_key"):
                        result["jira_key"] = original.get("jira_key", "")
                    results_by_index[idx] = result

    return [results_by_index.get(idx, records[idx]) for idx in range(len(records))]


def apply_fetch_metadata(metadata: dict[str, int]) -> None:
    """Vuelca diagnósticos de carga al session_state en el hilo principal."""
    for key, value in metadata.items():
        st.session_state[key] = value


def fetch_assets_backend(
    config: AppConfig,
    aql_query: str = "",
    *,
    discovered_type_ids: list[str] | None = None,
    all_schema_type_ids: list[str] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, int], list[dict[str, Any]]]:
    if not config.workspace_id:
        raise RuntimeError("Falta ASSETS_WORKSPACE_ID/JIRA_WORKSPACE_ID en el entorno.")

    debug_log(f"fetch_assets:start aql={aql_query!r}")
    auth, headers = build_auth_headers(config)
    error_events: list[dict[str, Any]] = []
    metadata: dict[str, int] = {
        "last_base_records_count": 0,
        "last_segmented_records_count": 0,
        "last_bruteforce_records_count": 0,
        "last_type_scan_checked": 0,
        "last_type_scan_hits": 0,
    }
    active_scope_type_ids = get_active_hardware_type_ids(discovered_type_ids)
    active_scope_type_id_set = {type_id for type_id in active_scope_type_ids if type_id}
    scan_scope_type_ids = get_schema_scan_type_ids(all_schema_type_ids, active_scope_type_ids)
    if ASSETS_SCOPE_MODE in {"hardware", "strict_hardware"}:
        scan_type_ids = list(active_scope_type_ids)
    else:
        scan_type_ids = list(scan_scope_type_ids)

    def fetch_by_full_query(full_ql_query: str) -> tuple[list[dict[str, Any]], Exception | None]:
        try:
            records = paginate_aql_sync(config, auth, headers, full_ql_query)
            return records, None
        except Exception as exc:
            return [], exc

    base_full_query = build_primary_aql(aql_query, type_ids=active_scope_type_ids)
    base_records, last_error = fetch_by_full_query(base_full_query)
    metadata["last_base_records_count"] = len(base_records)
    debug_log(f"fetch_assets:base_done count={len(base_records)} error={bool(last_error)}")

    base_is_sufficient = bool(base_records)
    recovery_allowed = (
        SEGMENTED_FETCH_ENABLED
        and not str(aql_query or "").strip()
        and len(base_records) < max(1, int(FORCE_FETCH_MIN_ASSETS))
    )
    if base_is_sufficient or not recovery_allowed:
        debug_log(
            "fetch_assets:using_base "
            f"count={len(base_records)} recovery_allowed={recovery_allowed}"
        )
        metadata["last_segmented_records_count"] = len(base_records)
        final_records = base_records
    else:
        debug_log(f"fetch_assets:recovery_start base_count={len(base_records)}")
        merged: dict[str, dict[str, Any]] = {}
        for record in base_records:
            key = str(record.get("object_id") or record.get("jira_key") or record.get("name") or id(record))
            merged[key] = record

        try:
            segmented_results: list[list[dict[str, Any]]] = []
            with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, min(len(scan_type_ids), 8))) as executor:
                futures = [
                    executor.submit(fetch_type_sync, config, auth, headers, type_id, aql_query, active_scope_type_ids)
                    for type_id in scan_type_ids
                ]
                for future in concurrent.futures.as_completed(futures):
                    try:
                        segmented_results.append(future.result())
                    except Exception as exc:
                        last_error = exc

            for segmented_records in segmented_results:
                for record in segmented_records:
                    key = str(record.get("object_id") or record.get("jira_key") or record.get("name") or id(record))
                    merged[key] = record
        except Exception as exc:
            last_error = exc

        final_records = list(merged.values())
        metadata["last_segmented_records_count"] = len(merged)

        should_force_bruteforce = (
            FORCE_FETCH_ENABLED
            and not str(aql_query or "").strip()
            and len(merged) < max(1, int(FORCE_FETCH_MIN_ASSETS))
        )
        if should_force_bruteforce:
            brute_results: list[list[dict[str, Any]]] = []

            def fetch_bruteforce_worker(type_id: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
                worker_errors: list[dict[str, Any]] = []
                rows = fetch_objects_by_type_bruteforce(config, auth, headers, type_id, error_sink=worker_errors)
                return rows, worker_errors

            try:
                with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, min(len(scan_type_ids), 8))) as executor:
                    futures = [executor.submit(fetch_bruteforce_worker, type_id) for type_id in scan_type_ids]
                    for future in concurrent.futures.as_completed(futures):
                        try:
                            rows, worker_errors = future.result()
                            error_events.extend(worker_errors)
                            brute_results.append(rows)
                        except Exception as exc:
                            last_error = exc
            except Exception as exc:
                last_error = exc

            for rows in brute_results:
                for record in rows:
                    key = str(record.get("object_id") or record.get("jira_key") or record.get("name") or id(record))
                    merged[key] = record

            metadata["last_bruteforce_records_count"] = max(0, len(merged) - int(metadata.get("last_segmented_records_count", 0)))
            final_records = list(merged.values())
            metadata["last_segmented_records_count"] = len(merged)

            if TYPE_SCAN_ENABLED and len(merged) < max(1, int(FORCE_FETCH_MIN_ASSETS)):
                scan_start = min(TYPE_SCAN_START, TYPE_SCAN_END)
                scan_end = max(TYPE_SCAN_START, TYPE_SCAN_END)
                scan_ids = [str(x) for x in range(scan_start, scan_end + 1)]
                skip = {str(x).strip() for x in scan_type_ids if str(x).strip()}
                scan_ids = [type_id for type_id in scan_ids if type_id not in skip]
                metadata["last_type_scan_checked"] = len(scan_ids)
                scan_hits = 0
                try:
                    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
                        future_map = {
                            executor.submit(fetch_bruteforce_worker, type_id): type_id
                            for type_id in scan_ids
                        }
                        for future in concurrent.futures.as_completed(future_map):
                            try:
                                rows, worker_errors = future.result()
                                error_events.extend(worker_errors)
                            except Exception:
                                continue
                            if not rows:
                                continue
                            scan_hits += 1
                            for record in rows:
                                key = str(record.get("object_id") or record.get("jira_key") or record.get("name") or id(record))
                                merged[key] = record
                except Exception as exc:
                    last_error = exc
                metadata["last_type_scan_hits"] = scan_hits
                metadata["last_bruteforce_records_count"] = max(0, len(merged) - int(metadata.get("last_segmented_records_count", 0)))
                final_records = list(merged.values())
                metadata["last_segmented_records_count"] = len(merged)

    if not final_records and last_error is not None:
        raise RuntimeError(f"No se pudo consultar Jira Assets: {last_error}")

    if ASSETS_SCOPE_MODE == "strict_hardware":
        final_records = [
            record for record in final_records
            if str(record.get("object_type_id") or "").strip() in active_scope_type_id_set
        ]

    debug_log(f"fetch_assets:before_enrich count={len(final_records)}")
    enriched = enrich_assets_with_object_details(config, auth, headers, final_records, error_sink=error_events)
    debug_log(f"fetch_assets:after_enrich count={len(enriched)}")
    debug_log(
        "fetch_assets:end "
        f"base={len(base_records)} final={len(final_records)} enriched={len(enriched)} "
        f"segmented={metadata['last_segmented_records_count']} "
        f"bruteforce={metadata['last_bruteforce_records_count']}"
    )
    return enriched, metadata, error_events


def fetch_assets(config: AppConfig, aql_query: str = "") -> list[dict[str, Any]]:
    assets, metadata, error_events = fetch_assets_backend(
        config,
        aql_query,
        discovered_type_ids=st.session_state.get("discovered_type_ids") or KNOWN_OBJECT_TYPE_IDS,
        all_schema_type_ids=st.session_state.get("all_schema_type_ids") or [],
    )
    apply_fetch_metadata(metadata)
    append_error_events(error_events)
    return assets


def compute_cache_hash(config: AppConfig, aql_query: str) -> str:
    """Calcula hash de caché por query, workspace y usuario Jira."""
    raw = f"{aql_query}|{config.workspace_id}|{config.jira_email}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def cached_fetch_assets(config: AppConfig, aql_query: str, ttl_minutes: int, *, force_live: bool = False) -> list[dict[str, Any]]:
    """Resuelve caché de assets con invalidación por hash y vencimiento."""
    discovered_snapshot = normalize_type_id_list(st.session_state.get("discovered_type_ids") or KNOWN_OBJECT_TYPE_IDS)
    all_schema_snapshot = normalize_type_id_list(st.session_state.get("all_schema_type_ids") or [])
    cache_hash = compute_cache_hash(config, aql_query)
    now = datetime.now()
    expiry = st.session_state.get("cache_expiry")
    current_hash = st.session_state.get("cache_hash")
    if (
        not force_live
        and st.session_state.get("assets")
        and current_hash == cache_hash
        and isinstance(expiry, datetime)
        and now < expiry
    ):
        return st.session_state["assets"]
    process_now = time.time()
    with PROCESS_FETCH_LOCK:
        cached = PROCESS_FETCH_RESULTS.get(cache_hash)
        if (
            not force_live
            and cached
            and (process_now - cached[0]) < min(int(ttl_minutes * 60), PROCESS_FETCH_CACHE_TTL_SECONDS)
        ):
            _, cached_assets, cached_metadata, cached_errors = cached
            debug_log(f"cached_fetch_assets:process_cache_hit hash={cache_hash[:8]}")
            apply_fetch_metadata(cached_metadata)
            append_error_events(copy.deepcopy(cached_errors))
            assets = copy.deepcopy(cached_assets)
            st.session_state["cache_hash"] = cache_hash
            st.session_state["cache_expiry"] = now + timedelta(minutes=ttl_minutes)
            return assets

    if not force_live:
        snapshot_assets, snapshot_metadata, snapshot_saved_at = load_assets_snapshot()
        if snapshot_assets:
            debug_log(f"cached_fetch_assets:snapshot_hit count={len(snapshot_assets)} hash={cache_hash[:8]}")
            apply_fetch_metadata({k: int(v or 0) for k, v in snapshot_metadata.items() if isinstance(v, (int, str))})
            st.session_state["cache_hash"] = cache_hash
            st.session_state["cache_expiry"] = now + timedelta(minutes=ttl_minutes)
            if isinstance(snapshot_saved_at, datetime):
                st.session_state["last_sync"] = snapshot_saved_at
            return copy.deepcopy(snapshot_assets)

    with PROCESS_FETCH_LOCK:
        future = PROCESS_FETCH_JOBS.get(cache_hash)
        if future is None:
            config_snapshot = config_to_cache_key(config)
            debug_log(f"cached_fetch_assets:submit hash={cache_hash[:8]}")
            future = PROCESS_FETCH_EXECUTOR.submit(
                fetch_assets_backend,
                config_from_cache_key(config_snapshot),
                aql_query,
                discovered_type_ids=discovered_snapshot,
                all_schema_type_ids=all_schema_snapshot,
            )
            PROCESS_FETCH_JOBS[cache_hash] = future
        else:
            debug_log(f"cached_fetch_assets:reuse_inflight hash={cache_hash[:8]}")

    try:
        assets, metadata, error_events = future.result()
    except Exception:
        with PROCESS_FETCH_LOCK:
            current_future = PROCESS_FETCH_JOBS.get(cache_hash)
            if current_future is future:
                PROCESS_FETCH_JOBS.pop(cache_hash, None)
        raise

    with PROCESS_FETCH_LOCK:
        PROCESS_FETCH_RESULTS[cache_hash] = (
            time.time(),
            copy.deepcopy(assets),
            dict(metadata),
            copy.deepcopy(error_events),
        )
        current_future = PROCESS_FETCH_JOBS.get(cache_hash)
        if current_future is future:
            PROCESS_FETCH_JOBS.pop(cache_hash, None)

    apply_fetch_metadata(metadata)
    append_error_events(error_events)
    if assets:
        save_assets_snapshot(assets, metadata)
    st.session_state["cache_hash"] = cache_hash
    st.session_state["cache_expiry"] = now + timedelta(minutes=ttl_minutes)
    return assets


def detect_category_from_prompt(prompt: str) -> str | None:
    text = normalize_lookup_key(prompt)
    tokens = lookup_tokens(prompt)
    if any(
        matches_lookup_keyword(text, phrase, lookup_token_set=tokens)
        for phrase in [
            "a quien",
            "a quién",
            "quien tiene",
            "quién tiene",
            "de quien",
            "de quién",
            "que tiene asignado",
            "qué tiene asignado",
        ]
    ):
        return None
    for alias, canonical in sorted(CATEGORY_ALIAS_TO_CANONICAL.items(), key=lambda item: len(normalize_lookup_key(item[0])), reverse=True):
        if normalize_lookup_key(alias) in GENERIC_CATEGORY_PROMPT_ALIASES:
            continue
        if matches_lookup_keyword(text, alias, lookup_token_set=tokens):
            return canonical
    return None


def detect_status_from_prompt(prompt: str) -> str | None:
    low = normalize_text(prompt)
    # "a quién está asignado" pregunta por dueño, no por estado.
    if any(
        phrase in low
        for phrase in [
            "a quien esta asignado",
            "a quién está asignado",
            "quien esta asignado",
            "quién está asignado",
            "quien tiene",
            "quién tiene",
            "que tiene asignado",
            "qué tiene asignado",
        ]
    ):
        return None
    text = normalize_text(prompt)
    for alias, canonical in ESTADO_NORMALIZATION.items():
        if alias in text:
            return canonical
    return None


def detect_country_from_prompt(prompt: str) -> str | None:
    text = normalize_lookup_key(prompt)
    tokens = lookup_tokens(prompt)
    has_email = "@" in (prompt or "")
    explicit_company = detect_company_from_prompt(prompt)
    for country, keywords in PAIS_KEYWORDS.items():
        for keyword in sorted(keywords, key=lambda value: len(normalize_lookup_key(value)), reverse=True):
            normalized_keyword = normalize_lookup_key(keyword)
            if len(normalized_keyword) <= 3:
                if has_email or explicit_company:
                    continue
            if matches_lookup_keyword(text, keyword, lookup_token_set=tokens):
                return country
    return None


def detect_company_from_prompt(prompt: str) -> str | None:
    text = normalize_lookup_key(prompt)
    tokens = lookup_tokens(prompt)
    for canonical, keywords in COMPANIA_KEYWORDS.items():
        for keyword in sorted(keywords, key=lambda value: len(normalize_lookup_key(value)), reverse=True):
            if matches_lookup_keyword(text, keyword, lookup_token_set=tokens):
                return canonical
    return None


def detect_entity_from_prompt(prompt: str) -> str | None:
    m = re.search(r"(?:entidad)\s+([A-Za-z0-9 ._\-/]+)", prompt or "", flags=re.IGNORECASE)
    if m:
        return m.group(1).strip(" .?")
    return None


def find_asset_by_serial(assets: list[dict[str, Any]], serial: str) -> dict[str, Any] | None:
    target = normalize_text(serial)
    for asset in assets:
        if normalize_text(asset.get("serial_number", "")) == target:
            return asset
    return None


def find_asset_by_hostname(assets: list[dict[str, Any]], hostname: str) -> dict[str, Any] | None:
    target = normalize_text(hostname)
    for asset in assets:
        if normalize_text(asset.get("hostname", "")) == target:
            return asset
    return None


def find_asset_by_identifier(assets: list[dict[str, Any]], identifier: str) -> dict[str, Any] | None:
    ident = normalize_text(identifier)
    if not ident:
        return None
    by_serial = find_asset_by_serial(assets, identifier)
    if by_serial:
        return by_serial
    by_host = find_asset_by_hostname(assets, identifier)
    if by_host:
        return by_host
    for asset in assets:
        if ident == normalize_text(asset.get("jira_key", "")):
            return asset
    for asset in assets:
        hostname_norm = normalize_text(asset.get("hostname", ""))
        if ident and ident in hostname_norm:
            return asset
    for asset in assets:
        if ident and ident in normalize_text(asset.get("name", "")):
            return asset
    return None


def extract_serial_candidate(prompt: str) -> str | None:
    text = (prompt or "").strip()
    # Patrón típico inventario: mezcla de letras/números con largo >= 6
    candidates = re.findall(r"[A-Za-z0-9][A-Za-z0-9._\-/]{5,}", text)
    if not candidates:
        return None
    # Prioriza tokens que tengan letras y números
    rich = [c for c in candidates if re.search(r"[A-Za-z]", c) and re.search(r"\d", c)]
    if rich:
        return rich[0].strip()
    return None


def filter_assets_by_assignee(assets: list[dict[str, Any]], person_query: str) -> list[dict[str, Any]]:
    q = normalize_text(person_query)
    if not q:
        return []
    out = []
    for asset in assets:
        assignee = normalize_text(asset.get("assigned_to", ""))
        if not assignee:
            continue
        if q in assignee:
            out.append(asset)
            continue
        # match más laxo por tokens (nombre/apellido/mail parcial)
        tokens = [t for t in re.findall(r"[a-z0-9@._-]+", q) if len(t) >= 3]
        if tokens and all(token in assignee for token in tokens):
            out.append(asset)
    return out


def local_chat_answer(assets: list[dict[str, Any]], prompt: str, prefiltered: bool = False) -> str:
    return answer_inventory_question(assets, prompt)


def flatten_asset_for_display(asset: dict[str, Any], include_all_attributes: bool = True) -> dict[str, Any]:
    row = {
        "Nombre": asset.get("name", ""),
        "Jira": asset.get("jira_key", ""),
        "Tipo": asset.get("object_type", ""),
        "Categoría": asset.get("category", ""),
        "Estado": asset.get("status", ""),
        "Asignado": asset.get("assigned_to", ""),
        "Serial": asset.get("serial_number", ""),
        "Hostname": asset.get("hostname", ""),
        "Modelo": asset.get("model", ""),
        "País": asset.get("country", ""),
        "Compañía": asset.get("company", ""),
        "Entidad": asset.get("entity", ""),
        "Proveedor": asset.get("provider", ""),
        "Fecha compra": asset.get("purchase_date", ""),
        "Fecha garantía": asset.get("warranty_date", ""),
        "Costo": asset.get("purchase_price", ""),
        "Creado": asset.get("created", ""),
    }
    if include_all_attributes:
        for attr_name in sorted(asset.get("attrs_by_name", {}).keys()):
            row[f"Atributo: {attr_name}"] = asset["attrs_by_name"][attr_name]
    return row


def parse_cost(value: str) -> float:
    raw = str(value or "").strip()
    if not raw:
        return 0.0
    cleaned = raw.replace(" ", "")
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def get_os_value(asset: dict[str, Any]) -> str:
    attrs = asset.get("attrs_by_name", {})
    for name, value in attrs.items():
        norm = normalize_lookup_key(name)
        if norm in {"os", "sistema operativo", "operating system"} or "operativo" in norm:
            return str(value)
    blob = " ".join([str(asset.get("name", "")), str(asset.get("model", ""))]).lower()
    if "macbook" in blob or "mac" in blob:
        return "macOS"
    if "windows" in blob:
        return "Windows"
    return "Sin dato"


def build_dependency_rows(assets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    edges = []
    relation_tokens = {"depende", "dependencia", "relacion", "related", "linked", "switch", "server", "servidor", "conect"}
    for a in assets:
        source = a.get("name") or a.get("jira_key") or "Sin fuente"
        for attr_name, attr_value in a.get("attrs_by_name", {}).items():
            norm = normalize_lookup_key(attr_name)
            if not any(token in norm for token in relation_tokens):
                continue
            targets = [t.strip() for t in str(attr_value).split("|") if t.strip()]
            for target in targets:
                edges.append({"Origen": source, "Relación": attr_name, "Destino": target})
    return edges


def get_dynamic_attr(asset: dict[str, Any], aliases: list[str]) -> str:
    attrs = asset.get("attrs_by_name", {})
    norm_map = {normalize_lookup_key(k): str(v) for k, v in attrs.items()}
    for alias in aliases:
        a = normalize_lookup_key(alias)
        if a in norm_map and norm_map[a].strip():
            return norm_map[a].strip()
    for key, value in attrs.items():
        nkey = normalize_lookup_key(key)
        if any(normalize_lookup_key(alias) in nkey for alias in aliases):
            if str(value).strip():
                return str(value).strip()
    return ""


def get_os_version(asset: dict[str, Any]) -> str:
    return get_dynamic_attr(asset, ["Version del SO", "Versión del SO", "Sistema Operativo", "OS", "Operating System"]) or get_os_value(asset)


def get_invoice_number(asset: dict[str, Any]) -> str:
    return get_dynamic_attr(asset, ["Numero de factura", "Número de factura", "Factura", "Invoice"])


def get_serial_value(asset: dict[str, Any]) -> str:
    base = str(asset.get("serial_number") or "").strip()
    if base:
        return base
    return get_dynamic_attr(asset, ["Serial Number", "Serial", "Nro de serie", "Número de serie"])


def get_hostname_value(asset: dict[str, Any]) -> str:
    base = str(asset.get("hostname") or "").strip()
    if base:
        return base
    return get_dynamic_attr(asset, ["Hostname", "Host name", "Nombre de host"])


def resolve_attr_id(asset: dict[str, Any], aliases: list[str], fallback_id: str = "") -> str:
    name_to_id = asset.get("attr_name_to_id", {}) or {}
    alias_norms = [normalize_lookup_key(a) for a in aliases if normalize_lookup_key(a)]

    for alias_norm in alias_norms:
        found = str(name_to_id.get(alias_norm, "")).strip()
        if found:
            return found

    for key, attr_id in name_to_id.items():
        key_norm = normalize_lookup_key(key)
        if not key_norm:
            continue
        if any(alias in key_norm or key_norm in alias for alias in alias_norms):
            found = str(attr_id or "").strip()
            if found:
                return found

    return str(fallback_id or "").strip()


def normalize_company(value: str) -> str:
    raw = normalize_lookup_key(value)
    if raw in COMPANIA_CANONICAS:
        return COMPANIA_CANONICAS[raw]
    for canonical, kws in COMPANIA_KEYWORDS.items():
        nkws = {normalize_lookup_key(k) for k in kws}
        if raw in nkws or any(k in raw for k in nkws if len(k) >= 4):
            return canonical
    return value.strip() or ""


def update_asset_attributes(config: AppConfig, object_id: str, object_type_id: str, attrs: list[dict[str, Any]]) -> tuple[bool, str]:
    auth, headers = build_auth_headers(config)
    payload_full = {"objectTypeId": str(object_type_id), "attributes": attrs}
    payload_attrs = {"attributes": attrs}

    url_sets = [
        (
            f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/{object_id}",
            f"{config.site}/gateway/api/jsm/assets/workspace/{config.workspace_id}/v1/object/{object_id}/attributes",
        ),
        (
            f"{config.site}/rest/servicedeskapi/assets/workspace/{config.workspace_id}/v1/object/{object_id}",
            f"{config.site}/rest/servicedeskapi/assets/workspace/{config.workspace_id}/v1/object/{object_id}/attributes",
        ),
    ]

    attempts: list[tuple[str, str, dict[str, Any]]] = []
    for object_url, attrs_url in url_sets:
        # Rutas de actualización completa del objeto.
        attempts.append(("PUT", object_url, payload_full))
        attempts.append(("PATCH", object_url, payload_full))
        # Rutas de actualización parcial de atributos.
        attempts.append(("PUT", attrs_url, payload_full))
        attempts.append(("PATCH", attrs_url, payload_full))
        attempts.append(("PUT", attrs_url, payload_attrs))
        attempts.append(("PATCH", attrs_url, payload_attrs))
        attempts.append(("POST", attrs_url, payload_attrs))

    errors: list[str] = []
    seen = set()
    for method, url, payload in attempts:
        sig = (method, url, tuple(sorted(payload.keys())))
        if sig in seen:
            continue
        seen.add(sig)
        try:
            response = jira_request_with_retry(method, url, auth=auth, headers=headers, json_payload=payload)
            if response.status_code in (200, 201):
                return True, f"Actualización aplicada ({method} {url.split('/workspace/')[-1]})"
        except Exception as exc:
            errors.append(f"{method} {url}: {exc}")
            continue

    short = " | ".join(errors[:3]) if errors else "sin detalle"
    return False, f"Error actualizando asset: {short}"


# ── 6. LÓGICA DE NEGOCIO (assign, unassign, status, bulk) ────────────────
def assign_asset(config: AppConfig, assets: list[dict[str, Any]], identifier: str, user_email_or_name: str) -> tuple[bool, str]:
    candidate = find_asset_by_identifier(assets, identifier)
    if not candidate:
        return False, f"No encontré activo `{identifier}`."
    auth, _ = build_auth_headers(config)
    assignee_value: str = user_email_or_name
    account_id = resolve_user_account_id(config, user_email_or_name, auth)
    if account_id:
        assignee_value = account_id
    assignee_attr_id = resolve_attr_id(
        candidate,
        ["Usuario asignado", "Asignado a", "Assigned To", "User assigned", "Asignación", "Asignacion"],
        ID_ASIGNACION,
    )
    status_attr_id = resolve_attr_id(candidate, ["Estado del activo", "Estado", "Status"], ID_ESTADO)
    attrs_payload = []
    if assignee_attr_id:
        attrs_payload.append({"objectTypeAttributeId": assignee_attr_id, "objectAttributeValues": [{"value": assignee_value}]})
    if status_attr_id:
        attrs_payload.append({"objectTypeAttributeId": status_attr_id, "objectAttributeValues": [{"value": "En uso"}]})
    if not attrs_payload:
        return False, f"No encontré atributos actualizables en `{candidate.get('jira_key') or candidate.get('name')}`."

    before_assignee = str(candidate.get("assigned_to") or "")
    ok, msg = update_asset_attributes(
        config,
        str(candidate.get("object_id", "")),
        str(candidate.get("object_type_id", "")),
        attrs_payload,
    )
    out = f"{msg}. Asset `{candidate.get('jira_key') or candidate.get('name')}`"
    if ok:
        log_movimiento(
            config,
            candidate,
            "ASIGNACION",
            "Usuario asignado",
            before_assignee,
            user_email_or_name,
            "OK",
            out,
            identifier,
        )
    return ok, out


def unassign_asset(config: AppConfig, assets: list[dict[str, Any]], identifier: str, target_status: str = "Stock usado") -> tuple[bool, str]:
    candidate = find_asset_by_identifier(assets, identifier)
    if not candidate:
        return False, f"No encontré activo `{identifier}`."

    assignee_attr_id = resolve_attr_id(
        candidate,
        ["Usuario asignado", "Asignado a", "Assigned To", "User assigned", "Asignación", "Asignacion"],
        ID_ASIGNACION,
    )
    status_attr_id = resolve_attr_id(candidate, ["Estado del activo", "Estado", "Status"], ID_ESTADO)

    attrs_payload = []
    if assignee_attr_id:
        attrs_payload.append({"objectTypeAttributeId": assignee_attr_id, "objectAttributeValues": []})
    if status_attr_id:
        attrs_payload.append({"objectTypeAttributeId": status_attr_id, "objectAttributeValues": [{"value": target_status}]})
    if not attrs_payload:
        return False, f"No encontré atributos actualizables en `{candidate.get('jira_key') or candidate.get('name')}`."

    before_assignee = str(candidate.get("assigned_to") or "")
    ok, msg = update_asset_attributes(
        config,
        str(candidate.get("object_id", "")),
        str(candidate.get("object_type_id", "")),
        attrs_payload,
    )
    out = f"{msg}. Asset `{candidate.get('jira_key') or candidate.get('name')}` (desasignado, estado: {target_status})"
    if ok:
        log_movimiento(
            config,
            candidate,
            "DESASIGNACION",
            "Usuario asignado",
            before_assignee,
            "",
            "OK",
            out,
            identifier,
        )
    return ok, out


def update_status(config: AppConfig, assets: list[dict[str, Any]], identifier: str, new_status: str) -> tuple[bool, str]:
    candidate = find_asset_by_identifier(assets, identifier)
    if not candidate:
        return False, f"No encontré activo `{identifier}`."
    status_attr_id = resolve_attr_id(candidate, ["Estado del activo", "Estado", "Status"], ID_ESTADO)
    if not status_attr_id:
        return False, f"No encontré atributo de estado en `{candidate.get('jira_key') or candidate.get('name')}`."
    before_status = str(candidate.get("status") or "")
    ok, msg = update_asset_attributes(
        config,
        str(candidate.get("object_id", "")),
        str(candidate.get("object_type_id", "")),
        [{"objectTypeAttributeId": status_attr_id, "objectAttributeValues": [{"value": new_status}]}],
    )
    out = f"{msg}. Estado => {new_status}. Asset `{candidate.get('jira_key') or candidate.get('name')}`"
    if ok:
        log_movimiento(
            config,
            candidate,
            "CAMBIO_ESTADO",
            "Estado del activo",
            before_status,
            new_status,
            "OK",
            out,
            identifier,
        )
    return ok, out


def bulk_update_location(
    config: AppConfig,
    assets: list[dict[str, Any]],
    serials_or_hosts: list[str],
    target_company: str,
    target_country: str,
) -> tuple[int, list[str]]:
    updated = 0
    errors: list[str] = []
    company = normalize_company(target_company) or target_company
    country = canonical_country(target_country)
    for ident in serials_or_hosts:
        asset = find_asset_by_identifier(assets, ident)
        if not asset:
            errors.append(f"{ident}: no encontrado")
            continue
        company_attr_id = resolve_attr_id(asset, ["Compañía", "Compania", "Company"], ID_COMPANIA)
        country_attr_id = resolve_attr_id(asset, ["País", "Pais", "Country"], ID_PAIS)
        attrs_payload = []
        if company_attr_id:
            attrs_payload.append({"objectTypeAttributeId": company_attr_id, "objectAttributeValues": [{"value": company}]})
        if country_attr_id:
            attrs_payload.append({"objectTypeAttributeId": country_attr_id, "objectAttributeValues": [{"value": country}]})
        if not attrs_payload:
            errors.append(f"{ident}: sin atributos de compañía/país para actualizar")
            continue
        ok, msg = update_asset_attributes(
            config,
            str(asset.get("object_id", "")),
            str(asset.get("object_type_id", "")),
            attrs_payload,
        )
        if ok:
            updated += 1
            log_movimiento(
                config,
                asset,
                "BULK_UPDATE",
                "Pais/Compania",
                "",
                f"{country}/{company}",
                "OK",
                "Actualización masiva de ubicación",
                ident,
            )
        else:
            errors.append(f"{ident}: {msg}")
    return updated, errors


def flag_missing_data(assets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for a in assets:
        missing = []
        if not str(a.get("serial_number", "")).strip():
            missing.append("Serial Number")
        if not str(a.get("purchase_price", "")).strip():
            missing.append("Costo")
        if not get_invoice_number(a):
            missing.append("Número de factura")
        if missing:
            rows.append(
                {
                    "Asset": a.get("name") or a.get("jira_key"),
                    "Jira": a.get("jira_key", ""),
                    "Hostname": a.get("hostname", ""),
                    "Faltantes": ", ".join(missing),
                }
            )
    return rows


def detect_duplicates(assets: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    serial_map: dict[str, list[dict[str, Any]]] = {}
    host_map: dict[str, list[dict[str, Any]]] = {}
    for a in assets:
        s = normalize_text(a.get("serial_number", ""))
        h = normalize_text(a.get("hostname", ""))
        if s:
            serial_map.setdefault(s, []).append(a)
        if h:
            host_map.setdefault(h, []).append(a)
    serial_dups = [{"key": k, "count": len(v), "assets": [x.get("jira_key") or x.get("name") for x in v]} for k, v in serial_map.items() if len(v) > 1]
    host_dups = [{"key": k, "count": len(v), "assets": [x.get("jira_key") or x.get("name") for x in v]} for k, v in host_map.items() if len(v) > 1]
    return {"serial_duplicates": serial_dups, "hostname_duplicates": host_dups}


def generate_handover_document(assets: list[dict[str, Any]], identifier: str) -> str:
    asset = find_asset_by_identifier(assets, identifier)
    if not asset:
        raise ValueError(f"No encontré activo `{identifier}`.")
    out_dir = Path("reportes_chat")
    out_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = out_dir / f"acta_entrega_{normalize_lookup_key(asset.get('hostname') or asset.get('serial_number') or 'asset').replace(' ', '_')}_{stamp}.md"
    content = [
        "# Acta de Entrega",
        "",
        f"- Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- Activo: {asset.get('name')}",
        f"- Jira: {asset.get('jira_key')}",
        f"- Hostname: {asset.get('hostname')}",
        f"- Serial: {asset.get('serial_number')}",
        f"- Usuario asignado: {asset.get('assigned_to') or 'Sin asignar'}",
        f"- Estado: {asset.get('status')}",
        f"- País: {asset.get('country')}",
        f"- Compañía: {asset.get('company')}",
        "",
        "Firma receptor: _____________________",
        "Firma IT: ___________________________",
    ]
    filename.write_text("\n".join(content), encoding="utf-8")
    return str(filename)


def predict_stock_depletion(assets: list[dict[str, Any]]) -> str:
    available = [a for a in assets if normalize_text(a.get("status")) in {normalize_text("Stock nuevo"), normalize_text("Stock usado")}]
    in_use_recent = 0
    cutoff = datetime.now().date() - timedelta(days=30)
    for a in assets:
        if normalize_text(a.get("status")) != normalize_text("En uso"):
            continue
        created = parse_date(str(a.get("created", "")))
        if created and created.date() >= cutoff:
            in_use_recent += 1
    rate = in_use_recent / 30 if in_use_recent > 0 else 0
    if rate <= 0:
        return "No hay suficiente señal reciente para estimar agotamiento."
    days = round(len(available) / rate, 1)
    return f"Con ritmo reciente (~{rate:.2f} asignaciones/día), el stock disponible se agotaría en ~{days} días."


def suggest_reallocation(assets: list[dict[str, Any]]) -> str:
    stock_by_company: dict[str, int] = {}
    for a in assets:
        if normalize_text(a.get("status")) not in {normalize_text("Stock nuevo"), normalize_text("Stock usado")}:
            continue
        comp = normalize_company(str(a.get("company", ""))) or "Sin compañía"
        stock_by_company[comp] = stock_by_company.get(comp, 0) + 1
    if not stock_by_company:
        return "No hay stock disponible para sugerir reasignación."
    ordered = sorted(stock_by_company.items(), key=lambda x: x[1], reverse=True)
    donor, donor_qty = ordered[0]
    receiver, receiver_qty = ordered[-1]
    if donor == receiver:
        return "No se detecta desequilibrio entre compañías."
    if donor_qty - receiver_qty < 5:
        return "No hay brecha significativa de stock entre compañías."
    return f"Sugerencia: mover parte del stock de `{donor}` ({donor_qty}) a `{receiver}` ({receiver_qty})."


def calculate_depreciation(assets: list[dict[str, Any]], years: int = 3) -> dict[str, Any]:
    total_original = 0.0
    total_book = 0.0
    rows = []
    for a in assets:
        cost = parse_cost(str(a.get("purchase_price", "")))
        if cost <= 0:
            continue
        purchase = parse_date(str(a.get("purchase_date", "")))
        if not purchase:
            continue
        age_days = max((datetime.now().date() - purchase.date()).days, 0)
        life_days = years * 365
        residual = 0.0
        if age_days >= life_days:
            book = residual
        else:
            book = cost * (1 - age_days / life_days)
        total_original += cost
        total_book += book
        rows.append({"asset": a.get("jira_key") or a.get("name"), "original": round(cost, 2), "book_value": round(book, 2)})
    return {"total_original": round(total_original, 2), "total_book_value": round(total_book, 2), "rows": rows}


def parse_bulk_location_action(prompt: str) -> tuple[list[str], str, str] | None:
    t = normalize_lookup_key(prompt)
    if not re.search(r"\b(?:bulk|lote|masiv\w*|actualiz\w*|cambi\w*|mov\w*|pas\w*)\b", t):
        return None
    ids = re.findall(r"[A-Za-z0-9][A-Za-z0-9._\-/]{5,}", prompt or "")
    ids = [i for i in ids if re.search(r"[A-Za-z]", i) and re.search(r"\d", i)]
    if not ids:
        return None
    country = detect_country_from_prompt(prompt) or ""
    company = detect_company_from_prompt(prompt) or ""
    if not country and not company:
        return None
    if not country and company:
        country = {"Bancar ARG": "Argentina", "Bancar COL": "Colombia", "Bancar MEX": "México"}.get(company, "")
    if not company and country:
        company = company_for_country(country)
    if not company or not country:
        return None
    return ids, company, country


def parse_filters_from_prompt(prompt: str) -> dict[str, str]:
    t = normalize_text(prompt)
    f: dict[str, str] = {}
    attribute_search = detect_attribute_search(prompt)
    category = None if attribute_search else detect_category_from_prompt(prompt)
    status = detect_status_from_prompt(prompt)
    country = detect_country_from_prompt(prompt)
    company = detect_company_from_prompt(prompt)
    entity = detect_entity_from_prompt(prompt)
    person = parse_assignee_query(prompt)
    identifier_context = any(x in t for x in ["serial", "sn", "equipo", "hostname", "host", "jira", "key"]) or bool(parse_assignee_of_identifier_query(prompt)) or len(t.split()) <= 3
    serial = extract_serial_candidate(prompt) if identifier_context else None
    if not serial:
        # Caso común: el usuario escribe solo el identificador (serial/hostname/key).
        token = (prompt or "").strip()
        if token and len(t.split()) <= 3 and re.search(r"[A-Za-z]", token) and re.search(r"\d", token):
            serial = token
    if not serial:
        # Soporta tokens solo-letras (ej: wksar, mbpro) para buscar por hostname/name.
        token = (prompt or "").strip()
        if len(t.split()) == 1 and re.fullmatch(r"[A-Za-z]{4,}", token or ""):
            serial = token
    if not serial:
        owner_of = parse_assignee_of_identifier_query(prompt)
        if owner_of:
            serial = owner_of
    if not serial and identifier_context:
        candidates = extract_identifier_candidates(prompt)
        if candidates:
            serial = candidates[0]
    if category:
        f["category"] = category
    if status:
        f["status"] = status
    if country:
        f["country"] = country
    if company:
        f["company"] = company
    if entity:
        f["entity"] = entity
    if person:
        f["assignee"] = person
    if serial:
        f["identifier"] = serial
    mail = re.search(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", prompt or "")
    if mail:
        f["assignee"] = mail.group(0)
    model = re.search(r"(?:modelo|model)\s+(.+)$", prompt, flags=re.IGNORECASE)
    if model and not attribute_search:
        f["model"] = model.group(1).strip().strip("\"'")
    return f


def apply_filters(assets: list[dict[str, Any]], filters: dict[str, str]) -> list[dict[str, Any]]:
    out = []
    ident = filters.get("identifier", "")
    ident_norm = normalize_text(ident)
    for a in assets:
        if filters.get("category") and normalize_text(a.get("category")) != normalize_text(filters["category"]):
            continue
        if filters.get("status"):
            stv = normalize_text(a.get("status"))
            wanted = normalize_text(filters["status"])
            if wanted == normalize_text("En uso"):
                if stv != wanted:
                    continue
            elif stv != wanted:
                continue
        if filters.get("country") and normalize_text(filters["country"]) not in normalize_text(a.get("country")):
            continue
        if filters.get("company") and normalize_text(filters["company"]) not in normalize_text(a.get("company")):
            continue
        if filters.get("entity") and normalize_text(filters["entity"]) not in normalize_text(a.get("entity")):
            continue
        if filters.get("assignee") and normalize_text(filters["assignee"]) not in normalize_text(a.get("assigned_to")):
            continue
        if filters.get("model") and normalize_text(filters["model"]) not in normalize_text(a.get("model")):
            continue
        if ident:
            exact_match = (
                ident_norm == normalize_text(get_serial_value(a))
                or ident_norm == normalize_text(get_hostname_value(a))
                or ident_norm == normalize_text(a.get("jira_key"))
            )
            partial_match = (
                ident_norm in normalize_text(get_hostname_value(a))
                or ident_norm in normalize_text(a.get("name"))
            )
            if not (exact_match or partial_match):
                continue
        out.append(a)
    return out


def search_assets_by_attribute(assets: list[dict[str, Any]], campo: str, operador: str, valor: str) -> list[dict[str, Any]]:
    """Busca activos por campo principal o atributo dinámico con operador lógico."""
    matches: list[dict[str, Any]] = []
    campo_norm = normalize_lookup_key(campo)
    valor_norm = normalize_text(valor)
    for asset in assets:
        if campo in asset:
            source_value = str(asset.get(campo, ""))
        else:
            source_value = ""
            for key, raw_value in (asset.get("attrs_by_name") or {}).items():
                if normalize_lookup_key(key) == campo_norm:
                    source_value = str(raw_value)
                    break
        src = normalize_text(source_value)
        ok = False
        if operador == "contiene":
            ok = valor_norm in src
        elif operador == "empieza_con":
            ok = src.startswith(valor_norm)
        elif operador == "termina_con":
            ok = src.endswith(valor_norm)
        elif operador == "igual_a":
            ok = src == valor_norm
        elif operador == "regex":
            ok = bool(re.search(valor, source_value, flags=re.IGNORECASE))
        elif operador == "no_contiene":
            ok = valor_norm not in src
        if ok:
            matches.append(asset)
    return matches


def detect_attribute_search(prompt: str) -> tuple[str, str, str] | None:
    """Detecta consultas libres por atributo en lenguaje natural."""
    pattern = r"(?:activos|equipos|assets)\s+(?:donde|con|que\s+t(?:ienen?|engan))\s+(\w[\w\s]*?)\s+(contiene|empieza\s+con|termina\s+con|es|igual\s+a)\s+['\"]?(.+)['\"]?$"
    m = re.search(pattern, prompt.strip(), flags=re.IGNORECASE)
    if not m:
        return None
    campo = m.group(1).strip()
    raw_op = normalize_lookup_key(m.group(2))
    valor = m.group(3).strip()
    mapping = {
        "contiene": "contiene",
        "empieza con": "empieza_con",
        "termina con": "termina_con",
        "es": "igual_a",
        "igual a": "igual_a",
    }
    return campo, mapping.get(raw_op, "contiene"), valor


def _read_json_file(path: Path, default: Any) -> Any:
    try:
        if not path.exists():
            return default
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        push_app_error("_read_json_file", f"{path.name}: {exc}")
        return default


def _write_json_file(path: Path, payload: Any) -> bool:
    try:
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return True
    except Exception as exc:
        push_app_error("_write_json_file", f"{path.name}: {exc}")
        return False


def load_assets_snapshot() -> tuple[list[dict[str, Any]], dict[str, int], datetime | None]:
    payload = _read_json_file(ASSETS_SNAPSHOT_FILE, {})
    if not isinstance(payload, dict):
        return [], {}, None
    assets = payload.get("assets") if isinstance(payload.get("assets"), list) else []
    metadata = payload.get("metadata") if isinstance(payload.get("metadata"), dict) else {}
    saved_at_raw = str(payload.get("saved_at") or "").strip()
    saved_at = None
    if saved_at_raw:
        try:
            saved_at = datetime.fromisoformat(saved_at_raw)
        except ValueError:
            saved_at = None
    return assets, metadata, saved_at


def save_assets_snapshot(assets: list[dict[str, Any]], metadata: dict[str, int]) -> bool:
    return _write_json_file(
        ASSETS_SNAPSHOT_FILE,
        {
            "saved_at": datetime.now().isoformat(timespec="seconds"),
            "assets": assets,
            "metadata": metadata,
        },
    )


def load_auto_assign_rules() -> list[dict[str, Any]]:
    data = _read_json_file(AUTO_ASSIGN_RULES_FILE, [])
    return data if isinstance(data, list) else []


def save_auto_assign_rules(rules: list[dict[str, Any]]) -> bool:
    return _write_json_file(AUTO_ASSIGN_RULES_FILE, rules)


def load_auto_assign_log() -> list[dict[str, Any]]:
    data = _read_json_file(AUTO_ASSIGN_LOG_FILE, [])
    if not isinstance(data, list):
        return []
    return data[-100:]


def save_auto_assign_log(rows: list[dict[str, Any]]) -> bool:
    return _write_json_file(AUTO_ASSIGN_LOG_FILE, rows[-100:])


def save_normalization_rules(rules: list[dict[str, Any]]) -> bool:
    return _write_json_file(RULES_FILE, rules)


def _get_asset_condition_value(asset: dict[str, Any], campo: str) -> str:
    direct = str(asset.get(campo) or "")
    if direct:
        return direct
    aliases = {
        "hostname": {"hostname", "host", "host name"},
        "country": {"pais", "país", "country", "pais region", "pais/region"},
        "category": {"categoria", "categoría", "category"},
        "company": {"compania", "compañia", "company"},
    }
    target_aliases = aliases.get(campo, {normalize_lookup_key(campo)})
    for key, raw in (asset.get("attrs_by_name") or {}).items():
        if normalize_lookup_key(key) in target_aliases and raw:
            return str(raw)
    return ""


def evaluar_regla_asignacion(asset: dict[str, Any], regla: ReglaAsignacionAuto) -> bool:
    value = _get_asset_condition_value(asset, regla.campo_condicion)
    if not value:
        return False
    src = normalize_text(value)
    trg = normalize_text(regla.valor_condicion)
    if regla.operador == "empieza_con":
        return src.startswith(trg)
    if regla.operador == "termina_con":
        return src.endswith(trg)
    if regla.operador == "contiene":
        return trg in src
    if regla.operador == "igual_a":
        return src == trg
    if regla.operador == "regex":
        try:
            return bool(re.search(regla.valor_condicion, value, re.IGNORECASE))
        except re.error:
            return False
    return False


def _asset_identifier(asset: dict[str, Any]) -> str:
    return str(asset.get("serial_number") or asset.get("hostname") or asset.get("jira_key") or asset.get("name") or "").strip()


def ejecutar_asignacion_automatica(config: AppConfig, asset: dict[str, Any], regla: ReglaAsignacionAuto) -> dict[str, Any]:
    identifier = _asset_identifier(asset)
    result = {"timestamp": datetime.now().isoformat(timespec="seconds"), "activo": identifier, "regla": regla.nombre, "ok": False, "detalle": ""}
    if not identifier:
        result["detalle"] = "Activo sin identificador utilizable."
        return result

    if regla.pais_destino and normalize_text(regla.pais_destino) != normalize_text(asset.get("country")):
        result["detalle"] = "País no coincide con regla."
        return result

    if regla.tipo_accion == "asignar_usuario":
        if not regla.usuario_destino.strip():
            result["detalle"] = "Regla sin usuario destino."
            return result
        ok, msg = assign_asset(config, [asset], identifier, regla.usuario_destino.strip())
        result["ok"] = ok
        result["detalle"] = msg
        return result

    if regla.tipo_accion == "asignar_por_stock":
        queue = [u.strip() for u in (regla.cola_usuarios or []) if str(u).strip()]
        if not queue:
            result["detalle"] = "Regla sin cola de usuarios."
            return result
        rr_index = int(st.session_state.get("auto_assign_rr_idx", {}).get(regla.nombre, 0))
        target_user = queue[rr_index % len(queue)]
        ok, msg = assign_asset(config, [asset], identifier, target_user)
        rr_map = st.session_state.get("auto_assign_rr_idx", {})
        rr_map[regla.nombre] = rr_index + 1
        st.session_state["auto_assign_rr_idx"] = rr_map
        result["ok"] = ok
        result["detalle"] = msg
        result["usuario"] = target_user
        return result

    if regla.tipo_accion == "cambiar_estado":
        if not regla.estado_destino.strip():
            result["detalle"] = "Regla sin estado destino."
            return result
        ok, msg = update_status(config, [asset], identifier, regla.estado_destino.strip())
        result["ok"] = ok
        result["detalle"] = msg
        return result

    result["detalle"] = "Tipo de acción no soportado."
    return result


def auto_assign_job(config: AppConfig) -> list[dict[str, Any]]:
    assets_fresh = fetch_assets(config)
    snapshot = _read_json_file(AUTO_ASSIGN_SNAPSHOT_FILE, [])
    prev_keys = {str(a.get("jira_key") or "") for a in snapshot if str(a.get("jira_key") or "")}

    nuevos = [a for a in assets_fresh if a.get("jira_key") and str(a.get("jira_key")) not in prev_keys]
    sin_asignar = [
        a
        for a in assets_fresh
        if not str(a.get("assigned_to") or "").strip()
        and normalize_text(a.get("status")) in {normalize_text("stock nuevo"), normalize_text("stock")}
    ]
    candidatos = nuevos + [a for a in sin_asignar if a not in nuevos]

    rules = [ReglaAsignacionAuto(**row) for row in load_auto_assign_rules() if isinstance(row, dict) and row.get("activa")]
    rules = sorted(rules, key=lambda r: int(r.prioridad or 0), reverse=True)
    resultados: list[dict[str, Any]] = []
    for activo in candidatos:
        for regla in rules:
            if evaluar_regla_asignacion(activo, regla):
                resultados.append(ejecutar_asignacion_automatica(config, activo, regla))
                break

    _write_json_file(AUTO_ASSIGN_SNAPSHOT_FILE, assets_fresh)
    if resultados:
        existing_log = load_auto_assign_log()
        save_auto_assign_log(existing_log + resultados)
    return resultados


def start_auto_assign_scheduler(config: AppConfig) -> None:
    if BackgroundScheduler is None:
        return
    if st.session_state.get("scheduler_running"):
        return
    scheduler = BackgroundScheduler(daemon=True)
    scheduler.add_job(
        auto_assign_job,
        trigger="interval",
        seconds=int(os.getenv("AUTO_ASSIGN_INTERVAL", "60")),
        args=[config],
        id="auto_assign",
        replace_existing=True,
    )
    scheduler.start()
    st.session_state["scheduler_running"] = True
    st.session_state["scheduler_instance"] = scheduler


def stop_auto_assign_scheduler() -> None:
    scheduler = st.session_state.get("scheduler_instance")
    if scheduler is not None:
        try:
            scheduler.shutdown(wait=False)
        except Exception:
            pass
    st.session_state["scheduler_running"] = False
    st.session_state["scheduler_instance"] = None


def parse_normalization_rule_from_prompt(prompt: str) -> ReglaNormalizacion | None:
    """Interpreta reglas de normalización masiva desde lenguaje natural."""
    text = prompt.strip()
    text = re.split(r"[,.]\s*(?:asignarle|asignar|poner|setear|actualizar)\b", text, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    text_norm = normalize_lookup_key(text)

    def _destino(destino_raw: str) -> tuple[str, str]:
        country_guess = canonical_country(destino_raw)
        if country_guess != "Sin país":
            return "country", country_guess
        company_guess = normalize_company(destino_raw)
        if company_guess:
            return "company", company_guess
        if normalize_text(destino_raw) in ESTADO_NORMALIZATION:
            return "status", canonical_status(destino_raw)
        return "model", destino_raw.strip()

    # 1) "todos los de bancar arg son de argentina"
    p_company = re.search(r"(?:todos\s+)?(?:los\s+)?de\s+(.+?)\s+son\s+(?:de\s+)?(.+)$", text, flags=re.IGNORECASE)
    if p_company:
        cond = p_company.group(1).strip()
        campo_destino, valor_nuevo = _destino(p_company.group(2).strip())
        cond_candidates = extract_identifier_candidates(cond)
        if cond_candidates:
            cond_norm = normalize_lookup_key(cond_candidates[0])
            return ReglaNormalizacion(
                campo_condicion="hostname",
                operador="contiene",
                valor_condicion=cond_norm,
                campo_a_modificar=campo_destino,
                valor_nuevo=valor_nuevo,
                descripcion=f"hostname contiene '{cond_norm}' => {campo_destino}='{valor_nuevo}'",
            )
        return ReglaNormalizacion(
            campo_condicion="company",
            operador="contiene",
            valor_condicion=cond,
            campo_a_modificar=campo_destino,
            valor_nuevo=valor_nuevo,
            descripcion=f"company contiene '{cond}' => {campo_destino}='{valor_nuevo}'",
        )

    # 2) Regla explícita por hostname/serial/modelo/company.
    p_explicit = re.search(
        r"(?:todos\s+)?(?:los\s+)?([\w\s]+?)\s+que\s+(empiec[ea]n?\s+con|contien[ea]n?|terminen?\s+con|sean?\s+igual(?:es)?\s+a)\s+['\"]?(\S+)['\"]?\s+son\s+(?:de\s+)?(.+)$",
        text,
        flags=re.IGNORECASE,
    )
    if p_explicit:
        campo_raw = normalize_lookup_key(p_explicit.group(1))
        op_raw = normalize_lookup_key(p_explicit.group(2))
        valor = p_explicit.group(3).strip()
        destino = p_explicit.group(4).strip()
        campo_alias = {
            "host": "hostname",
            "hosts": "hostname",
            "hostname": "hostname",
            "hostnames": "hostname",
            "serial": "serial_number",
            "seriales": "serial_number",
            "serie": "serial_number",
            "modelo": "model",
            "model": "model",
            "compania": "company",
            "compañia": "company",
            "company": "company",
        }
        campo = campo_alias.get(campo_raw, campo_raw or "hostname")
        op_map = {
            "empiecen con": "empieza_con",
            "empieza con": "empieza_con",
            "contienen": "contiene",
            "contiene": "contiene",
            "terminen con": "termina_con",
            "termina con": "termina_con",
            "sean igual a": "igual_a",
            "igual a": "igual_a",
        }
        operador = op_map.get(op_raw, "contiene")
        campo_destino, valor_nuevo = _destino(destino)
        return ReglaNormalizacion(
            campo_condicion=campo,
            operador=operador,
            valor_condicion=valor,
            campo_a_modificar=campo_destino,
            valor_nuevo=valor_nuevo,
            descripcion=f"{campo} {operador} '{valor}' => {campo_destino}='{valor_nuevo}'",
        )

    # 3) Token libre con forma de identificador: hostname contiene.
    p_token = re.search(r"(?:todos\s+)?(?:los\s+)?([A-Za-z]{3,}[0-9]*[A-Za-z0-9._-]*)\s+son\s+(?:de\s+)?(.+)$", text, flags=re.IGNORECASE)
    if p_token:
        token = p_token.group(1).strip()
        if not looks_like_inventory_identifier(token):
            return None
        campo_destino, valor_nuevo = _destino(p_token.group(2).strip())
        return ReglaNormalizacion(
            campo_condicion="hostname",
            operador="contiene",
            valor_condicion=normalize_lookup_key(token),
            campo_a_modificar=campo_destino,
            valor_nuevo=valor_nuevo,
            descripcion=f"hostname contiene '{normalize_lookup_key(token)}' => {campo_destino}='{valor_nuevo}'",
        )

    p_simple = re.search(
        r"(?:los|las|equipos|hosts?|hostnames?)\s+(?:que\s+)?(?:empiezan?\s+con\s+|empiecen\s+con\s+|empezando\s+con\s+)?([A-Z]{2,}[A-Z0-9]*)\s+son\s+de\s+(.+)$",
        text,
        flags=re.IGNORECASE,
    )
    if p_simple:
        token = p_simple.group(1).strip()
        destino = p_simple.group(2).strip()
        campo_destino, valor_nuevo = _destino(destino)
        return ReglaNormalizacion(
            campo_condicion="hostname",
            operador="empieza_con",
            valor_condicion=normalize_lookup_key(token),
            campo_a_modificar=campo_destino,
            valor_nuevo=valor_nuevo,
            descripcion=f"hostname empieza_con '{normalize_lookup_key(token)}' => {campo_destino}='{valor_nuevo}'",
        )

    return None


def evaluar_regla(asset: dict[str, Any], regla: ReglaNormalizacion) -> bool:
    """Evalúa si un activo cumple la condición de una regla."""
    if not regla.campo_condicion:
        return False

    campo = regla.campo_condicion
    target = regla.valor_condicion.strip() if regla.valor_condicion else ""

    # ── PRIORIDAD 1: any_text (blob completo) ──
    if campo == "any_text":
        partes = [
            str(asset.get("name") or ""),
            str(asset.get("hostname") or ""),
            str(asset.get("serial_number") or ""),
            str(asset.get("company") or ""),
            str(asset.get("country") or ""),
            str(asset.get("status") or ""),
            str(asset.get("jira_key") or ""),
            str(asset.get("assigned_to") or ""),
        ]
        for v in (asset.get("attrs_by_name") or {}).values():
            partes.append(str(v or ""))
        for v in (asset.get("attrs_by_id") or {}).values():
            partes.append(str(v or ""))
        valor_campo = " ".join(partes)

    # ── PRIORIDAD 2: campo específico ──
    else:
        # a) Buscar directo en el dict del asset
        valor_campo = str(asset.get(campo) or "")

        # b) Si vacío, buscar en attrs_by_name con alias_map fijo
        if not valor_campo:
            alias_map = {
                "hostname": {"hostname", "host", "host name", "nombre de host", "computername", "computer name"},
                "company": {"compania", "compañia", "company", "empresa"},
                "country": {"pais", "pais region", "país", "country", "pais/region"},
                "serial_number": {"serial", "serial number", "nro serie", "numero de serie"},
                "model": {"modelo", "model", "nombre del modelo"},
                "status": {"estado", "estado del activo", "status"},
                "assigned_to": {"asignacion", "asignación", "assigned to", "usuario", "usuario asignado"},
            }
            aliases = alias_map.get(campo, {normalize_lookup_key(campo)})
            for key, raw_val in (asset.get("attrs_by_name") or {}).items():
                if normalize_lookup_key(key) in aliases and raw_val:
                    valor_campo = str(raw_val)
                    break

        # c) Si sigue vacío y es hostname, buscar cualquier key que contenga "host"
        if not valor_campo and campo == "hostname":
            for key, raw_val in (asset.get("attrs_by_name") or {}).items():
                if "host" in normalize_lookup_key(key) and raw_val:
                    valor_campo = str(raw_val)
                    break

        # d) Último fallback para hostname: usar name
        if not valor_campo and campo == "hostname":
            valor_campo = str(asset.get("name") or "")

    if campo == "hostname":
        host_candidates = []
        host_main = str(asset.get("hostname") or "")
        if host_main:
            host_candidates.append(host_main)
        host_candidates.append(str(asset.get("name") or ""))
        host_candidates.append(str(asset.get("jira_key") or ""))
        for key, raw_val in (asset.get("attrs_by_name") or {}).items():
            key_norm = normalize_lookup_key(key)
            if "host" in key_norm or "computername" in key_norm:
                host_candidates.append(str(raw_val or ""))
        valor_campo = " ".join(v for v in host_candidates if str(v).strip())

    if not valor_campo:
        return False

    src = valor_campo.lower()
    trg = target.lower()

    if regla.operador == "empieza_con":
        if campo == "hostname":
            candidates = []
            for part in valor_campo.split():
                if part:
                    candidates.append(part)
            return any(candidate.lower().startswith(trg) for candidate in candidates)
        return src.startswith(trg) or any(p.lower().startswith(trg) for p in valor_campo.split() if p)
    if regla.operador == "termina_con":
        return src.endswith(trg) or any(
            p.lower().endswith(trg) for p in valor_campo.split() if p
        )
    if regla.operador == "contiene":
        return trg in src
    if regla.operador == "igual_a":
        return normalize_text(valor_campo) == normalize_text(target)
    if regla.operador == "regex":
        try:
            return bool(re.search(target, valor_campo, re.IGNORECASE))
        except re.error:
            return False
    if regla.operador == "no_contiene":
        return trg not in src
    return False


def aplicar_regla(config: AppConfig, assets: list[dict[str, Any]], regla: ReglaNormalizacion, dry_run: bool = False) -> tuple[int, list[str]]:
    """Aplica una regla de normalización en modo simulación o ejecución real."""
    affected = 0
    errors: list[str] = []
    for asset in assets:
        if not evaluar_regla(asset, regla):
            continue
        affected += 1
        if dry_run:
            continue
        alias_map = {
            "country": (["Pais", "País", "Country"], ID_PAIS),
            "company": (["Compañía", "Compania", "Company"], ID_COMPANIA),
            "status": (["Estado del activo", "Estado", "Status"], ID_ESTADO),
            "model": (["Nombre del modelo", "Modelo", "Model"], ID_MODELO),
            "provider": (["Proveedor", "Provider"], ID_PROVEEDOR),
            "serial_number": (["Serial Number", "Serial"], ID_SERIAL),
            "hostname": (["Hostname"], ID_HOSTNAME),
            "entity": (["Entidad del activo", "Entidad", "Entity"], ID_ENTIDAD),
        }
        aliases, fallback = alias_map.get(regla.campo_a_modificar, ([regla.campo_a_modificar], ""))
        attr_id = resolve_attr_id(asset, aliases, fallback)
        if not attr_id:
            errors.append(f"{asset.get('jira_key') or asset.get('name')}: atributo no resoluble")
            continue
        ok, msg = update_asset_attributes(
            config,
            str(asset.get("object_id", "")),
            str(asset.get("object_type_id", "")),
            [{"objectTypeAttributeId": attr_id, "objectAttributeValues": [{"value": regla.valor_nuevo}]}],
        )
        if not ok:
            errors.append(f"{asset.get('jira_key') or asset.get('name')}: {msg}")
            continue
        log_movimiento(
            config,
            asset,
            "REGLA_NORMALIZACION",
            regla.campo_a_modificar,
            str(asset.get(regla.campo_a_modificar) or ""),
            regla.valor_nuevo,
            "OK",
            regla.descripcion,
        )
    return affected, errors


def answer_inventory_question(assets: list[dict[str, Any]], prompt: str) -> str:
    """Responde preguntas de inventario con despacho por prioridad de intención."""
    t = normalize_text(prompt)
    if parse_nl_dashboard_request(prompt):
        return build_dashboard_chat_payload(assets, prompt)
    filters = parse_filters_from_prompt(prompt)
    mail_match = re.search(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", prompt)
    if mail_match and not filters.get("assignee"):
        filters["assignee"] = mail_match.group(0)
    selected = list(apply_filters(assets, filters)) if filters else list(assets)

    def _owner_by_identifier() -> str | None:
        ident = parse_assignee_of_identifier_query(prompt)
        if not ident:
            return None
        asset = find_asset_by_identifier(assets, ident)
        if not asset:
            return "❌ No encontré ese activo para resolver el dueño."
        return f"ℹ️ El activo **{asset.get('jira_key') or asset.get('name') or ident}** está asignado a **{asset.get('assigned_to') or 'Sin asignar'}**."

    def _assignee_assets() -> str | None:
        if not filters.get("assignee"):
            return None
        if not selected:
            return "ℹ️ No encontré activos asignados para ese usuario."
        lines = [f"ℹ️ Encontré **{len(selected)}** activos del usuario:"]
        for asset in selected[:40]:
            lines.append(
                f"- **{asset.get('jira_key') or asset.get('name')}** | {asset.get('hostname') or 'Sin hostname'} | {asset.get('status')} | {asset.get('country')}"
            )
        return "\n".join(lines)

    def _identifier_details() -> str | None:
        ident = filters.get("identifier")
        if not ident:
            return None
        asset = find_asset_by_identifier(assets, ident)
        if not asset:
            ident_norm = normalize_text(ident)
            partials = [
                a
                for a in assets
                if ident_norm in normalize_text(f"{a.get('hostname', '')} {a.get('name', '')}")
            ]
            if not partials:
                return "❌ No encontré ese activo."
            lines = [f"ℹ️ No hubo match exacto para `{ident}`. Coincidencias parciales: **{len(partials)}**"]
            for row in partials[:25]:
                lines.append(
                    f"- **{row.get('jira_key') or row.get('name')}** | Hostname: {row.get('hostname') or 'Sin hostname'} | "
                    f"Serial: {row.get('serial_number') or 'Sin serial'} | Estado: {row.get('status') or 'Sin estado'}"
                )
            return "\n".join(lines)
        return (
            f"ℹ️ Activo encontrado: **{asset.get('jira_key') or asset.get('name') or ident}**\n"
            f"- Nombre: {asset.get('name') or 'Sin nombre'}\n"
            f"- Serial: {get_serial_value(asset) or 'Sin serial'}\n"
            f"- Hostname: {get_hostname_value(asset) or 'Sin hostname'}\n"
            f"- Asignado: {asset.get('assigned_to') or 'Sin asignar'}\n"
            f"- Estado: {asset.get('status') or 'Sin estado'}\n"
            f"- País: {asset.get('country') or 'Sin país'}\n"
            f"- Modelo: {asset.get('model') or 'Sin modelo'}"
        )

    def _attribute_search() -> str | None:
        found = detect_attribute_search(prompt)
        if not found:
            return None
        campo, operador, valor = found
        rows = search_assets_by_attribute(assets, campo, operador, valor)
        if not rows:
            return "ℹ️ No encontré activos para esa búsqueda por atributo."
        lines = [f"ℹ️ Encontré **{len(rows)}** activos donde `{campo}` {operador} `{valor}`:"]
        for asset in rows[:40]:
            lines.append(f"- {asset.get('jira_key') or asset.get('name')} | {asset.get('status')} | {asset.get('assigned_to') or 'Sin asignar'}")
        selected[:] = rows
        return "\n".join(lines)

    def _summary() -> str | None:
        if "resumen ejecutivo" not in t:
            return None
        total = len(selected)
        in_use = sum(1 for a in selected if normalize_text(a.get("status")) == normalize_text("en uso"))
        stock = sum(1 for a in selected if normalize_text(a.get("status")) in {normalize_text("stock nuevo"), normalize_text("stock usado")})
        by_country = Counter((a.get("country") or "Sin país") for a in selected)
        missing = len(flag_missing_data(selected))
        dup = detect_duplicates(selected)
        soon = 0
        now = datetime.now().date()
        for a in selected:
            w = parse_date(str(a.get("warranty_date", "")).split("|")[0].strip())
            if w and w.date() < now:
                soon += 1
        pct = round((in_use / max(total, 1)) * 100, 2)
        return (
            f"ℹ️ Inventario: **{total}** activos, en uso **{in_use} ({pct}%)**, disponibles **{stock}**. "
            f"Países principales: {', '.join(f'{k}:{v}' for k, v in by_country.most_common(3))}. "
            f"Alertas: garantías vencidas **{soon}**, duplicados **{len(dup['serial_duplicates']) + len(dup['hostname_duplicates'])}**, faltantes **{missing}**."
        )

    def _most_expensive() -> str | None:
        if "mas caro" not in t and "más caro" not in t:
            return None
        priced = [(a, parse_cost(str(a.get("purchase_price", "")))) for a in selected]
        priced = [row for row in priced if row[1] > 0]
        if not priced:
            return "❌ No hay costos cargados para resolver esa consulta."
        asset, cost = sorted(priced, key=lambda x: x[1], reverse=True)[0]
        return (
            f"ℹ️ Activo más caro: **{asset.get('name') or asset.get('jira_key')}**\n"
            f"- Serial: {get_serial_value(asset) or 'Sin serial'}\n"
            f"- Modelo: {asset.get('model') or 'Sin modelo'}\n"
            f"- Costo: **{round(cost, 2)}**\n"
            f"- País: {asset.get('country') or 'Sin país'}"
        )

    def _top_users() -> str | None:
        if "top 5 usuarios" not in t:
            return None
        by_user: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for asset in selected:
            user = str(asset.get("assigned_to") or "").strip()
            if user:
                by_user[user].append(asset)
        if not by_user:
            return "ℹ️ No encontré usuarios con activos asignados."
        lines = ["ℹ️ Top 5 usuarios con más equipos:"]
        for user, rows in sorted(by_user.items(), key=lambda x: len(x[1]), reverse=True)[:5]:
            models = Counter((r.get("model") or "Sin modelo") for r in rows)
            lines.append(f"- **{user}** | {len(rows)} equipos | modelos: {', '.join(m for m, _ in models.most_common(3))}")
        return "\n".join(lines)

    def _compare_stock() -> str | None:
        if "comparar stock" not in t:
            return None
        by_country: dict[str, dict[str, int]] = defaultdict(lambda: {"total": 0, "en_uso": 0, "stock": 0, "sin_asignar": 0})
        for asset in selected:
            country = asset.get("country") or "Sin país"
            by_country[country]["total"] += 1
            if normalize_text(asset.get("status")) == normalize_text("en uso"):
                by_country[country]["en_uso"] += 1
            if normalize_text(asset.get("status")) in {normalize_text("stock nuevo"), normalize_text("stock usado")}:
                by_country[country]["stock"] += 1
            if not str(asset.get("assigned_to") or "").strip():
                by_country[country]["sin_asignar"] += 1
        lines = ["ℹ️ País | Total | En uso | Stock | Sin asignar"]
        for country, vals in sorted(by_country.items(), key=lambda x: x[1]["total"], reverse=True):
            lines.append(f"- {country} | {vals['total']} | {vals['en_uso']} | {vals['stock']} | {vals['sin_asignar']}")
        return "\n".join(lines)

    def _inventory_of_user() -> str | None:
        if "inventario completo de" not in t:
            return None
        who = parse_assignee_query(prompt) or prompt.split("de", 1)[-1].strip(" ?.")
        rows = filter_assets_by_assignee(assets, who)
        if not rows:
            return "ℹ️ No encontré activos para ese usuario."
        lines = [f"ℹ️ Inventario completo de **{who}** ({len(rows)} activos):"]
        for asset in rows[:60]:
            lines.append(
                f"- **{asset.get('jira_key') or asset.get('name')}** | Hostname: {get_hostname_value(asset) or 'N/A'} | Serial: {get_serial_value(asset) or 'N/A'} | "
                f"Estado: {asset.get('status')} | País: {asset.get('country')} | Modelo: {asset.get('model')} | Compañía: {asset.get('company')}"
            )
        return "\n".join(lines)

    def _financial() -> str | None:
        if not any(k in t for k in ["costo", "gasto", "inversion", "inversión", "depreciacion", "depreciación"]):
            return None
        if "depreci" in t or "valor contable" in t:
            dep = calculate_depreciation(selected)
            return f"ℹ️ Depreciación estimada: valor original **{dep['total_original']}**, valor contable **{dep['total_book_value']}**."
        total_cost = round(sum(parse_cost(str(a.get("purchase_price", ""))) for a in selected), 2)
        avg = round(total_cost / max(len([a for a in selected if parse_cost(str(a.get('purchase_price', ''))) > 0]), 1), 2)
        return f"ℹ️ Costo total: **{total_cost}** | costo promedio: **{avg}**."

    def _warranty() -> str | None:
        if not any(k in t for k in ["garantia", "garantía", "vence", "vencida", "proximo mes", "próximo mes"]):
            return None
        now = datetime.now().date()
        soon_rows = []
        expired_rows = []
        for asset in selected:
            w = parse_date(str(asset.get("warranty_date", "")).split("|")[0].strip())
            if not w:
                continue
            delta = (w.date() - now).days
            if delta < 0:
                expired_rows.append(asset)
            elif delta <= 45:
                soon_rows.append(asset)
        lines = [f"ℹ️ Garantías vencidas: **{len(expired_rows)}** | próximas (<=45 días): **{len(soon_rows)}**"]
        for asset in (expired_rows + soon_rows)[:30]:
            lines.append(f"- {asset.get('jira_key') or asset.get('name')} | {asset.get('warranty_date')} | {asset.get('status')}")
        return "\n".join(lines)

    def _data_quality() -> str | None:
        if not any(k in t for k in ["faltante", "duplicado", "sin serial", "sin hostname", "sin costo", "sin número de factura", "sin numero de factura"]):
            return None
        missing = flag_missing_data(selected)
        dup = detect_duplicates(selected)
        return (
            f"ℹ️ Calidad de datos: faltantes **{len(missing)}**, duplicados serial **{len(dup['serial_duplicates'])}**, "
            f"duplicados hostname **{len(dup['hostname_duplicates'])}**."
        )

    def _os_distribution() -> str | None:
        if not any(k in t for k in ["so", "sistema operativo", "windows", "macos", "versión del so", "version del so"]):
            return None
        counts: dict[str, int] = {}
        for asset in selected:
            os_value = get_os_version(asset) or "Sin dato"
            counts[os_value] = counts.get(os_value, 0) + 1
        lines = ["ℹ️ Distribución de SO:"] + [f"- {k}: {v}" for k, v in sorted(counts.items(), key=lambda x: x[1], reverse=True)[:10]]
        return "\n".join(lines)

    def _hostname_missing() -> str | None:
        if "hostname" not in t or "sin" not in t:
            return None
        missing_rows = [a for a in selected if not str(get_hostname_value(a)).strip()]
        return f"ℹ️ Hay **{len(missing_rows)}** activos sin hostname."

    def _generic_count() -> str | None:
        count_keywords = ["cuantos", "cuántos", "cantidad", "total", "hay", "tenemos", "existen"]
        if not any(k in t for k in count_keywords):
            return None
        category_info = f" de {filters['category']}" if filters.get("category") else ""
        country_info = f" en {filters['country']}" if filters.get("country") else ""
        status_info = f" con estado {filters['status']}" if filters.get("status") else ""
        preview = ""
        if selected:
            preview = "\n".join(
                [
                    f"- {a.get('jira_key') or a.get('name')} | {a.get('status')} | {a.get('assigned_to') or 'Sin asignar'}"
                    for a in selected[:10]
                ]
            )
        return f"ℹ️ Hay **{len(selected)}** activos{category_info}{country_info}{status_info}.\n\n{preview}".strip()

    def _general() -> str:
        lines = [f"ℹ️ Encontré **{len(selected)}** activos. Muestra:"]
        for asset in selected[:20]:
            lines.append(f"- {asset.get('jira_key') or asset.get('name')} | {asset.get('status')} | {asset.get('assigned_to') or 'Sin asignar'} | {asset.get('country')}")
        return "\n".join(lines)

    intent_handlers: list[tuple[int, Any, Any]] = [
        (100, lambda: parse_assignee_of_identifier_query(prompt) is not None, _owner_by_identifier),
        (95, lambda: bool(filters.get("assignee")), _assignee_assets),
        (90, lambda: bool(filters.get("identifier")), _identifier_details),
        (88, lambda: detect_attribute_search(prompt) is not None, _attribute_search),
        (80, lambda: "resumen ejecutivo" in t, _summary),
        (75, lambda: "mas caro" in t or "más caro" in t, _most_expensive),
        (70, lambda: "top 5 usuarios" in t, _top_users),
        (65, lambda: "comparar stock" in t, _compare_stock),
        (60, lambda: "inventario completo de" in t, _inventory_of_user),
        (55, lambda: any(k in t for k in ["costo", "gasto", "inversion", "inversión", "depreciacion", "depreciación"]), _financial),
        (50, lambda: any(k in t for k in ["garantia", "garantía", "vence", "vencida"]), _warranty),
        (45, lambda: any(k in t for k in ["faltante", "duplicado", "sin serial", "sin hostname"]), _data_quality),
        (40, lambda: any(k in t for k in ["so", "sistema operativo", "windows", "macos"]), _os_distribution),
        (35, lambda: "hostname" in t and "sin" in t, _hostname_missing),
        (30, lambda: any(k in t for k in ["cuantos", "cuántos", "cantidad", "total"]), _generic_count),
        (20, lambda: True, _general),
    ]
    for _, condition_fn, handler_fn in sorted(intent_handlers, key=lambda item: item[0], reverse=True):
        if condition_fn():
            result = handler_fn()
            if result:
                st.session_state["last_chat_result_count"] = len(selected)
                return result
    st.session_state["last_chat_result_count"] = len(selected)
    return _general()


def parse_nl_dashboard_request(prompt: str) -> dict[str, Any]:
    lookup = normalize_lookup_key(prompt)
    tokens = lookup_tokens(prompt)
    show_spend = any(keyword in lookup for keyword in ["gasto", "costo", "inversion", "depreciacion", "valor contable"])
    show_geo = any(keyword in lookup for keyword in ["pais", "geograf", "ubicacion", "region"])
    show_quality = any(keyword in lookup for keyword in ["calidad", "quality", "cobertura", "faltantes", "completitud"])
    show_stock = any(keyword in lookup for keyword in ["stock", "critico", "periferic", "disponibilidad"])
    signal_count = sum(bool(flag) for flag in [show_spend, show_geo, show_quality, show_stock])
    explicit_dashboard = any(matches_lookup_keyword(lookup, hint, lookup_token_set=tokens) for hint in DASHBOARD_EXPLICIT_HINTS)
    visual_intent = any(matches_lookup_keyword(lookup, hint, lookup_token_set=tokens) for hint in DASHBOARD_VISUAL_HINTS)
    aggregation_intent = any(matches_lookup_keyword(lookup, hint, lookup_token_set=tokens) for hint in DASHBOARD_AGGREGATION_HINTS)
    if not (explicit_dashboard or (visual_intent and signal_count >= 1) or signal_count >= 2 or (aggregation_intent and signal_count >= 1)):
        return {}
    if explicit_dashboard and signal_count == 0:
        show_spend = True
        show_geo = True
        show_quality = True
        show_stock = True
    return {
        "show_spend": show_spend,
        "show_geo": show_geo,
        "show_quality": show_quality,
        "show_stock": show_stock,
        "filters": parse_filters_from_prompt(prompt),
        "raw": prompt,
    }


def build_dashboard_chat_payload(assets: list[dict[str, Any]], prompt: str) -> str:
    request = parse_nl_dashboard_request(prompt)
    working = apply_filters(assets, request.get("filters", {})) if request else assets
    total = len(working)
    if total == 0:
        return encode_chat_payload("No encontré activos para ese dashboard.")
    if pd is None or px is None or pio is None:
        return encode_chat_payload("Instalá `pandas` y `plotly` para dashboards visuales.")

    in_use = sum(1 for a in working if normalize_text(a.get("status", "")) == "en uso")
    stock = sum(1 for a in working if "stock" in normalize_text(a.get("status", "")))
    sin_serial = sum(1 for a in working if not str(a.get("serial_number", "")).strip())
    sin_asignar = sum(1 for a in working if not str(a.get("assigned_to") or "").strip())
    costo = round(sum(parse_cost(str(a.get("purchase_price", ""))) for a in working), 2)

    kpi_text = (
        f"**Dashboard — {total} activos**\n\n"
        f"| Métrica | Valor | % |\n|---|---|---|\n"
        f"| En uso | {in_use} | {round(in_use / max(total, 1) * 100, 1)}% |\n"
        f"| Stock disponible | {stock} | {round(stock / max(total, 1) * 100, 1)}% |\n"
        f"| Sin asignar | {sin_asignar} | {round(sin_asignar / max(total, 1) * 100, 1)}% |\n"
        f"| Sin serial | {sin_serial} | {round(sin_serial / max(total, 1) * 100, 1)}% |\n"
        f"| Costo total | ${costo:,.0f} | — |"
    )

    layout = dict(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(size=12),
        margin=dict(t=50, b=30, l=10, r=10),
    )
    color_map = {
        "En uso": "#0f766e",
        "Stock nuevo": "#334155",
        "Stock usado": "#78716c",
        "Asignado al edificio": "#64748b",
        "Sin estado": "#cbd5e1",
    }
    show_all = not any(bool(request.get(key)) for key in ["show_spend", "show_geo", "show_quality", "show_stock"])

    charts: list[dict[str, Any]] = []
    geo_df = pd.DataFrame(
        [{"País": a.get("country") or "Sin país", "Compañía": a.get("company") or "Sin compañía", "Categoría": a.get("category") or "Sin cat"} for a in working]
    )
    if request.get("show_geo") or show_all:
        sun_df = geo_df.groupby(["País", "Compañía", "Categoría"]).size().reset_index(name="Cantidad")
        fig1 = px.sunburst(
            sun_df,
            path=["País", "Compañía", "Categoría"],
            values="Cantidad",
            title="Distribución País → Compañía → Categoría",
            color_discrete_sequence=["#111827", "#334155", "#475569", "#64748b", "#94a3b8", "#cbd5e1"],
        )
        fig1.update_traces(textinfo="label+percent parent")
        fig1.update_layout(**layout)
        charts.append({"title": "Distribución geográfica", "figure_json": fig1.to_json()})

    estado_df = pd.DataFrame([{"Categoría": a.get("category") or "Sin cat", "Estado": a.get("status") or "Sin estado"} for a in working])
    if request.get("show_stock") or show_all:
        est_count = estado_df.groupby(["Categoría", "Estado"]).size().reset_index(name="Cantidad")
        fig2 = px.bar(
            est_count,
            x="Categoría",
            y="Cantidad",
            color="Estado",
            barmode="stack",
            title="Estado por Categoría de activo",
            color_discrete_map=color_map,
            text_auto=True,
        )
        fig2.update_layout(**layout, xaxis=dict(gridcolor="rgba(0,0,0,0.06)"), yaxis=dict(gridcolor="rgba(0,0,0,0.06)"))
        charts.append({"title": "Estado por categoría", "figure_json": fig2.to_json()})

    quality = {
        "Serial": sum(1 for a in working if get_serial_value(a)),
        "Hostname": sum(1 for a in working if get_hostname_value(a)),
        "Modelo": sum(1 for a in working if a.get("model")),
        "Costo": sum(1 for a in working if parse_cost(str(a.get("purchase_price", "")))),
        "País": sum(1 for a in working if a.get("country") and a.get("country") != "Sin país"),
        "Garantía": sum(1 for a in working if a.get("warranty_date")),
    }
    if request.get("show_quality") or show_all:
        q_df = pd.DataFrame([{"Campo": k, "Score": round(v / max(total, 1) * 100, 1)} for k, v in quality.items()]).sort_values("Score")
        fig3 = px.bar(
            q_df,
            x="Score",
            y="Campo",
            orientation="h",
            title="Data Quality Score (%)",
            text="Score",
            color="Score",
            color_continuous_scale=["#e7e5e4", "#94a3b8", "#0f766e"],
            range_color=[0, 100],
        )
        fig3.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
        fig3.update_layout(**layout, xaxis=dict(range=[0, 115]), coloraxis_showscale=False)
        charts.append({"title": "Data Quality", "figure_json": fig3.to_json()})

    costo_rows = [(a.get("country") or "Sin país", parse_cost(str(a.get("purchase_price", "")))) for a in working]
    costo_rows = [(p, c) for p, c in costo_rows if c > 0]
    if costo_rows and (request.get("show_spend") or show_all):
        costo_por_pais: dict[str, float] = {}
        for pais, amount in costo_rows:
            costo_por_pais[pais] = costo_por_pais.get(pais, 0) + amount
        cp_df = pd.DataFrame([{"País": k, "Inversión": round(v, 0)} for k, v in costo_por_pais.items()]).sort_values("Inversión", ascending=False)
        fig4 = px.bar(cp_df, x="País", y="Inversión", title="Inversión total por País", text_auto=".2s", color="Inversión", color_continuous_scale=["#e7e5e4", "#334155"])
        fig4.update_layout(**layout, coloraxis_showscale=False)
        charts.append({"title": "Inversión por país", "figure_json": fig4.to_json()})

    return encode_chat_payload(kpi_text, charts)


def run_nl_coverage_test(assets: list[dict[str, Any]]) -> dict[str, Any]:
    base_questions = [
        "¿Cuántas laptops tenemos en stock nuevo?",
        "¿Qué portátiles están en stock usado?",
        "¿Cuántas laptops están en uso?",
        "¿Qué laptop tiene asignada Juan Pérez?",
        "¿Cuál es el hostname del equipo de Juan Pérez?",
        "¿A qué usuario pertenece la laptop con hostname WKSCO091L?",
        "¿Cuántas laptops hay en Colombia?",
        "¿Cuántas laptops de Bancar ARG están en stock nuevo?",
        "¿Qué versión de SO es la más común en Bancar ARG?",
        "¿Cuál es el costo total por país?",
    ]
    questions = []
    for i in range(10):
        questions.extend(base_questions)
    ok = 0
    failures = []
    for q in questions:
        try:
            out = answer_inventory_question(assets, q)
            if isinstance(out, str) and out.strip():
                ok += 1
            else:
                failures.append({"q": q, "error": "respuesta vacía"})
        except Exception as exc:
            failures.append({"q": q, "error": str(exc)})
    return {"total": len(questions), "ok": ok, "failures": failures[:10]}


def ensure_session_state() -> None:
    """Inicializa el estado de sesión de Streamlit."""
    st.session_state.setdefault("assets", [])
    st.session_state.setdefault("last_sync", None)
    st.session_state.setdefault("chat_history", [])
    st.session_state.setdefault("openai_history", [])
    st.session_state.setdefault("last_error", "")
    st.session_state.setdefault("error_log", [])
    st.session_state.setdefault("aql_input", "")
    st.session_state.setdefault("schema_bridge", {})
    st.session_state.setdefault("critical_threshold", 10)
    st.session_state.setdefault("use_ai_compact", True)
    st.session_state.setdefault("auto_clear_after_action", True)
    st.session_state.setdefault("last_action_result", "")
    st.session_state.setdefault("pending_action", None)
    st.session_state.setdefault("theme_mode", "Oscuro ")
    st.session_state.setdefault("insights_prompt", "")
    st.session_state.setdefault("cache_ttl_minutes", 10)
    st.session_state.setdefault("cache_hash", "")
    st.session_state.setdefault("cache_expiry", None)
    st.session_state.setdefault("action_log", [])
    st.session_state.setdefault("movimientos", [])
    st.session_state.setdefault("reglas_guardadas", [])
    st.session_state.setdefault("last_chat_result_count", 0)
    st.session_state.setdefault("global_filter_countries", [])
    st.session_state.setdefault("global_filter_companies", [])
    st.session_state.setdefault("anomaly_report", {})
    st.session_state.setdefault("last_load_seconds", 0.0)
    st.session_state.setdefault("last_aql_executed", "")
    st.session_state.setdefault("auto_assign_rules", [])
    st.session_state.setdefault("auto_assign_log", [])
    st.session_state.setdefault("auto_assign_rr_idx", {})
    st.session_state.setdefault("scheduler_running", False)
    st.session_state.setdefault("scheduler_instance", None)
    st.session_state.setdefault("discovered_type_ids", [])
    st.session_state.setdefault("all_schema_type_ids", [])
    st.session_state.setdefault("type_discovery_source", "fallback")
    st.session_state.setdefault("type_discovery_error", "")
    st.session_state.setdefault("last_base_records_count", 0)
    st.session_state.setdefault("last_segmented_records_count", 0)
    st.session_state.setdefault("last_bruteforce_records_count", 0)
    st.session_state.setdefault("last_type_scan_checked", 0)
    st.session_state.setdefault("last_type_scan_hits", 0)
    st.session_state.setdefault("auto_reset_empty_once", False)
    st.session_state.setdefault("last_dashboard_prompt", "")
    st.session_state.setdefault("last_dashboard_text", "")
    st.session_state.setdefault("last_dashboard_charts", [])
    st.session_state.setdefault("last_dashboard_updated_at", "")
    if not st.session_state["movimientos"]:
        movement_path = MOVEMENTS_FILE
        if movement_path.exists():
            rows: list[dict[str, Any]] = []
            try:
                for line in movement_path.read_text(encoding="utf-8", errors="ignore").splitlines():
                    raw = line.strip()
                    if not raw:
                        continue
                    try:
                        rows.append(json.loads(raw))
                    except json.JSONDecodeError:
                        continue
                st.session_state["movimientos"] = rows[-500:]
            except OSError:
                pass
    if not st.session_state["reglas_guardadas"]:
        reglas_path = RULES_FILE
        if reglas_path.exists():
            try:
                data = json.loads(reglas_path.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    st.session_state["reglas_guardadas"] = data
            except (json.JSONDecodeError, OSError):
                pass
    if not st.session_state["auto_assign_rules"]:
        st.session_state["auto_assign_rules"] = load_auto_assign_rules()
    if not st.session_state["auto_assign_log"]:
        st.session_state["auto_assign_log"] = load_auto_assign_log()


def refresh_assets(config: AppConfig, aql_query: str = "", *, force_live: bool = False) -> None:
    """Refresca inventario desde Jira aplicando caché configurable."""
    debug_log(f"refresh_assets:start aql={aql_query!r} force_live={force_live}")
    with st.spinner("Consultando Jira Assets de Uala..."):
        cached_discovered = normalize_type_id_list(st.session_state.get("discovered_type_ids") or [])
        if not cached_discovered:
            st.session_state["discovered_type_ids"] = normalize_type_id_list(KNOWN_OBJECT_TYPE_IDS)
            st.session_state["type_discovery_source"] = "fallback"
            st.session_state["type_discovery_error"] = ""

        cached_all_schema = normalize_type_id_list(st.session_state.get("all_schema_type_ids") or [])
        if not cached_all_schema:
            st.session_state["all_schema_type_ids"] = get_schema_scan_type_ids(
                all_schema_type_ids=KNOWN_OBJECT_TYPE_IDS,
                discovered_type_ids=st.session_state.get("discovered_type_ids") or KNOWN_OBJECT_TYPE_IDS,
            )
        ttl = int(st.session_state.get("cache_ttl_minutes", 10))
        t0 = time.perf_counter()
        st.session_state.assets = cached_fetch_assets(config, aql_query, ttl, force_live=force_live)
        st.session_state.last_load_seconds = round(time.perf_counter() - t0, 3)
        st.session_state.last_aql_executed = build_primary_aql(
            aql_query,
            type_ids=get_active_hardware_type_ids(st.session_state.get("discovered_type_ids") or KNOWN_OBJECT_TYPE_IDS),
        )
        st.session_state.last_sync = datetime.now()
        st.session_state.anomaly_report = run_anomaly_detection(st.session_state.assets)
        st.session_state.last_error = ""
    debug_log(
        f"refresh_assets:end assets={len(st.session_state.get('assets', []))} "
        f"seconds={st.session_state.get('last_load_seconds', 0.0)}"
    )


def apply_theme() -> None:
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500&display=swap');
    :root {
        --uala-bg: #f5f5f4;
        --uala-surface: #ffffff;
        --uala-surface-soft: #fafaf9;
        --uala-border: #e7e5e4;
        --uala-border-strong: #d6d3d1;
        --uala-text: #111827;
        --uala-text-muted: #6b7280;
        --uala-text-soft: #9ca3af;
        --uala-accent: #1f2937;
        --uala-accent-soft: #475569;
        --uala-success: #0f766e;
        --uala-warning: #a16207;
        --uala-danger: #991b1b;
    }
    [data-testid="stSidebar"] { display: none !important; }
    [data-testid="collapsedControl"] { display: none !important; }
    .stApp > header { display: none !important; }
    .stApp {
        background: var(--uala-bg);
        color: var(--uala-text);
        font-family: 'Inter', var(--font-sans);
    }
    [data-testid="stAppViewContainer"] {
        background: var(--uala-bg);
    }
    [data-testid="stAppViewContainer"] * {
        color: inherit;
    }
    .uala-topbar {
        display: flex;
        align-items: center;
        justify-content: space-between;
        background: var(--uala-surface);
        border: 1px solid var(--uala-border);
        border-radius: 14px;
        padding: 0 20px;
        min-height: 56px;
        margin-bottom: 12px;
        gap: 12px;
        box-shadow: 0 1px 2px rgba(17, 24, 39, 0.04);
    }
    .uala-brand { display: flex; align-items: center; gap: 10px; }
    .uala-brand-icon {
        width: 32px; height: 32px; border-radius: 10px;
        background: var(--uala-accent);
        display: flex; align-items: center; justify-content: center;
        font-size: 13px; font-weight: 500; color: white;
    }
    .uala-brand-name { font-size: 14px; font-weight: 600; color: var(--uala-text); }
    .uala-brand-sub { font-size: 11px; color: var(--uala-text-muted); }
    .uala-nav { display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }
    .uala-topbar-right { display: flex; align-items: center; gap: 8px; }
    .uala-sync-badge {
        font-size: 10px; color: var(--uala-text-muted);
        background: var(--uala-surface-soft);
        border: 1px solid var(--uala-border);
        border-radius: 999px; padding: 5px 10px;
        display: flex; align-items: center; gap: 5px;
    }
    .uala-sync-dot { width: 6px; height: 6px; border-radius: 50%; background: var(--uala-success); }
    .uala-sync-dot.warn { background: var(--uala-warning); }
    .uala-alert-pill {
        font-size: 10px; padding: 5px 10px; border-radius: 999px;
        background: #fef2f2; border: 1px solid #fecaca; color: var(--uala-danger);
        display: flex; align-items: center; gap: 4px;
    }
    .uala-alert-pill.ok { background: #f5f5f4; border-color: var(--uala-border); color: var(--uala-text-muted); }
    .uala-filterbar {
        display: flex; align-items: center; gap: 12px; flex-wrap: wrap;
        background: var(--uala-surface);
        border: 1px solid var(--uala-border);
        border-radius: 14px; padding: 8px 20px;
        margin-bottom: 12px;
        box-shadow: 0 1px 2px rgba(17, 24, 39, 0.04);
    }
    .uala-hero {
        background: var(--uala-surface);
        border: 1px solid var(--uala-border);
        border-radius: 16px;
        padding: 1.15rem 1.35rem;
        margin-bottom: 1rem;
        box-shadow: 0 1px 2px rgba(17, 24, 39, 0.04);
    }
    .uala-kicker { font-size: 0.72rem; letter-spacing: .12em; text-transform: uppercase; color: var(--uala-text-muted); margin-bottom: .45rem; }
    .uala-title { font-size: 2rem; line-height: 1.05; color: var(--uala-text); margin-bottom: .35rem; font-weight: 600; }
    .uala-sub { color: var(--uala-text-muted); font-size: .95rem; }
    div[data-testid="metric-container"] {
        background: var(--uala-surface) !important;
        border: 1px solid var(--uala-border) !important;
        border-radius: 14px !important;
        padding: 12px 16px !important;
        box-shadow: none !important;
    }
    div[data-testid="metric-container"] label {
        font-size: 11px !important;
        color: var(--uala-text-muted) !important;
        text-transform: uppercase;
        letter-spacing: 0.04em;
    }
    div[data-testid="metric-container"] [data-testid="stMetricValue"] {
        font-size: 24px !important;
        font-weight: 600 !important;
        color: var(--uala-text) !important;
        text-shadow: none !important;
    }
    [data-testid="stChatMessage"] {
        background: var(--uala-surface) !important;
        border: 1px solid var(--uala-border) !important;
        border-radius: 14px !important;
        box-shadow: none !important;
        max-width: 100% !important;
    }
    .stTabs [data-baseweb="tab-list"] { gap: 6px; border-bottom: 1px solid var(--uala-border); }
    .stTabs [data-baseweb="tab"] {
        border-radius: 10px 10px 0 0 !important;
        border: 1px solid transparent !important;
        background: transparent !important;
        color: var(--uala-text-muted) !important;
        font-size: 13px;
        padding: 8px 16px;
    }
    .stTabs [aria-selected="true"] {
        background: var(--uala-surface) !important;
        border-color: var(--uala-border) !important;
        border-bottom-color: var(--uala-surface) !important;
        color: var(--uala-text) !important;
        font-weight: 600 !important;
    }
    .stButton > button {
        border-radius: 12px !important;
        border: 1px solid var(--uala-border-strong) !important;
        background: var(--uala-surface) !important;
        color: var(--uala-text) !important;
        font-size: 13px !important;
        font-weight: 500 !important;
        box-shadow: none !important;
        padding: 8px 16px !important;
    }
    .stButton > button:hover {
        background: var(--uala-surface-soft) !important;
        border-color: var(--uala-accent-soft) !important;
    }
    .stButton > button[kind="primary"] {
        background: var(--uala-accent) !important;
        color: white !important;
        border-color: var(--uala-accent) !important;
    }
    .stButton > button p { color: inherit !important; }
    .stTextInput > div > div > input,
    .stTextArea textarea,
    [data-testid="stChatInput"] textarea,
    [data-baseweb="input"] input,
    [data-testid="stDateInput"] input {
        background: var(--uala-surface) !important;
        color: var(--uala-text) !important;
        border-radius: 12px !important;
    }
    .stTextInput > div > div,
    .stTextArea > div > div,
    [data-baseweb="input"] > div,
    [data-testid="stDateInput"] > div > div,
    div[data-baseweb="select"] > div {
        background: var(--uala-surface) !important;
        border: 1px solid var(--uala-border-strong) !important;
        border-radius: 12px !important;
        box-shadow: none !important;
        color: var(--uala-text) !important;
    }
    div[data-baseweb="select"] * {
        color: var(--uala-text) !important;
    }
    .stMultiSelect [data-baseweb="tag"] {
        background: var(--uala-surface-soft) !important;
        border: 1px solid var(--uala-border) !important;
        color: var(--uala-text) !important;
        border-radius: 999px !important;
    }
    input::placeholder, textarea::placeholder {
        color: var(--uala-text-soft) !important;
        opacity: 1 !important;
    }
    .uala-confirm-bar {
        background: #fffbeb;
        border: 1px solid #fcd34d;
        border-radius: 12px;
        padding: 10px 16px;
        display: flex; align-items: center; justify-content: space-between;
        margin: 8px 0;
    }
    .uala-confirm-text { font-size: 13px; color: #92400e; }
    [data-testid="stDataFrame"] {
        border: 1px solid var(--uala-border) !important;
        border-radius: 14px !important;
        overflow: hidden;
        background: var(--uala-surface) !important;
    }
    .stAlert {
        border-radius: 12px !important;
        border: 1px solid var(--uala-border) !important;
    }
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    </style>
    """, unsafe_allow_html=True)


# ── 10. UI STREAMLIT (páginas, sidebar, branding) ─────────────────────────
def render_branding(config: AppConfig) -> None:
    workspace_label = escape_html_text(config.workspace_id or "no configurado")
    st.markdown(
        f"""
        <div class="uala-hero">
            <div class="uala-kicker">Uala Asset Control</div>
            <div class="uala-title">Inventario Uala (Jira Assets)</div>
            <div class="uala-sub">Workspace: {workspace_label} · Esquema: {SCHEMA_ID}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_setup_screen() -> None:
    st.title("Configuración inicial")
    st.warning("Faltan variables obligatorias. Configurá el entorno antes de usar la app.")
    with st.expander("Variables necesarias"):
        st.code(
            "JIRA_EMAIL=tu-email@bancar.com\n"
            "JIRA_API_TOKEN=tu-api-token\n"
            "ASSETS_WORKSPACE_ID=tu-workspace-id\n"
            "JIRA_SITE=https://bancar.atlassian.net\n"
            "OPENAI_API_KEY=sk-..."
        )
    st.info("En Streamlit Cloud: Settings -> Secrets.")
    st.info("En local: crear archivo .env con esas variables.")


def render_topbar(config: AppConfig, current_page: str, assets: list[dict[str, Any]]) -> str:
    pages = ["Chat", "Activos", "Insights", "Auditoría", "Movimientos", "Scripts", "Extra"]
    last_sync = st.session_state.get("last_sync")
    sync_text = last_sync.strftime("%H:%M:%S") if last_sync else "sin sync"
    anomaly = st.session_state.get("anomaly_report", {})
    total_anomaly = int(anomaly.get("total", 0))
    garantia_vencida = int(anomaly.get("garantia_vencida_en_uso", 0))
    alert_class = "ok" if total_anomaly == 0 else ""
    alert_text = "Sin anomalías" if total_anomaly == 0 else f"{total_anomaly} anomalías"
    workspace_label = escape_html_text(config.workspace_id[:8] if config.workspace_id else "no conf.")
    alert_label = escape_html_text(alert_text)
    garantia_label = escape_html_text(f" · {garantia_vencida} garantías" if total_anomaly else "")

    st.markdown(
        f"""
        <div class="uala-topbar">
            <div class="uala-brand">
                <div class="uala-brand-icon">U</div>
                <div>
                    <div class="uala-brand-name">Uala Asset Control</div>
                    <div class="uala-brand-sub">Workspace {workspace_label} · Esquema {SCHEMA_ID}</div>
                </div>
            </div>
            <div class="uala-topbar-right">
                <div class="uala-alert-pill {alert_class}">{alert_label}{garantia_label}</div>
                <div class="uala-sync-badge"><div class="uala-sync-dot {'' if last_sync else 'warn'}"></div>{sync_text}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    nav_cols = st.columns(len(pages))
    for idx, page in enumerate(pages):
        label = f"• {page}" if page == current_page else page
        if nav_cols[idx].button(label, key=f"top_nav_{page}", use_container_width=True):
            st.query_params["page"] = page
            st.rerun()
    return current_page


def render_filterbar(config: AppConfig) -> None:
    assets_raw = st.session_state.get("assets", [])
    countries = sorted({str(a.get("country") or "Sin país") for a in assets_raw if a.get("country") and a.get("country") != "Sin país"})
    companies = sorted({str(a.get("company") or "") for a in assets_raw if a.get("company")})
    persisted_countries = normalize_type_id_list([])
    persisted_countries = [value for value in (st.session_state.get("global_filter_countries", []) or []) if value in countries]
    persisted_companies = [value for value in (st.session_state.get("global_filter_companies", []) or []) if value in companies]
    if persisted_countries != list(st.session_state.get("global_filter_countries", []) or []):
        st.session_state.global_filter_countries = persisted_countries
    if persisted_companies != list(st.session_state.get("global_filter_companies", []) or []):
        st.session_state.global_filter_companies = persisted_companies
    st.markdown('<div class="uala-filterbar"></div>', unsafe_allow_html=True)
    col_filters, col_companies, col_aql, col_actions = st.columns([3, 3, 2, 1])
    with col_filters:
        selected_countries = st.multiselect(
            "País",
            options=countries,
            default=persisted_countries,
            label_visibility="collapsed",
            placeholder="Filtrar por país...",
        )
        st.session_state.global_filter_countries = selected_countries
    with col_companies:
        selected_companies = st.multiselect(
            "Compañía",
            options=companies,
            default=persisted_companies,
            label_visibility="collapsed",
            placeholder="Filtrar por compañía...",
        )
        st.session_state.global_filter_companies = selected_companies
    with col_aql:
        aql_val = st.text_input(
            "AQL",
            value=st.session_state.get("aql_input", ""),
            placeholder="AQL adicional...",
            label_visibility="collapsed",
        )
        st.session_state.aql_input = aql_val.strip()
    with col_actions:
        col_ref, col_clr = st.columns(2)
        if col_ref.button("↻", help="Refrescar inventario", use_container_width=True):
            st.session_state.cache_hash = ""
            st.session_state.cache_expiry = None
            refresh_assets(config, st.session_state.aql_input, force_live=True)
            st.rerun()
        if col_clr.button("✕", help="Limpiar filtros", use_container_width=True):
            st.session_state.global_filter_countries = []
            st.session_state.global_filter_companies = []
            st.session_state.aql_input = ""
            st.rerun()
    raw_count = len(assets_raw)
    filtered_count = len(apply_global_filter(assets_raw))
    if raw_count != filtered_count:
        st.caption(f"Mostrando {filtered_count} de {raw_count} activos · filtro activo")
    st.divider()


def render_insights(assets: list[dict[str, Any]]) -> None:
    st.subheader("Insights")
    nl_prompt = st.text_input(
        "Filtrar dashboard por lenguaje natural",
        value=st.session_state.get("insights_prompt", ""),
        placeholder="Ej: activos de Bancar ARG con garantía por vencer",
    )
    st.session_state.insights_prompt = nl_prompt
    dashboard_spec = parse_nl_dashboard_request(nl_prompt) if nl_prompt else {}
    working = apply_filters(assets, dashboard_spec.get("filters", {})) if dashboard_spec else assets
    total = len(working)
    if total == 0:
        st.warning("No hay activos para los filtros actuales.")
        return

    in_use = sum(1 for a in working if normalize_text(a.get("status", "")) == "en uso")
    stock_nuevo = sum(1 for a in working if normalize_text(a.get("status", "")) == "stock nuevo")
    stock_usado = sum(1 for a in working if normalize_text(a.get("status", "")) == "stock usado")
    sin_asignar = sum(1 for a in working if not str(a.get("assigned_to") or "").strip())
    sin_serial = sum(1 for a in working if not str(get_serial_value(a)).strip())
    garantia_vencida = 0
    garantia_45 = 0
    today = datetime.now().date()
    for a in working:
        w = parse_date(str(a.get("warranty_date", "")).split("|")[0].strip())
        if not w:
            continue
        delta = (w.date() - today).days
        if delta < 0:
            garantia_vencida += 1
        elif delta <= 45:
            garantia_45 += 1
    costo_total = round(sum(parse_cost(str(a.get("purchase_price", ""))) for a in working), 2)
    dep = calculate_depreciation(working)

    r1c1, r1c2, r1c3, r1c4 = st.columns(4)
    r1c1.metric("Total activos", total)
    r1c2.metric("En uso", in_use, delta=f"{round(in_use / max(total, 1) * 100, 1)}%")
    r1c3.metric("Stock nuevo", stock_nuevo)
    r1c4.metric("Stock usado", stock_usado)

    r2c1, r2c2, r2c3, r2c4 = st.columns(4)
    r2c1.metric("Sin asignar", sin_asignar, delta="revisar" if sin_asignar > 0 else "ok", delta_color="inverse" if sin_asignar > 0 else "normal")
    r2c2.metric("Sin serial", sin_serial, delta=f"{round(sin_serial / max(total, 1) * 100, 1)}%", delta_color="inverse" if sin_serial > 0 else "normal")
    r2c3.metric("Garantía vencida", garantia_vencida, delta="urgente" if garantia_vencida > 0 else "ok", delta_color="inverse" if garantia_vencida > 0 else "normal")
    r2c4.metric("Garantía ≤45 días", garantia_45, delta="atención" if garantia_45 > 0 else "ok", delta_color="inverse" if garantia_45 > 0 else "normal")

    st.divider()
    if pd is None or px is None:
        st.info("Instalá `pandas` y `plotly` para gráficos.")
        return

    tab_geo, tab_estado, tab_quality, tab_financiero, tab_stock = st.tabs(["Geografía", "Estados", "Calidad de datos", "Financiero", "Stock crítico"])
    with tab_geo:
        geo_rows = [{"País": a.get("country") or "Sin país", "Compañía": a.get("company") or "Sin compañía"} for a in working]
        df_geo = pd.DataFrame(geo_rows)
        geo_count = df_geo.groupby(["País", "Compañía"]).size().reset_index(name="Cantidad")
        fig = px.bar(geo_count, x="País", y="Cantidad", color="Compañía", barmode="group", title="Activos por País y Compañía", text_auto=True, color_discrete_sequence=["#111827", "#334155", "#475569", "#64748b", "#94a3b8"])
        fig.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", font=dict(size=12), margin=dict(t=40, b=30, l=0, r=0), legend=dict(orientation="h", yanchor="bottom", y=1.02), xaxis=dict(gridcolor="rgba(0,0,0,0.06)"), yaxis=dict(gridcolor="rgba(0,0,0,0.06)"))
        st.plotly_chart(fig, use_container_width=True)

    with tab_estado:
        rows_es = [{"Categoría": a.get("category") or "Sin cat", "Estado": a.get("status") or "Sin estado"} for a in working]
        df_es = pd.DataFrame(rows_es)
        count_es = df_es.groupby(["Categoría", "Estado"]).size().reset_index(name="Cantidad")
        color_map = {"En uso": "#0f766e", "Stock nuevo": "#334155", "Stock usado": "#78716c", "Asignado al edificio": "#64748b", "Sin estado": "#cbd5e1"}
        fig2 = px.bar(count_es, x="Categoría", y="Cantidad", color="Estado", barmode="stack", title="Estado por Categoría", color_discrete_map=color_map, text_auto=True)
        fig2.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", margin=dict(t=40, b=30, l=0, r=0), xaxis=dict(gridcolor="rgba(0,0,0,0.06)"), yaxis=dict(gridcolor="rgba(0,0,0,0.06)"))
        st.plotly_chart(fig2, use_container_width=True)

    with tab_quality:
        quality_scores = {
            "Serial": sum(1 for a in working if get_serial_value(a)),
            "Hostname": sum(1 for a in working if get_hostname_value(a)),
            "Modelo": sum(1 for a in working if a.get("model")),
            "Estado válido": sum(1 for a in working if normalize_text(a.get("status", "")) in {normalize_text(v) for v in ESTADO_NORMALIZATION.values()}),
            "Costo cargado": sum(1 for a in working if parse_cost(str(a.get("purchase_price", "")))),
            "País asignado": sum(1 for a in working if a.get("country") and a.get("country") != "Sin país"),
            "Garantía cargada": sum(1 for a in working if a.get("warranty_date")),
            "Asignado/Stock ok": sum(1 for a in working if a.get("assigned_to") or "stock" in normalize_text(a.get("status", ""))),
        }
        df_q = pd.DataFrame([{"Campo": k, "Cobertura": round(v / max(total, 1) * 100, 1), "Faltantes": total - v} for k, v in quality_scores.items()]).sort_values("Cobertura")
        fig3 = px.bar(df_q, x="Cobertura", y="Campo", orientation="h", title="Cobertura de atributos clave (%)", text="Cobertura", color="Cobertura", color_continuous_scale=["#e7e5e4", "#94a3b8", "#0f766e"], range_color=[0, 100], custom_data=["Faltantes"])
        fig3.update_traces(texttemplate="%{text:.1f}%", textposition="outside", hovertemplate="<b>%{y}</b><br>Cobertura: %{x:.1f}%<br>Faltantes: %{customdata[0]}<extra></extra>")
        fig3.update_layout(xaxis=dict(range=[0, 115], gridcolor="rgba(0,0,0,0.06)"), yaxis=dict(gridcolor="rgba(0,0,0,0)"), coloraxis_showscale=False, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", margin=dict(t=40, b=20, l=0, r=60))
        st.plotly_chart(fig3, use_container_width=True)
        worst = df_q[df_q["Cobertura"] < 80].sort_values("Cobertura")
        if not worst.empty:
            st.caption(f"Campos con cobertura menor al 80% — {len(worst)} campos")
            st.dataframe(worst[["Campo", "Cobertura", "Faltantes"]], use_container_width=True, hide_index=True)

    with tab_financiero:
        costo_por_pais: dict[str, float] = {}
        costo_por_cat: dict[str, float] = {}
        for a in working:
            amount = parse_cost(str(a.get("purchase_price", "")))
            if amount <= 0:
                continue
            pais = a.get("country") or "Sin país"
            cat = a.get("category") or "Sin categoría"
            costo_por_pais[pais] = costo_por_pais.get(pais, 0) + amount
            costo_por_cat[cat] = costo_por_cat.get(cat, 0) + amount
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            if costo_por_pais:
                df_fp = pd.DataFrame([{"País": k, "Costo": round(v, 2)} for k, v in costo_por_pais.items()]).sort_values("Costo", ascending=False)
                fig_fp = px.bar(df_fp, x="País", y="Costo", title="Inversión por País", text_auto=".2s", color="Costo", color_continuous_scale=["#e7e5e4", "#334155"])
                fig_fp.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", coloraxis_showscale=False, margin=dict(t=40, b=20))
                st.plotly_chart(fig_fp, use_container_width=True)
        with col_f2:
            if costo_por_cat:
                df_fc = pd.DataFrame([{"Categoría": k, "Costo": round(v, 2)} for k, v in costo_por_cat.items()]).sort_values("Costo", ascending=False)
                fig_fc = px.pie(df_fc, names="Categoría", values="Costo", title="Inversión por Categoría", color_discrete_sequence=["#111827", "#334155", "#475569", "#64748b", "#94a3b8", "#cbd5e1"])
                fig_fc.update_traces(textinfo="label+percent")
                fig_fc.update_layout(paper_bgcolor="rgba(0,0,0,0)", margin=dict(t=40, b=20))
                st.plotly_chart(fig_fc, use_container_width=True)
        f1, f2, f3 = st.columns(3)
        f1.metric("Costo total inventario", f"${costo_total:,.0f}")
        con_precio = [a for a in working if parse_cost(str(a.get("purchase_price", ""))) > 0]
        avg = round(costo_total / max(len(con_precio), 1), 2)
        f2.metric("Costo promedio", f"${avg:,.0f}")
        f3.metric("Valor contable actual", f"${dep.get('total_book_value', 0):,.0f}", delta=f"-${round(costo_total - dep.get('total_book_value', 0), 0):,.0f} depreciado")

    with tab_stock:
        st.markdown("**Stock disponible por categoría**")
        critical_threshold = int(st.session_state.get("critical_threshold", 10))
        cat_stock: dict[str, int] = {}
        for a in working:
            if normalize_text(a.get("status", "")) not in {normalize_text("stock nuevo"), normalize_text("stock usado")}:
                continue
            cat = a.get("category") or "Sin categoría"
            cat_stock[cat] = cat_stock.get(cat, 0) + 1
        if not cat_stock:
            st.info("No hay stock disponible para los filtros actuales.")
        else:
            for cat, qty in sorted(cat_stock.items(), key=lambda x: x[1]):
                if qty == 0:
                    color, label = "#991b1b", "CRÍTICO"
                elif qty <= critical_threshold:
                    color, label = "#a16207", "BAJO"
                else:
                    color, label = "#0f766e", "OK"
                c1, c2, c3 = st.columns([3, 1, 1])
                c1.write(cat)
                c2.markdown(f"<span style='color:{color};font-weight:500;font-size:12px;'>{label}</span>", unsafe_allow_html=True)
                c3.write(str(qty))


def colored_metric(container: Any, label: str, value: Any, alert_threshold: int | float | None = None, inverse: bool = False) -> None:
    if alert_threshold is None:
        container.metric(label, value)
        return
    numeric_val = value if isinstance(value, (int, float)) else 0
    if inverse:
        delta_color = "inverse" if numeric_val > alert_threshold else "normal"
        delta = "⚠️ revisar" if numeric_val > alert_threshold else "✅ ok"
    else:
        delta_color = "normal" if numeric_val >= alert_threshold else "inverse"
        delta = "✅ ok" if numeric_val >= alert_threshold else "⚠️ bajo"
    container.metric(label, value, delta=delta, delta_color=delta_color)


def build_contextual_suggestions(assets: list[dict[str, Any]], anomaly: dict[str, Any]) -> list[str]:
    suggestions = ["Resumen ejecutivo del inventario"]
    if anomaly.get("en_uso_sin_asignado", 0) > 0:
        suggestions.append(f"Mostrame los {anomaly['en_uso_sin_asignado']} equipos en uso sin usuario asignado")
    if anomaly.get("serial_duplicado", 0) > 0:
        suggestions.append(f"Hay {anomaly['serial_duplicado']} seriales duplicados — mostrálos")
    if anomaly.get("garantia_vencida_en_uso", 0) > 0:
        suggestions.append(f"Equipos en uso con garantía vencida ({anomaly['garantia_vencida_en_uso']})")
    sin_serial = sum(1 for a in assets if not get_serial_value(a))
    if sin_serial > 0:
        suggestions.append(f"Mostrame los {sin_serial} activos sin número de serie")
    countries = sorted({a.get("country") for a in assets if a.get("country") and a.get("country") != "Sin país"})
    for country in countries[:2]:
        suggestions.append(f"Stock disponible en {country}")
    con_costo = sum(1 for a in assets if parse_cost(str(a.get("purchase_price", ""))) > 0)
    if con_costo > 10:
        suggestions.append("¿Cuál es el costo total del inventario por país?")
    seen: set[str] = set()
    unique: list[str] = []
    for item in suggestions:
        if item not in seen:
            seen.add(item)
            unique.append(item)
    return unique[:6]


def _build_contextual_suggestions(assets: list[dict[str, Any]], anomaly: dict[str, Any]) -> list[str]:
    return build_contextual_suggestions(assets, anomaly)


def _execute_pending_action(config: AppConfig, assets: list[dict[str, Any]], pending_action: dict[str, Any] | None = None) -> tuple[bool, str]:
    pending_action = pending_action or st.session_state.get("pending_action")
    if not pending_action:
        return False, "No hay acción pendiente."

    action = pending_action.get("action")
    ok = False
    answer = "Acción no válida."
    if action == "assign":
        ok, answer = assign_asset(config, assets, pending_action["identifier"], pending_action["assignee"])
    elif action == "unassign":
        ok, answer = unassign_asset(config, assets, pending_action["identifier"], pending_action.get("target_status", "Stock usado"))
    elif action == "status":
        ok, answer = update_status(config, assets, pending_action["identifier"], pending_action["new_status"])
    elif action == "bulk":
        updated, errors = bulk_update_location(config, assets, pending_action["identifiers"], pending_action["company"], pending_action["country"])
        ok = updated > 0
        answer = f"Bulk update aplicado. Actualizados: {updated}/{len(pending_action['identifiers'])}."
        if errors:
            answer += " " + " | ".join(errors[:5])
    elif action == "regla":
        regla = pending_action["regla"]
        rule_obj = ReglaNormalizacion(**regla) if isinstance(regla, dict) else regla
        updated, errors = aplicar_regla(config, assets, rule_obj, dry_run=False)
        ok = updated > 0
        answer = f"Regla aplicada. Afectados: {updated}."
        if errors:
            answer += " " + " | ".join(errors[:5])

    st.session_state.pending_action = None
    return ok, answer


def _render_pending_action_block(config: AppConfig, all_assets: list[dict[str, Any]]) -> None:
    pending = st.session_state.get("pending_action")
    if not pending:
        return
    pending_summary = escape_html_text(pending.get("summary", ""))

    st.markdown(
        f"""
        <div class="uala-confirm-bar">
            <div class="uala-confirm-text">
                ⚡ <strong>Acción pendiente:</strong> {pending_summary}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    col_ok, col_no, _ = st.columns([1, 1, 4])
    if col_ok.button("✅ Confirmar", key="btn_confirm_pending", type="primary", use_container_width=True):
        ok, msg = _execute_pending_action(config, all_assets, pending)
        st.session_state.pending_action = None
        if ok:
            refresh_assets(config, st.session_state.get("aql_input", ""), force_live=True)
            st.success(f"✅ {msg}")
        else:
            st.error(f"❌ {msg}")
    if col_no.button("✕ Cancelar", key="btn_cancel_pending", use_container_width=True):
        st.session_state.pending_action = None
        st.session_state.chat_history.append({"role": "assistant", "content": "Acción cancelada."})
        st.rerun()


def _process_chat_prompt(config: AppConfig, all_assets: list[dict[str, Any]], filtered_assets: list[dict[str, Any]], prompt: str) -> None:
    st.session_state.chat_history.append({"role": "user", "content": prompt})
    regla = parse_normalization_rule_from_prompt(prompt)
    if regla:
        affected, _ = aplicar_regla(config, all_assets, regla, dry_run=True)
        sample = [str(a.get("jira_key") or a.get("hostname") or a.get("name", "")) for a in all_assets if evaluar_regla(a, regla)][:5]
        preview = (
            f"**Regla detectada**\n\n"
            f"- Condición: `{regla.campo_condicion}` {regla.operador} `{regla.valor_condicion}`\n"
            f"- Acción: `{regla.campo_a_modificar}` → `{regla.valor_nuevo}`\n"
            f"- **Activos afectados: {affected}**\n"
            f"- Muestra: {', '.join(sample) if sample else 'sin coincidencias'}\n\n"
            f"_Usá el botón Confirmar para aplicar._"
        )
        st.session_state.pending_action = {
            "action": "regla",
            "regla": regla.__dict__,
            "summary": f"'{regla.descripcion}' — {affected} activos afectados",
        }
        st.session_state.chat_history.append({"role": "assistant", "content": preview})
        push_openai_history(prompt, preview)
        return

    assignment = parse_assignment_action(prompt)
    if assignment:
        identifier, assignee = assignment
        st.session_state.pending_action = {"action": "assign", "identifier": identifier, "assignee": assignee, "summary": f"Asignar `{identifier}` a `{assignee}`"}
        st.session_state.chat_history.append({"role": "assistant", "content": f"Asignación detectada: `{identifier}` → `{assignee}`. Confirmá con el botón."})
        return
    unassign = parse_unassign_action(prompt)
    if unassign:
        identifier, target_status = unassign
        st.session_state.pending_action = {"action": "unassign", "identifier": identifier, "target_status": target_status, "summary": f"Desasignar `{identifier}` → estado `{target_status}`"}
        st.session_state.chat_history.append({"role": "assistant", "content": f"Desasignación detectada: `{identifier}`. Confirmá con el botón."})
        return
    status_change = parse_status_change_action(prompt)
    if status_change:
        identifier, new_status = status_change
        st.session_state.pending_action = {"action": "status", "identifier": identifier, "new_status": new_status, "summary": f"Cambiar estado de `{identifier}` → `{new_status}`"}
        st.session_state.chat_history.append({"role": "assistant", "content": f"Cambio de estado detectado: `{identifier}` → `{new_status}`. Confirmá con el botón."})
        return
    bulk = parse_bulk_location_action(prompt)
    if bulk:
        identifiers, company, country = bulk
        st.session_state.pending_action = {"action": "bulk", "identifiers": identifiers, "company": company, "country": country, "summary": f"Bulk update {len(identifiers)} activos → {country}/{company}"}
        st.session_state.chat_history.append({"role": "assistant", "content": f"Actualización masiva detectada: {len(identifiers)} activos → {country}/{company}. Confirmá con el botón."})
        return

    dashboard_req = parse_nl_dashboard_request(prompt)
    if dashboard_req:
        response = answer_inventory_question(filtered_assets, prompt)
    else:
        suggested_aql, notes = build_aql_from_prompt(prompt)
        if suggested_aql:
            try:
                fetched = fetch_assets(config, suggested_aql)
                use_ai = st.session_state.get("use_ai_compact", True)
                answer = ai_compact_answer(config, prompt, fetched, notes, prefiltered=True) if use_ai and config.openai_api_key else answer_inventory_question(fetched, prompt)
                response = f"_AQL: `{suggested_aql}`_ · {len(fetched)} activos\n\n{answer}"
            except RuntimeError as exc:
                response = f"Falló AQL: {exc}\n\n{answer_inventory_question(filtered_assets, prompt)}"
        else:
            response = answer_inventory_question(filtered_assets, prompt)
    remember_dashboard_response(prompt, response)
    st.session_state.chat_history.append({"role": "assistant", "content": response})
    push_openai_history(prompt, decode_chat_payload(response)[0])


def render_chat_dashboard_panel() -> None:
    restore_dashboard_state_from_history()
    charts = st.session_state.get("last_dashboard_charts") or []
    if not charts:
        return
    prompt = str(st.session_state.get("last_dashboard_prompt") or "").strip()
    text = str(st.session_state.get("last_dashboard_text") or "").strip()
    updated = str(st.session_state.get("last_dashboard_updated_at") or "").strip()

    st.divider()
    head_col, clear_col = st.columns([6, 1])
    head_col.markdown("### Dashboard solicitado")
    if clear_col.button("Limpiar", key="clear_last_dashboard", use_container_width=True):
        st.session_state["last_dashboard_prompt"] = ""
        st.session_state["last_dashboard_text"] = ""
        st.session_state["last_dashboard_charts"] = []
        st.session_state["last_dashboard_updated_at"] = ""
        st.rerun()
    if prompt:
        st.caption(f"Consulta: {prompt}")
    if updated:
        st.caption(f"Última actualización: {updated}")
    intro, table_df = parse_chat_response_for_table(text)
    if intro:
        st.markdown(intro)
    if table_df is not None:
        st.dataframe(table_df, use_container_width=True, hide_index=True)
    for chart in charts:
        if pio is None:
            st.caption(chart.get("title", ""))
            continue
        try:
            fig = pio.from_json(chart.get("figure_json", ""))
            st.plotly_chart(fig, use_container_width=True)
        except Exception:
            st.caption(chart.get("title", ""))


# ── 9. EXPORTACIONES (Excel) ──────────────────────────────────────────────
def render_assets_page(assets: list[dict[str, Any]]) -> None:
    st.subheader("Listado de activos Uala")
    c1, c2, c3 = st.columns([1.5, 1, 1])

    with c1:
        search = st.text_input("Buscar", placeholder="Nombre, serial, asignado, Jira...")

    statuses = sorted({a.get("status") or "Sin estado" for a in assets})
    categories = sorted({a.get("category") or "Sin categoría" for a in assets})

    with c2:
        status_filter = st.selectbox("Estado", ["Todos"] + statuses)
    with c3:
        category_filter = st.selectbox("Categoría", ["Todas"] + categories)

    filtered = []
    term = normalize_text(search)
    terms = [normalize_text(x) for x in search.split(",") if normalize_text(x)]
    chip_html = "".join(
        [
            f"<span style='padding:4px 10px;border:1px solid #94a3b8;border-radius:999px;margin-right:6px'>{escape_html_text(t)}</span>"
            for t in terms
        ]
    )
    if chip_html:
        st.markdown(f"<div>Filtros activos: {chip_html}</div>", unsafe_allow_html=True)
    for asset in assets:
        if status_filter != "Todos" and (asset.get("status") or "Sin estado") != status_filter:
            continue
        if category_filter != "Todas" and (asset.get("category") or "Sin categoría") != category_filter:
            continue

        haystack = " | ".join([
            str(asset.get("name", "")),
            str(asset.get("serial_number", "")),
            str(asset.get("assigned_to", "")),
            str(asset.get("jira_key", "")),
            str(asset.get("model", "")),
            str(asset.get("country", "")),
            str(asset.get("company", "")),
        ])
        attr_blob = " | ".join([f"{k}:{v}" for k, v in (asset.get("attrs_by_name") or {}).items()])
        full_blob = normalize_text(f"{haystack} | {attr_blob}")
        if terms and not all(t in full_blob for t in terms):
            continue
        found_in = "principal"
        if terms and all(t in normalize_text(attr_blob) for t in terms):
            found_in = "attrs_by_name"
        row = dict(asset)
        row["_found_in"] = found_in
        filtered.append(row)

    st.caption(f"{len(filtered)} activos encontrados")
    table = [flatten_asset_for_display(a, include_all_attributes=True) for a in filtered]
    for idx, a in enumerate(filtered):
        table[idx]["Encontrado en"] = a.get("_found_in", "principal")
    if Workbook is not None and st.button("📥 Exportar a Excel", use_container_width=False):
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        header_fill = PatternFill(start_color="003262", end_color="003262", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary["A1"] = f"Exportado: {now}"
        total = len(filtered)
        en_uso = sum(1 for a in filtered if normalize_text(a.get("status")) == normalize_text("En uso"))
        stock_nuevo = sum(1 for a in filtered if normalize_text(a.get("status")) == normalize_text("Stock nuevo"))
        stock_usado = sum(1 for a in filtered if normalize_text(a.get("status")) == normalize_text("Stock usado"))
        sin_asignar = sum(1 for a in filtered if not str(a.get("assigned_to") or "").strip())
        resumen_rows = [
            ("Total activos", total),
            ("En uso", en_uso),
            ("Stock nuevo", stock_nuevo),
            ("Stock usado", stock_usado),
            ("Sin asignar", sin_asignar),
            ("Costo total", round(sum(parse_cost(str(a.get("purchase_price", ""))) for a in filtered), 2)),
        ]
        for idx, (k, v) in enumerate(resumen_rows, start=3):
            ws_summary[f"A{idx}"] = k
            ws_summary[f"B{idx}"] = v
        by_country = Counter((a.get("country") or "Sin país") for a in filtered)
        by_company = Counter((a.get("company") or "Sin compañía") for a in filtered)
        start_country = len(resumen_rows) + 5
        ws_summary[f"A{start_country}"] = "Por país"
        ws_summary[f"A{start_country}"].font = header_font
        ws_summary[f"A{start_country}"].fill = header_fill
        row_ptr = start_country + 1
        for k, v in by_country.items():
            ws_summary[f"A{row_ptr}"] = k
            ws_summary[f"B{row_ptr}"] = v
            row_ptr += 1
        ws_summary[f"D{start_country}"] = "Por compañía"
        ws_summary[f"D{start_country}"].font = header_font
        ws_summary[f"D{start_country}"].fill = header_fill
        row_ptr = start_country + 1
        for k, v in by_company.items():
            ws_summary[f"D{row_ptr}"] = k
            ws_summary[f"E{row_ptr}"] = v
            row_ptr += 1

        by_type: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for a in filtered:
            by_type[str(a.get("object_type") or "Sin tipo")[:31]].append(a)
        for object_type, rows in by_type.items():
            ws = wb.create_sheet(title=object_type or "Sin tipo")
            records = [flatten_asset_for_display(r, include_all_attributes=True) for r in rows]
            headers = list(records[0].keys()) if records else []
            for c_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
            for r_idx, rec in enumerate(records, start=2):
                for c_idx, h in enumerate(headers, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=str(rec.get(h, "")))
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        for col in ws_summary.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_summary.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        st.download_button(
            "Descargar Excel",
            data=out.getvalue(),
            file_name=f"inventario_uala_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )
    html_report = generate_html_report(filtered)
    st.download_button(
        "📄 Exportar HTML imprimible",
        data=html_report.encode("utf-8"),
        file_name=f"reporte_inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html",
        mime="text/html",
    )
    st.dataframe(table, use_container_width=True, hide_index=True)


def render_extra_page(config: AppConfig, assets: list[dict[str, Any]]) -> None:
    st.subheader("Extra analítico")
    st.caption("Este bloque es adicional y no reemplaza el inventario Jira.")

    with st.expander("🔍 Diagnóstico de carga", expanded=False):
        raw_assets = st.session_state.get("assets", [])
        by_type = defaultdict(lambda: {"count": 0, "types": set()})
        unknown_status = set()
        missing_type = 0
        missing_serial = 0
        missing_hostname = 0
        for a in raw_assets:
            type_id = str(a.get("object_type_id") or "")
            type_name = str(a.get("object_type") or "")
            if not type_id:
                missing_type += 1
            by_type[type_id]["count"] += 1
            if type_name:
                by_type[type_id]["types"].add(type_name)
            if not str(get_serial_value(a)).strip():
                missing_serial += 1
            if not str(get_hostname_value(a)).strip():
                missing_hostname += 1
            status_raw = normalize_text(a.get("status"))
            allowed = {normalize_text(v) for v in ESTADO_NORMALIZATION.values()}
            if status_raw and status_raw not in allowed:
                unknown_status.add(str(a.get("status")))
        diag_rows = []
        for k, v in by_type.items():
            diag_rows.append({"type_id": k or "(sin id)", "cantidad": v["count"], "tipos": ", ".join(sorted(v["types"]))})
        st.dataframe(diag_rows, use_container_width=True, hide_index=True)
        total_assets = max(len(raw_assets), 1)
        st.write(f"Sin object_type_id: **{missing_type}**")
        st.write(f"Sin serial: **{missing_serial}** ({round((missing_serial / total_assets) * 100, 2)}%)")
        st.write(f"Sin hostname: **{missing_hostname}** ({round((missing_hostname / total_assets) * 100, 2)}%)")
        st.write(f"Status desconocido: {', '.join(sorted(unknown_status)) if unknown_status else 'Ninguno'}")
        discovered = st.session_state.get("discovered_type_ids", []) or []
        all_schema_types = st.session_state.get("all_schema_type_ids", []) or []
        active_scope = get_active_hardware_type_ids()
        source = st.session_state.get("type_discovery_source", "fallback")
        discovery_error = st.session_state.get("type_discovery_error", "")
        st.write(
            f"Tipos esquema total: **{len(all_schema_types)}** | descubiertos bajo {GENERAL_HARDWARE_TYPE_ID}: **{len(discovered)}** | "
            f"hardcodeados: **{len(KNOWN_OBJECT_TYPE_IDS)}** | scope activo (incluye {GENERAL_HARDWARE_TYPE_ID}): **{len(active_scope)}** | fuente: **{source}**"
        )
        st.write(
            f"Registros etapa base AQL: **{int(st.session_state.get('last_base_records_count', 0))}** | "
            f"merge base+typeId: **{int(st.session_state.get('last_segmented_records_count', 0))}** | "
            f"sumados por brute force: **{int(st.session_state.get('last_bruteforce_records_count', 0))}**"
        )
        st.write(
            f"Type scan rango {TYPE_SCAN_START}-{TYPE_SCAN_END}: "
            f"checados **{int(st.session_state.get('last_type_scan_checked', 0))}** | "
            f"hits **{int(st.session_state.get('last_type_scan_hits', 0))}**"
        )
        if discovered:
            st.caption(f"Descendientes de {GENERAL_HARDWARE_TYPE_ID}: {', '.join(discovered)}")
        if discovery_error:
            st.caption(f"Detalle discovery fallback: {discovery_error}")
        st.caption("Cada objeto cargado incluye `attrs_by_name` y `attrs_by_id` con todos los atributos devueltos por Jira.")
        st.write(f"Último AQL ejecutado: `{st.session_state.get('last_aql_executed', '')}`")
        st.write(f"Tiempo última carga: **{st.session_state.get('last_load_seconds', 0.0)}s**")

    with st.expander("Puente Jira: Object Types y atributos", expanded=False):
        if st.button("Cargar mapeo de atributos Jira", use_container_width=True):
            st.session_state.schema_bridge = fetch_schema_bridge(config)
        bridge = st.session_state.get("schema_bridge", {})
        if bridge:
            rows = []
            for type_id, attrs in bridge.items():
                for attr in attrs:
                    rows.append(
                        {
                            "object_type_id": type_id,
                            "attribute_id": str(attr.get("id", "")),
                            "attribute_name": str(attr.get("name", "")),
                            "type": str((attr.get("defaultType") or {}).get("name") or ""),
                            "required": int(attr.get("minimumCardinality", 0) or 0) > 0,
                        }
                    )
            st.dataframe(rows, use_container_width=True, hide_index=True, height=300)
        else:
            st.info("Todavía no cargaste el mapeo de atributos.")

    risks = []
    for a in assets:
        score = 0
        reasons = []
        if not str(a.get("serial_number", "")).strip():
            score += 40
            reasons.append("sin serial")
        if not str(a.get("assigned_to", "")).strip():
            score += 20
            reasons.append("sin asignación")
        if normalize_text(a.get("status")) in {"stock usado", "usado"}:
            score += 10
            reasons.append("stock usado")
        warranty = parse_date(str(a.get("warranty_date", "")).split("|")[0].strip())
        if warranty and warranty.date() < datetime.now().date():
            score += 30
            reasons.append("garantía vencida")
        if score > 0:
            risks.append({"Activo": a.get("name", ""), "Jira": a.get("jira_key", ""), "Riesgo": score, "Motivos": ", ".join(reasons)})

    risks = sorted(risks, key=lambda x: x["Riesgo"], reverse=True)
    st.dataframe(risks[:50], use_container_width=True, hide_index=True)

    st.subheader("Ciclo de vida / garantías")
    soon_rows = []
    today = datetime.now().date()
    for a in assets:
        warranty = parse_date(str(a.get("warranty_date", "")).split("|")[0].strip())
        if not warranty:
            continue
        days = (warranty.date() - today).days
        if 0 <= days <= 45:
            soon_rows.append(
                {
                    "Activo": a.get("name", ""),
                    "Jira": a.get("jira_key", ""),
                    "Garantía": a.get("warranty_date", ""),
                    "Días restantes": days,
                }
            )
    st.dataframe(soon_rows, use_container_width=True, hide_index=True)
    st.subheader("📅 Timeline de activo")
    timeline_id = st.text_input("Serial/Hostname para timeline", value="")
    if timeline_id:
        asset = find_asset_by_identifier(assets, timeline_id)
        if not asset:
            st.info("No encontré ese activo para timeline.")
        else:
            purchase = str(asset.get("purchase_date") or "Sin fecha compra")
            created = str(asset.get("created") or "Sin fecha alta")
            status = str(asset.get("status") or "Sin estado")
            assigned = str(asset.get("assigned_to") or "Sin asignar")
            warranty = parse_date(str(asset.get("warranty_date", "")).split("|")[0].strip())
            if warranty:
                delta = (warranty.date() - datetime.now().date()).days
                if delta < 0:
                    warr_color = "#991b1b"
                    warr_label = "Garantía vencida"
                elif delta <= 45:
                    warr_color = "#a16207"
                    warr_label = f"Garantía vence en {delta} días"
                else:
                    warr_color = "#0f766e"
                    warr_label = "Garantía vigente"
            else:
                warr_color = "#6b7280"
                warr_label = "Sin dato de garantía"
            html = f"""
            <div style='border-left:4px solid #94a3b8;padding-left:16px'>
              <div style='margin:8px 0'><span style='color:#475569;font-weight:700'>● Compra</span> — {escape_html_text(purchase)}</div>
              <div style='margin:8px 0'><span style='color:#0f766e;font-weight:700'>● Alta Jira</span> — {escape_html_text(created)}</div>
              <div style='margin:8px 0'><span style='color:#334155;font-weight:700'>● Estado/Asignación</span> — {escape_html_text(status)} / {escape_html_text(assigned)}</div>
              <div style='margin:8px 0'><span style='color:{warr_color};font-weight:700'>● Garantía</span> — {escape_html_text(warr_label)}</div>
            </div>
            """
            st.markdown(html, unsafe_allow_html=True)


def score_asset(asset: dict[str, Any]) -> int:
    """Calcula score de completitud para auditoría."""
    score = 0
    if str(get_serial_value(asset)).strip():
        score += 20
    if str(get_hostname_value(asset)).strip():
        score += 15
    if str(asset.get("model") or "").strip():
        score += 15
    if normalize_text(asset.get("status")) in {normalize_text(v) for v in ESTADO_NORMALIZATION.values()}:
        score += 10
    if str(asset.get("assigned_to") or "").strip() or normalize_text(asset.get("status")) in {normalize_text("stock nuevo"), normalize_text("stock usado")}:
        score += 10
    if str(asset.get("purchase_date") or "").strip():
        score += 10
    if parse_cost(str(asset.get("purchase_price", ""))) > 0:
        score += 10
    if str(asset.get("warranty_date") or "").strip():
        score += 10
    return score


def render_auditoria_page(config: AppConfig, assets: list[dict[str, Any]]) -> None:
    """Renderiza la página de auditoría operativa y de datos."""
    st.subheader("Auditoría")
    rows = []
    for a in assets:
        score = score_asset(a)
        action = "🟢 OK" if score > 70 else "🟡 Completar datos" if score >= 40 else "🔴 Revisar urgente"
        rows.append(
            {
                "Jira": a.get("jira_key"),
                "Nombre": a.get("name"),
                "Score": score,
                "Acción sugerida": action,
                "Serial": get_serial_value(a),
                "Hostname": get_hostname_value(a),
                "Estado": a.get("status"),
            }
        )
    st.dataframe(rows, use_container_width=True, hide_index=True)
    compare: dict[tuple[str, str], dict[str, Any]] = {}
    for a in assets:
        key = (str(a.get("country") or "Sin país"), str(a.get("company") or "Sin compañía"))
        compare.setdefault(key, {"total": 0, "en_uso": 0, "score_sum": 0, "vencidas": 0, "sin_serial": 0, "costo": 0.0})
        compare[key]["total"] += 1
        compare[key]["score_sum"] += score_asset(a)
        compare[key]["costo"] += parse_cost(str(a.get("purchase_price", "")))
        if normalize_text(a.get("status")) == normalize_text("en uso"):
            compare[key]["en_uso"] += 1
        if not get_serial_value(a):
            compare[key]["sin_serial"] += 1
        w = parse_date(str(a.get("warranty_date", "")).split("|")[0].strip())
        if w and w.date() < datetime.now().date():
            compare[key]["vencidas"] += 1
    compare_rows = []
    for (country, company), vals in compare.items():
        compare_rows.append(
            {
                "País": country,
                "Compañía": company,
                "Total": vals["total"],
                "% En uso": round((vals["en_uso"] / max(vals["total"], 1)) * 100, 2),
                "Score promedio": round(vals["score_sum"] / max(vals["total"], 1), 2),
                "Garantías vencidas": vals["vencidas"],
                "Sin serial": vals["sin_serial"],
                "Costo total": round(vals["costo"], 2),
            }
        )
    st.markdown("**Comparativa países/compañías**")
    st.dataframe(compare_rows, use_container_width=True, hide_index=True)
    st.markdown("**Activos huérfanos (En uso sin asignado)**")
    orphan_rows = [a for a in assets if normalize_text(a.get("status")) == normalize_text("en uso") and not str(a.get("assigned_to") or "").strip()]
    for orphan in orphan_rows:
        c1, c2 = st.columns([4, 1])
        c1.write(f"{orphan.get('jira_key')} | {orphan.get('name')} | {orphan.get('status')}")
        if c2.button("Corregir", key=f"fix_orphan_{orphan.get('object_id')}"):
            ok, msg = update_status(config, assets, str(orphan.get("jira_key") or orphan.get("hostname") or orphan.get("serial_number")), "Stock usado")
            if ok:
                st.success(f"✅ {msg}")
            else:
                st.error(f"❌ {msg}")
    st.markdown("**Historial de acciones de sesión**")
    st.dataframe(st.session_state.get("action_log", [])[-200:], use_container_width=True, hide_index=True)


def render_movimientos_page(assets: list[dict[str, Any]]) -> None:
    """Renderiza la página de historial de movimientos con filtros y exportación."""
    st.subheader("Movimientos")
    data = st.session_state.get("movimientos", [])
    if not data:
        st.info("No hay movimientos registrados todavía.")
        return
    if pd is None:
        st.write(data[-100:])
        return
    df = pd.DataFrame(data)
    df["timestamp_dt"] = pd.to_datetime(df["timestamp"], errors="coerce")
    min_date = df["timestamp_dt"].dt.date.min()
    max_date = df["timestamp_dt"].dt.date.max()
    c1, c2 = st.columns(2)
    date_from = c1.date_input("Desde", value=min_date)
    date_to = c2.date_input("Hasta", value=max_date)
    action_types = ["Todos"] + sorted(df["tipo_accion"].dropna().unique().tolist())
    action_filter = st.selectbox("Tipo de acción", action_types)
    free = st.text_input("Búsqueda libre", "")
    filtered = df[(df["timestamp_dt"].dt.date >= date_from) & (df["timestamp_dt"].dt.date <= date_to)]
    if action_filter != "Todos":
        filtered = filtered[filtered["tipo_accion"] == action_filter]
    if free.strip():
        q = normalize_text(free)
        filtered = filtered[
            filtered.apply(
                lambda row: q in normalize_text(" ".join([str(row.get("identificador", "")), str(row.get("jira_key", "")), str(row.get("usuario_asignado", ""))])),
                axis=1,
            )
        ]
    mt1, mt2, mt3, mt4, mt5 = st.columns(5)
    mt1.metric("Total movimientos", len(filtered))
    mt2.metric("Asignaciones", int((filtered["tipo_accion"] == "ASIGNACION").sum()))
    mt3.metric("Desasignaciones", int((filtered["tipo_accion"] == "DESASIGNACION").sum()))
    mt4.metric("Cambios de estado", int((filtered["tipo_accion"] == "CAMBIO_ESTADO").sum()))
    mt5.metric("Errores", int((filtered["resultado"] == "ERROR").sum()))
    show_cols = [
        "timestamp",
        "tipo_accion",
        "identificador",
        "jira_key",
        "campo_modificado",
        "valor_anterior",
        "valor_nuevo",
        "usuario_asignado",
        "ejecutado_por",
        "resultado",
    ]
    st.dataframe(filtered[show_cols], use_container_width=True, hide_index=True)
    if Workbook is not None:
        out = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"
        for col_idx, col in enumerate(show_cols, start=1):
            ws.cell(row=1, column=col_idx, value=col)
        for row_idx, row in enumerate(filtered[show_cols].fillna("").values.tolist(), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=str(value))
        wb.save(out)
        out.seek(0)
        st.download_button(
            "📥 Exportar movimientos",
            data=out.getvalue(),
            file_name=f"movimientos_uala_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def build_asset_attributes_payload(row: dict[str, Any]) -> tuple[str, list[dict[str, Any]]]:
    """Construye payload de atributos para alta/modificación desde fila tabular."""
    row_lookup = build_row_lookup(row)
    mapping = {
        "name": ID_NAME,
        "hostname": ID_HOSTNAME,
        "model": ID_MODELO,
        "purchase_date": ID_FECHA_COMPRA,
        "status": ID_ESTADO,
        "entity": ID_ENTIDAD,
        "warranty_date": ID_FECHA_GARANTIA,
        "cost": ID_COSTO,
        "serial": ID_SERIAL,
        "country": ID_PAIS,
        "assignment": ID_ASIGNACION,
        "provider": ID_PROVEEDOR,
        "category": ID_CATEGORIA,
        "company": ID_COMPANIA,
    }
    resolved_values = {
        field_name: get_row_value_by_aliases(row_lookup, MASS_UPLOAD_COLUMN_ALIASES[field_name])
        for field_name in mapping
    }
    if resolved_values["country"] and not resolved_values["company"]:
        resolved_values["company"] = company_for_country(resolved_values["country"])
    attrs: list[dict[str, Any]] = []
    for field_name, attr_id in mapping.items():
        value = resolved_values[field_name]
        if value:
            attrs.append({"objectTypeAttributeId": str(attr_id), "objectAttributeValues": [{"value": value}]})
    category_value = resolved_values["category"]
    type_id = CATEGORY_TO_TYPE_ID.get(canonical_category(category_value), KNOWN_OBJECT_TYPE_IDS[0])
    return type_id, attrs


def build_mass_upload_template_bytes() -> bytes:
    """Genera una plantilla Excel para altas masivas de assets."""
    if Workbook is None or Font is None or PatternFill is None or Alignment is None or get_column_letter is None:
        raise RuntimeError("openpyxl no está disponible para generar la plantilla.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Carga masiva"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASS_UPLOAD_TEMPLATE_HEADERS))}2"

    header_fill = PatternFill(start_color="003262", end_color="003262", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    required_fill = PatternFill(start_color="D4A12A", end_color="D4A12A", fill_type="solid")
    example_fill = PatternFill(start_color="F6F8FC", end_color="F6F8FC", fill_type="solid")

    for col_idx, header in enumerate(MASS_UPLOAD_TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = required_fill if header in MASS_UPLOAD_REQUIRED_HEADER_SET else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

        example_value = MASS_UPLOAD_TEMPLATE_EXAMPLE_ROW.get(header, "")
        example_cell = ws.cell(row=2, column=col_idx, value=example_value)
        example_cell.fill = example_fill
        example_cell.alignment = Alignment(vertical="center")

    for col_idx, header in enumerate(MASS_UPLOAD_TEMPLATE_HEADERS, start=1):
        values = [header]
        if header in MASS_UPLOAD_REQUIRED_HEADER_SET:
            values[0] = f"{header} *"
        max_len = max(len(str(item or "")) for item in values + [MASS_UPLOAD_TEMPLATE_EXAMPLE_ROW.get(header, "")])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, 18), 34)

    ws_lists = wb.create_sheet("Listas")
    ws_lists.sheet_state = "hidden"
    for col_idx, (header, values) in enumerate(MASS_UPLOAD_TEMPLATE_LISTS.items(), start=1):
        ws_lists.cell(row=1, column=col_idx, value=header)
        for row_idx, value in enumerate(values, start=2):
            ws_lists.cell(row=row_idx, column=col_idx, value=value)
        if DataValidation is not None:
            col_letter = get_column_letter(col_idx)
            dv = DataValidation(
                type="list",
                formula1=f"=Listas!${col_letter}$2:${col_letter}${len(values) + 1}",
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            target_col = MASS_UPLOAD_TEMPLATE_HEADERS.index(header) + 1
            dv.add(f"{get_column_letter(target_col)}2:{get_column_letter(target_col)}1000")

    ws_help = wb.create_sheet("Instrucciones")
    instructions = [
        ("Objetivo", "Completar una fila por asset y luego subir el archivo desde Scripts > Carga masiva."),
        ("Campos obligatorios", ", ".join(MASS_UPLOAD_REQUIRED_HEADERS)),
        ("Fechas", "Usar formato YYYY-MM-DD."),
        ("Usuario asignado", "Si el activo está asignado, cargar el mail corporativo del usuario."),
        ("Compañía", "Si se deja vacía y el país está informado, la app la deriva automáticamente."),
        ("Referencia", "La fila 2 contiene un ejemplo listo para copiar o reemplazar."),
    ]
    ws_help.column_dimensions["A"].width = 22
    ws_help.column_dimensions["B"].width = 90
    for row_idx, (title, detail) in enumerate(instructions, start=1):
        title_cell = ws_help.cell(row=row_idx, column=1, value=title)
        title_cell.font = Font(bold=True)
        detail_cell = ws_help.cell(row=row_idx, column=2, value=detail)
        detail_cell.alignment = Alignment(wrap_text=True, vertical="top")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def resolve_mass_update_identifier(row: dict[str, Any]) -> str:
    row_lookup = build_row_lookup(row)
    return get_row_value_by_aliases(row_lookup, MASS_UPDATE_IDENTIFIER_ALIASES)


def render_scripts_page(config: AppConfig, assets: list[dict[str, Any]]) -> None:
    """Renderiza la página de scripts: carga, modificación y reglas de normalización."""
    st.subheader("Scripts")
    tab_config, tab1, tab2, tab3, tab4 = st.tabs(["⚙️ Configuración", "📥 Carga masiva", "✏️ Modificación masiva", "⚙️ Reglas de normalización", "🤖 Asignación automática"])
    with tab_config:
        st.markdown("**Opciones generales**")
        c1, c2 = st.columns(2)
        st.session_state.critical_threshold = c1.slider("Umbral periféricos críticos", 1, 50, int(st.session_state.get("critical_threshold", 10)))
        st.session_state.cache_ttl_minutes = c2.slider("TTL caché (minutos)", 1, 60, int(st.session_state.get("cache_ttl_minutes", 10)))
        st.session_state.use_ai_compact = st.checkbox("Usar IA compacta", value=bool(st.session_state.get("use_ai_compact", True)))
        st.session_state.auto_clear_after_action = st.checkbox("Limpiar chat tras acciones", value=bool(st.session_state.get("auto_clear_after_action", True)))
        st.markdown("**Log de errores HTTP**")
        error_log = st.session_state.get("error_log", [])
        if error_log:
            st.dataframe(error_log[:20], use_container_width=True, hide_index=True)
        else:
            st.success("Sin errores registrados")
        st.markdown("**Últimos movimientos**")
        moves = st.session_state.get("movimientos", [])
        if moves:
            st.dataframe(moves[-10:][::-1], use_container_width=True, hide_index=True)
        else:
            st.info("Sin movimientos registrados")
    with tab1:
        if Workbook is not None:
            st.download_button(
                "📥 Descargar plantilla de alta",
                data=build_mass_upload_template_bytes(),
                file_name="plantilla_alta_assets.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_mass_upload_template",
            )
            st.caption("La plantilla incluye columnas estándar, fila de ejemplo y listas para tipo, estado, país y compañía.")
        uploaded = st.file_uploader("Subir Excel para carga masiva", type=["xlsx", "xls"], key="mass_upload")
        if uploaded is not None and pd is not None:
            frame = pd.read_excel(uploaded)
            st.dataframe(frame.head(10), use_container_width=True, hide_index=True)
            if st.button("🚀 Ejecutar carga masiva", key="run_mass_upload"):
                progress = st.progress(0)
                total_rows = len(frame)
                results = []
                for idx, row in frame.fillna("").iterrows():
                    row_dict = row.to_dict()
                    type_id, attrs, issues = build_asset_create_payload(config, row_dict)
                    if issues:
                        ok, msg = False, " | ".join(issues[:3])
                    elif attrs:
                        ok, msg = create_asset_from_payload(config, type_id, attrs)
                    else:
                        ok, msg = False, "Sin atributos válidos para crear"
                    if ok:
                        row_lookup = build_row_lookup(row_dict)
                        log_movimiento(
                            config,
                            None,
                            "CARGA_MASIVA",
                            "asset",
                            "",
                            get_row_value_by_aliases(row_lookup, MASS_UPLOAD_COLUMN_ALIASES["name"]),
                            "OK",
                            msg,
                            get_row_value_by_aliases(row_lookup, MASS_UPLOAD_COLUMN_ALIASES["serial"]),
                        )
                    results.append({"fila": idx + 1, "ok": ok, "detalle": msg, "type_id": type_id})
                    progress.progress(min((idx + 1) / max(total_rows, 1), 1.0))
                st.dataframe(results, use_container_width=True, hide_index=True)
                if st.button("Refrescar inventario", key="refresh_after_load"):
                    refresh_assets(config, st.session_state.aql_input, force_live=True)
    with tab2:
        uploaded_mod = st.file_uploader("Subir Excel para modificación masiva", type=["xlsx", "xls"], key="mass_update")
        if uploaded_mod is not None and pd is not None:
            frame_mod = pd.read_excel(uploaded_mod).fillna("")
            st.dataframe(frame_mod.head(10), use_container_width=True, hide_index=True)
            simulate = st.checkbox("🔍 Modo simulación", value=True, key="simulate_mass_update")
            if st.button("✏️ Ejecutar modificación masiva", key="run_mass_update"):
                progress = st.progress(0)
                results = []
                total_rows = len(frame_mod)
                for idx, row in frame_mod.iterrows():
                    row_dict = row.to_dict()
                    identifier = resolve_mass_update_identifier(row_dict)
                    asset = find_asset_by_identifier(assets, identifier)
                    if not asset:
                        results.append({"fila": idx + 1, "identificador": identifier, "ok": False, "detalle": "No encontrado"})
                        continue
                    type_id, attrs = build_asset_attributes_payload(row_dict)
                    if simulate:
                        results.append({"fila": idx + 1, "identificador": identifier, "ok": True, "detalle": f"Simulación ({len(attrs)} attrs)"})
                    elif not attrs:
                        results.append({"fila": idx + 1, "identificador": identifier, "ok": True, "detalle": "Sin cambios para aplicar"})
                    else:
                        ok, msg = update_asset_attributes(config, str(asset.get("object_id", "")), str(asset.get("object_type_id") or type_id), attrs)
                        if ok:
                            log_movimiento(config, asset, "MODIFICACION_MASIVA", "multiple", "", "updated", "OK", msg, identifier)
                        results.append({"fila": idx + 1, "identificador": identifier, "ok": ok, "detalle": msg})
                    progress.progress(min((idx + 1) / max(total_rows, 1), 1.0))
                st.dataframe(results, use_container_width=True, hide_index=True)
    with tab3:
        campos_base = ["hostname", "serial_number", "model", "status", "country", "company", "provider", "entity"]
        campo_cond = st.selectbox("Campo condición", campos_base + ["Atributo personalizado..."], key="rule_cond_field")
        if campo_cond == "Atributo personalizado...":
            campo_cond = st.text_input("Nombre atributo condición", key="rule_cond_custom")
        operador = st.selectbox("Operador", ["empieza_con", "contiene", "termina_con", "igual_a", "regex", "no_contiene"], key="rule_operator")
        valor_cond = st.text_input("Valor condición", key="rule_cond_value")
        campo_mod = st.selectbox("Campo a modificar", ["country", "company", "status", "model", "provider", "Atributo personalizado..."], key="rule_mod_field")
        if campo_mod == "Atributo personalizado...":
            campo_mod = st.text_input("Nombre atributo a modificar", key="rule_mod_custom")
        valor_nuevo = st.text_input("Valor nuevo", key="rule_new_value")
        descripcion = st.text_input("Descripción", key="rule_description")
        regla_actual = ReglaNormalizacion(
            campo_condicion=campo_cond,
            operador=operador,
            valor_condicion=valor_cond,
            campo_a_modificar=campo_mod,
            valor_nuevo=valor_nuevo,
            descripcion=descripcion or f"{campo_cond} {operador} {valor_cond} => {campo_mod}={valor_nuevo}",
        )
        if st.button("🔍 Vista previa", key="preview_rule"):
            affected, errors = aplicar_regla(config, assets, regla_actual, dry_run=True)
            st.info(f"Afectados: {affected}")
            if errors:
                st.write(errors)
        if st.button("💾 Guardar regla", key="save_rule"):
            reglas = st.session_state.setdefault("reglas_guardadas", [])
            reglas.append(regla_actual.__dict__)
            st.session_state["reglas_guardadas"] = reglas
            if save_normalization_rules(reglas):
                st.success("Regla guardada.")
            else:
                st.error("No se pudo persistir la regla en disco.")
        if st.button("⚡ Aplicar regla", key="apply_rule"):
            updated, errors = aplicar_regla(config, assets, regla_actual, dry_run=False)
            st.success(f"Afectados: {updated}")
            if errors:
                st.error(" | ".join(errors[:10]))
        st.markdown("**Reglas guardadas**")
        for idx, rule in enumerate(st.session_state.get("reglas_guardadas", [])):
            c1, c2, c3 = st.columns([6, 1, 1])
            c1.write(f"{idx+1}. {rule.get('descripcion', '')}")
            if c2.button("Aplicar", key=f"apply_saved_{idx}"):
                updated, errors = aplicar_regla(config, assets, ReglaNormalizacion(**rule), dry_run=False)
                st.success(f"Regla aplicada. Afectados: {updated}")
                if errors:
                    st.error(" | ".join(errors[:5]))
            if c3.button("Eliminar", key=f"del_saved_{idx}"):
                reglas = st.session_state.get("reglas_guardadas", [])
                st.session_state["reglas_guardadas"] = [r for i, r in enumerate(reglas) if i != idx]
                if not save_normalization_rules(st.session_state["reglas_guardadas"]):
                    st.error("No se pudo persistir la eliminación de la regla.")
    with tab4:
        scheduler_enabled = st.toggle(
            "Activar scheduler automático",
            value=bool(st.session_state.get("scheduler_running", False)),
            key="auto_assign_scheduler_toggle",
        )
        if scheduler_enabled and not st.session_state.get("scheduler_running"):
            if BackgroundScheduler is None:
                st.error("APScheduler no está instalado. Instalá `apscheduler`.")
            else:
                start_auto_assign_scheduler(config)
                st.success("Scheduler iniciado.")
        if not scheduler_enabled and st.session_state.get("scheduler_running"):
            stop_auto_assign_scheduler()
            st.info("Scheduler detenido.")

        st.markdown("### Nueva regla")
        col_a, col_b = st.columns(2)
        with col_a:
            nombre = st.text_input("Nombre regla", key="auto_rule_nombre")
            campo_cond = st.selectbox("Campo condición", ["hostname", "country", "category", "company"], key="auto_rule_campo")
            operador = st.selectbox("Operador", ["empieza_con", "contiene", "termina_con", "igual_a", "regex"], key="auto_rule_operador")
            valor_cond = st.text_input("Valor condición", key="auto_rule_valor")
            prioridad = st.number_input("Prioridad", min_value=1, max_value=999, value=100, step=1, key="auto_rule_prio")
        with col_b:
            tipo_accion = st.selectbox("Tipo acción", ["asignar_usuario", "asignar_por_stock", "cambiar_estado"], key="auto_rule_accion")
            usuario_destino = st.text_input("Usuario destino (email)", key="auto_rule_usuario")
            cola_usuarios = st.text_area("Cola usuarios (uno por línea)", key="auto_rule_queue")
            estado_destino = st.text_input("Estado destino", value="En uso", key="auto_rule_estado")
            pais_destino = st.text_input("País filtro opcional", key="auto_rule_pais")
        descripcion = st.text_input("Descripción", key="auto_rule_desc")

        regla_tmp = ReglaAsignacionAuto(
            nombre=nombre.strip() or f"Regla {datetime.now().strftime('%H:%M:%S')}",
            activa=True,
            campo_condicion=campo_cond,
            operador=operador,
            valor_condicion=valor_cond.strip(),
            tipo_accion=tipo_accion,
            usuario_destino=usuario_destino.strip(),
            perfil_destino="",
            pais_destino=pais_destino.strip(),
            cola_usuarios=[line.strip() for line in cola_usuarios.splitlines() if line.strip()],
            estado_destino=estado_destino.strip(),
            prioridad=int(prioridad),
            descripcion=descripcion.strip() or f"{campo_cond} {operador} '{valor_cond}' -> {tipo_accion}",
        )

        ctest, csave, crun = st.columns(3)
        if ctest.button("🧪 Probar regla", key="auto_rule_test"):
            matches = [a for a in assets if evaluar_regla_asignacion(a, regla_tmp)]
            st.info(f"Coincidencias: {len(matches)}")
            if matches:
                st.write([_asset_identifier(a) for a in matches[:10]])
        if csave.button("💾 Guardar regla automática", key="auto_rule_save"):
            rules = st.session_state.get("auto_assign_rules", [])
            rules.append(regla_tmp.__dict__)
            st.session_state["auto_assign_rules"] = rules
            if save_auto_assign_rules(rules):
                st.success("Regla automática guardada.")
            else:
                st.error("No se pudo persistir la regla automática.")
        if crun.button("▶ Ejecutar job ahora", key="auto_rule_run_now"):
            resultados = auto_assign_job(config)
            st.session_state["auto_assign_log"] = load_auto_assign_log()
            st.info(f"Job ejecutado. Resultados: {len(resultados)}")

        st.markdown("### Reglas guardadas")
        rules = st.session_state.get("auto_assign_rules", [])
        for idx, row in enumerate(rules):
            r = ReglaAsignacionAuto(**row)
            c1, c2, c3 = st.columns([6, 1, 1])
            c1.write(f"{idx+1}. [{'ON' if r.activa else 'OFF'}] {r.nombre} | p={r.prioridad} | {r.descripcion}")
            if c2.button("Toggle", key=f"auto_rule_toggle_{idx}"):
                rules[idx]["activa"] = not bool(rules[idx].get("activa", True))
                st.session_state["auto_assign_rules"] = rules
                if not save_auto_assign_rules(rules):
                    st.error("No se pudo persistir el cambio de estado de la regla.")
            if c3.button("Eliminar", key=f"auto_rule_del_{idx}"):
                st.session_state["auto_assign_rules"] = [x for i, x in enumerate(rules) if i != idx]
                if not save_auto_assign_rules(st.session_state["auto_assign_rules"]):
                    st.error("No se pudo persistir la eliminación de la regla.")
                st.rerun()

        st.markdown("### Log (últimos 100)")
        auto_log = load_auto_assign_log()
        st.session_state["auto_assign_log"] = auto_log
        if auto_log:
            st.dataframe(auto_log[::-1], use_container_width=True, hide_index=True)
        else:
            st.caption("Sin ejecuciones todavía.")


def render_chat_page(config: AppConfig, assets: list[dict[str, Any]]) -> None:
    all_assets = st.session_state.get("assets", assets)
    total = len(assets)
    in_use = sum(1 for a in assets if normalize_text(a.get("status", "")) == "en uso")
    stock = sum(1 for a in assets if "stock" in normalize_text(a.get("status", "")))
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total", total)
    m2.metric("En uso", f"{in_use} ({round(in_use / max(total, 1) * 100, 1)}%)")
    m3.metric("Stock", stock)
    m4.metric("Sin asignar", sum(1 for a in assets if not str(a.get("assigned_to") or "").strip()))
    _render_pending_action_block(config, all_assets)

    suggestions = _build_contextual_suggestions(assets, st.session_state.get("anomaly_report", {}))
    cols = st.columns(min(max(len(suggestions), 1), 4))
    for i, sug in enumerate(suggestions[:4]):
        if cols[i].button(sug, use_container_width=True, key=f"sug_{i}"):
            _process_chat_prompt(config, all_assets, assets, sug)
            st.rerun()

    for msg in st.session_state.get("chat_history", []):
        with st.chat_message(msg["role"]):
            plain, charts = decode_chat_payload(msg.get("content", ""))
            intro, table_df = parse_chat_response_for_table(plain)
            if intro:
                st.markdown(intro)
            if charts:
                st.caption("Dashboard visual actualizado en la sección de abajo.")
            if table_df is not None:
                st.dataframe(table_df, use_container_width=True, hide_index=True)
    render_chat_dashboard_panel()

    prompt = st.chat_input("Preguntá por los activos de Uala...")
    if prompt:
        _process_chat_prompt(config, all_assets, assets, prompt)
        st.rerun()


# ── 11. ENTRYPOINT ────────────────────────────────────────────────────────
def main() -> None:
    st.set_page_config(page_title="Uala Assets", page_icon="U", layout="wide", initial_sidebar_state="collapsed")
    ensure_session_state()
    config = load_config()
    apply_theme()

    missing = []
    if not config.jira_email:
        missing.append("JIRA_EMAIL")
    if not config.jira_api_token:
        missing.append("JIRA_API_TOKEN")
    if not config.workspace_id:
        missing.append("ASSETS_WORKSPACE_ID / JIRA_WORKSPACE_ID")

    if missing:
        render_setup_screen()
        st.error("Faltan variables requeridas: " + ", ".join(missing))
        st.stop()

    params = st.query_params
    page = params.get("page", "Chat")
    valid_pages = ["Chat", "Activos", "Insights", "Auditoría", "Movimientos", "Scripts", "Extra"]
    if page not in valid_pages:
        page = "Chat"
    debug_log(
        f"main:page={page} assets_before={len(st.session_state.get('assets', []))} "
        f"aql={st.session_state.get('aql_input', '')!r}"
    )

    raw_assets = st.session_state.assets
    render_topbar(config, page, raw_assets)
    render_filterbar(config)
    if not raw_assets:
        render_branding(config)

    if not st.session_state.assets:
        try:
            refresh_assets(config, st.session_state.aql_input)
        except Exception as exc:
            st.session_state.last_error = str(exc)

    if st.session_state.last_error:
        st.error(st.session_state.last_error)
        st.stop()

    if not st.session_state.assets and not st.session_state.get("auto_reset_empty_once", False):
        st.session_state["auto_reset_empty_once"] = True
        st.session_state["cache_hash"] = ""
        st.session_state["cache_expiry"] = None
        st.session_state["aql_input"] = ""
        st.session_state["global_filter_countries"] = []
        st.session_state["global_filter_companies"] = []
        try:
            refresh_assets(config, "", force_live=True)
        except Exception as exc:
            st.session_state.last_error = str(exc)
        st.rerun()

    raw_assets = st.session_state.assets
    if raw_assets:
        st.session_state["auto_reset_empty_once"] = False
    assets = apply_global_filter(raw_assets)
    debug_log(f"main:render page={page} raw_assets={len(raw_assets)} visible_assets={len(assets)}")
    if raw_assets and not assets:
        st.warning("No hay activos visibles con los filtros actuales. Limpié país/compañía o usá el botón ✕.")

    if page == "Chat":
        render_chat_page(config, assets)
    elif page == "Activos":
        render_assets_page(assets)
    elif page == "Insights":
        render_insights(assets)
    elif page == "Auditoría":
        render_auditoria_page(config, assets)
    elif page == "Movimientos":
        render_movimientos_page(assets)
    elif page == "Scripts":
        render_scripts_page(config, assets)
    else:
        render_extra_page(config, assets)


if __name__ == "__main__":
    main()
